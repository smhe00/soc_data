from __future__ import annotations

from datetime import datetime, timezone
import os
import json
from pathlib import Path
from tempfile import SpooledTemporaryFile
from tempfile import NamedTemporaryFile
from typing import Any

from sqlalchemy import delete, text
from fastapi import BackgroundTasks, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel
from sqlmodel import Field, Session, SQLModel, create_engine, select


BASE_DIR = Path(__file__).resolve().parent
DATABASE_DIR = BASE_DIR / "databases"
DEFAULT_DATABASE_PATH = BASE_DIR / "soc_3dic.db"
ACTIVE_DATABASE_PATH = Path(os.getenv("SOC_DB_PATH", DEFAULT_DATABASE_PATH)).expanduser().resolve()
TEMPLATE_PATH = BASE_DIR.parent / "templates" / "soc_import_template.xlsx"

engine = create_engine(f"sqlite:///{ACTIVE_DATABASE_PATH}", connect_args={"check_same_thread": False})


def database_id(path: Path) -> str:
    return path.stem


def database_label(path: Path) -> str:
    return "Demo database" if path.resolve() == DEFAULT_DATABASE_PATH.resolve() else path.stem.replace("_", " ")


def database_paths() -> list[Path]:
    DATABASE_DIR.mkdir(parents=True, exist_ok=True)
    paths = [DEFAULT_DATABASE_PATH]
    paths.extend(sorted(DATABASE_DIR.glob("*.db")))
    unique: dict[str, Path] = {}
    for path in paths:
        unique[str(path.resolve())] = path.resolve()
    return list(unique.values())


def database_path_from_id(db_id: str) -> Path:
    cleaned = db_id.strip()
    if not cleaned:
        raise HTTPException(status_code=400, detail="Database id is required.")
    for path in database_paths():
        if database_id(path) == cleaned:
            return path
    raise HTTPException(status_code=404, detail=f"Database not found: {db_id}")


def switch_database(path: Path, create_if_missing: bool = False) -> None:
    global engine, ACTIVE_DATABASE_PATH
    resolved = path.resolve()
    if create_if_missing:
        resolved.parent.mkdir(parents=True, exist_ok=True)
        resolved.touch(exist_ok=True)
    if not resolved.exists():
        raise HTTPException(status_code=404, detail=f"Database file not found: {database_id(resolved)}")
    engine.dispose()
    ACTIVE_DATABASE_PATH = resolved
    engine = create_engine(f"sqlite:///{ACTIVE_DATABASE_PATH}", connect_args={"check_same_thread": False})


def now_iso() -> str:
    return datetime.now(timezone.utc).date().isoformat()


class Project(SQLModel, table=True):
    id: str = Field(primary_key=True)
    name: str
    product_family: str
    generation: str
    owner: str
    phase: str
    description: str = ""
    created_at: str
    updated_at: str


class ImplOption(SQLModel, table=True):
    __tablename__ = "imploption"
    id: str = Field(primary_key=True)
    project_id: str = Field(foreign_key="project.id")
    name: str
    impl_type: str
    process_combo: str
    description: str = ""
    status: str
    created_at: str
    updated_at: str


class ModuleDefinition(SQLModel, table=True):
    id: str = Field(primary_key=True)
    name: str
    module_type: str
    ip_owner: str
    reuse_class: str
    description: str = ""


class LogicalComponent(SQLModel, table=True):
    id: str = Field(primary_key=True)
    project_id: str = Field(foreign_key="project.id")
    parent_id: str | None = Field(default=None, foreign_key="logicalcomponent.id")
    module_definition_id: str | None = Field(default=None, foreign_key="moduledefinition.id")
    name: str
    instance_type: str
    resource_type: str
    function_domain: str
    hierarchy_path: str
    logical_instance_count: int = 1
    owner_team: str = "Architecture Team"
    visibility_level: str = "team"
    description: str = ""
    created_at: str
    updated_at: str


class ProcessNode(SQLModel, table=True):
    id: str = Field(primary_key=True)
    foundry: str
    node_name: str
    logic_density_mtr_per_mm2: float
    sram_density_mb_per_mm2: float
    logic_area_scale: float = 1
    sram_area_scale: float = 1
    block_area_scale: float = 1
    voltage_nominal: float
    cost_factor: float
    maturity_level: str
    description: str = ""


class Tier(SQLModel, table=True):
    id: str = Field(primary_key=True)
    impl_option_id: str = Field(foreign_key="imploption.id")
    tier_index: int
    name: str
    process_id: str = Field(foreign_key="processnode.id")
    role: str
    orientation: str
    thickness_um: float = 0
    area_limit_mm2: float
    description: str = ""


class PhysicalPartition(SQLModel, table=True):
    id: str = Field(primary_key=True)
    impl_option_id: str = Field(foreign_key="imploption.id")
    logical_component_id: str = Field(foreign_key="logicalcomponent.id")
    tier_id: str = Field(foreign_key="tier.id")
    partition_name: str
    partition_type: str
    resource_category: str = "block"
    physical_instance_count: int = 1
    partition_ratio: float = 1
    content_share: float = 1
    description: str = ""


class Metric(SQLModel, table=True):
    id: str = Field(primary_key=True)
    impl_option_id: str = Field(foreign_key="imploption.id")
    subject_type: str
    subject_id: str
    metric_name: str
    metric_value: str
    metric_unit: str = ""
    metric_category: str = ""
    value_type: str = "number"
    corner: str = "typical"
    workload: str = "nominal"
    confidence: str = "draft"
    source_note: str = ""
    created_at: str


class ResponsibilityAssignment(SQLModel, table=True):
    id: str = Field(primary_key=True)
    project_id: str = Field(foreign_key="project.id")
    impl_option_id: str = Field(foreign_key="imploption.id")
    user_id: str
    team_name: str
    logical_component_id: str = Field(foreign_key="logicalcomponent.id")
    scope_type: str = "subtree"
    can_read: bool = True
    can_write: bool = True


class ImplOptionDetail(SQLModel, table=True):
    __tablename__ = "imploptiondetail"
    impl_option_id: str = Field(primary_key=True, foreign_key="imploption.id")
    implementation_type: str
    status: str = "draft"
    version: int = 1
    updated_at: str


class ImplementationTier(SQLModel, table=True):
    id: str = Field(primary_key=True)
    impl_option_id: str = Field(foreign_key="imploption.id")
    tier_id: str
    tier_index: int
    name: str
    process: str
    role: str
    thickness_um: float = 0


class ImplementationInterface(SQLModel, table=True):
    id: str = Field(primary_key=True)
    impl_option_id: str = Field(foreign_key="imploption.id")
    from_tier_id: str
    to_tier_id: str
    orientation: str
    interconnect: str
    hb_pitch_um: float = 0
    upper_tsv_pitch_um: float = 0
    upper_tsv_keepout_um: float = 0
    lower_tsv_pitch_um: float = 0
    lower_tsv_keepout_um: float = 0
    description: str = ""


class ImplementationPackageEscape(SQLModel, table=True):
    impl_option_id: str = Field(primary_key=True, foreign_key="imploption.id")
    bottom_tier_id: str
    requires_tsv: bool = False
    pitch_um: float = 0
    keepout_um: float = 0
    description: str = ""


class ApplicationScenario(SQLModel, table=True):
    __tablename__ = "applicationscenario"
    id: str = Field(primary_key=True)
    project_id: str = Field(foreign_key="project.id")
    name: str
    category: str
    description: str = ""


class PhysicalMapping(SQLModel, table=True):
    __tablename__ = "physicalmapping"
    id: str = Field(primary_key=True)
    impl_option_id: str = Field(foreign_key="imploption.id")
    name: str
    mapping_version: str
    description: str = ""
    mapping_json: str = ""


class OperatingPointSet(SQLModel, table=True):
    __tablename__ = "operatingpointset"
    id: str = Field(primary_key=True)
    project_id: str = Field(foreign_key="project.id")
    name: str
    description: str = ""
    op_json: str = ""


class PowerObservation(SQLModel, table=True):
    __tablename__ = "powerobservation"
    id: str = Field(primary_key=True)
    project_id: str = Field(foreign_key="project.id")
    impl_option_id: str = Field(foreign_key="imploption.id")
    physical_mapping_id: str = Field(foreign_key="physicalmapping.id")
    application_scenario_id: str = Field(foreign_key="applicationscenario.id")
    operating_point_set_id: str = Field(foreign_key="operatingpointset.id")
    
    scope_type: str
    scope_id: str | None = Field(default=None, foreign_key="logicalcomponent.id")
    scope_name: str
    
    use_case_name: str | None = None
    time_window_name: str | None = None
    statistic_type: str = "average"
    power_type: str = "total"
    power_value_w: float = 0.0
    
    development_stage: str | None = None
    source_type: str | None = None
    confidence: str | None = None
    is_additive: bool = True
    
    context_json: str | None = None
    note: str | None = None


class ApplicationScenarioSelection(SQLModel, table=True):
    __tablename__ = "applicationscenarioselection"
    id: str = Field(primary_key=True)
    project_id: str = Field(foreign_key="project.id")
    impl_option_id: str = Field(foreign_key="imploption.id")
    physical_mapping_id: str = Field(foreign_key="physicalmapping.id")
    application_scenario_id: str = Field(foreign_key="applicationscenario.id")
    component_id: str = Field(foreign_key="logicalcomponent.id")
    component_name: str
    use_case_name: str = "Default"
    operating_point_set_id: str = Field(foreign_key="operatingpointset.id")
    included: bool = True
    note: str = ""


class PartitionInput(BaseModel):
    id: str
    tier_id: str
    partition_name: str
    partition_type: str
    resource_category: str = "block"
    physical_instance_count: int
    content_share: float | None = None
    partition_ratio: float | None = None
    description: str = ""


class ComponentDetailUpdate(BaseModel):
    impl_option_id: str = "S2"
    team: str | None = None
    logical_instance_count: int
    partitions: list[PartitionInput]
    signal_count_total: int | None = None
    logic_area: float | None = None
    sram_area: float | None = None
    block_area: float | None = None
    power: float | None = None


class LogicalComponentInput(BaseModel):
    id: str | None = None
    project_id: str
    parent_id: str | None = None
    module_definition_id: str | None = None
    name: str
    instance_type: str = "block"
    resource_type: str = "logic"
    function_domain: str = "General"
    logical_instance_count: int = 1
    owner_team: str = "Architecture Team"
    visibility_level: str = "team"
    description: str = ""
    impl_option_id: str = "S2"
    team: str | None = None


class LogicalComponentDeleteInput(BaseModel):
    impl_option_id: str = "S2"
    team: str | None = None
    cascade: bool = True


class PowerObservationCreate(BaseModel):
    project_id: str
    impl_option_id: str
    physical_mapping_id: str
    application_scenario_id: str = "AS_default"
    operating_point_set_id: str
    
    scope_type: str = "component"
    scope_id: str | None = None
    scope_name: str
    
    use_case_name: str | None = None
    power_value_w: float = 0.0
    
    time_window_name: str | None = "steady_state"
    statistic_type: str = "average"
    power_type: str = "total"
    development_stage: str | None = "architecture_estimate"
    confidence: str | None = "draft"
    is_additive: bool = True
    note: str | None = None


class ModulePowerUseCaseInput(BaseModel):
    project_id: str
    impl_option_id: str
    physical_mapping_id: str
    component_id: str
    component_name: str
    use_case_name: str = "Default"
    operating_point_set_id: str | None = None
    operating_point_set_name: str | None = None
    power_value_w: float
    confidence: str | None = "draft"
    note: str | None = None


class ApplicationScenarioInput(BaseModel):
    project_id: str
    name: str
    category: str = "Custom"
    description: str | None = ""


class ApplicationScenarioSelectionInput(BaseModel):
    component_id: str
    component_name: str
    use_case_name: str = "Default"
    operating_point_set_id: str
    included: bool = True
    note: str | None = None


class ApplicationScenarioCompositionUpdate(BaseModel):
    project_id: str
    impl_option_id: str
    physical_mapping_id: str
    application_scenario_id: str
    selections: list[ApplicationScenarioSelectionInput]


class DatabaseCreateInput(BaseModel):
    name: str
    seed_demo: bool = False


class DatabaseSelectInput(BaseModel):
    id: str


class ImplementationTierInput(BaseModel):
    id: str
    name: str
    process: str
    role: str
    thickness_um: float


class ImplementationInterfaceInput(BaseModel):
    id: str
    from_tier_id: str
    to_tier_id: str
    orientation: str
    interconnect: str
    hb_pitch_um: float = 0
    upper_tsv_pitch_um: float = 0
    upper_tsv_keepout_um: float = 0
    lower_tsv_pitch_um: float = 0
    lower_tsv_keepout_um: float = 0
    description: str = ""


class ImplementationPackageEscapeInput(BaseModel):
    bottom_tier_id: str
    requires_tsv: bool = False
    pitch_um: float = 0
    keepout_um: float = 0
    description: str = ""


class ImplOptionDetailUpdate(BaseModel):
    implementation_type: str
    status: str = "draft"
    tiers: list[ImplementationTierInput]
    interfaces: list[ImplementationInterfaceInput]
    package_escape: ImplementationPackageEscapeInput


app = FastAPI(title="SoC Cross-Die Database API", version="0.2.0")
cors_origins = [
    origin.strip()
    for origin in os.getenv("CORS_ORIGINS", "http://localhost:5173,http://127.0.0.1:5173").split(",")
    if origin.strip()
]
cors_origin_regex = os.getenv("CORS_ORIGIN_REGEX", r"https?://[^/]+:5173")

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_origin_regex=cors_origin_regex,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def create_db_and_tables() -> None:
    SQLModel.metadata.create_all(engine)


def ensure_sqlite_schema_compatibility() -> None:
    with engine.begin() as connection:
        rows = connection.execute(text("PRAGMA table_info(logicalcomponent)")).fetchall()
        columns = {row[1] for row in rows}
        if "owner_team" not in columns:
            connection.execute(text("ALTER TABLE logicalcomponent ADD COLUMN owner_team VARCHAR DEFAULT 'Architecture Team'"))
        if "visibility_level" not in columns:
            connection.execute(text("ALTER TABLE logicalcomponent ADD COLUMN visibility_level VARCHAR DEFAULT 'team'"))
        partition_rows = connection.execute(text("PRAGMA table_info(physicalpartition)")).fetchall()
        partition_columns = {row[1] for row in partition_rows}
        if "content_share" not in partition_columns:
            connection.execute(text("ALTER TABLE physicalpartition ADD COLUMN content_share FLOAT DEFAULT 1"))
            connection.execute(
                text(
                    "UPDATE physicalpartition "
                    "SET content_share = CASE WHEN partition_type = 'full' THEN 1 ELSE partition_ratio END"
                )
            )
        if "resource_category" not in partition_columns:
            connection.execute(text("ALTER TABLE physicalpartition ADD COLUMN resource_category VARCHAR DEFAULT 'block'"))
        process_rows = connection.execute(text("PRAGMA table_info(processnode)")).fetchall()
        process_columns = {row[1] for row in process_rows}
        if "logic_area_scale" not in process_columns:
            connection.execute(text("ALTER TABLE processnode ADD COLUMN logic_area_scale FLOAT DEFAULT 1"))
        if "sram_area_scale" not in process_columns:
            connection.execute(text("ALTER TABLE processnode ADD COLUMN sram_area_scale FLOAT DEFAULT 1"))
        if "block_area_scale" not in process_columns:
            connection.execute(text("ALTER TABLE processnode ADD COLUMN block_area_scale FLOAT DEFAULT 1"))
        connection.execute(text("UPDATE physicalpartition SET partition_type = 'partial' WHERE partition_type = 'residual'"))
        connection.execute(
            text(
                "UPDATE physicalpartition "
                "SET logical_component_id = ("
                "SELECT parent_id FROM logicalcomponent WHERE logicalcomponent.id = physicalpartition.logical_component_id"
                ") "
                "WHERE logical_component_id IN ("
                "SELECT id FROM logicalcomponent WHERE instance_type = 'parent_residual' AND parent_id IS NOT NULL"
                ")"
            )
        )
        connection.execute(
            text(
                "DELETE FROM metric "
                "WHERE subject_type = 'logical_component' "
                "AND subject_id IN (SELECT id FROM logicalcomponent WHERE instance_type = 'parent_residual')"
            )
        )
        connection.execute(text("DELETE FROM logicalcomponent WHERE instance_type = 'parent_residual'"))


def number_or_zero(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def metric_id(row: dict[str, Any]) -> str:
    return (
        f"{row['impl_option_id']}-{row['subject_type']}-{row['subject_id']}-"
        f"{row['metric_name']}-{row['corner']}-{row['workload']}"
    )


def seed_data() -> None:
    create_db_and_tables()
    created = now_iso()
    with Session(engine) as session:
        demo_impl_options = ["S1", "S2", "S3"]
        session.exec(delete(Metric).where(Metric.impl_option_id.in_(demo_impl_options)))
        session.exec(delete(PhysicalPartition).where(PhysicalPartition.impl_option_id.in_(demo_impl_options)))
        session.exec(delete(Tier).where(Tier.impl_option_id.in_(demo_impl_options)))
        session.exec(delete(ResponsibilityAssignment).where(ResponsibilityAssignment.project_id == "P001"))
        session.exec(delete(LogicalComponent).where(LogicalComponent.project_id == "P001"))
        session.exec(delete(ModuleDefinition).where(ModuleDefinition.id.like("MD_%")))
        session.exec(delete(ImplOption).where(ImplOption.id.in_(demo_impl_options)))
        session.exec(delete(Project).where(Project.id.in_(["P001", "P002"])))
        session.exec(delete(ApplicationScenarioSelection))
        session.exec(delete(PowerObservation))
        session.exec(delete(OperatingPointSet))
        session.exec(delete(PhysicalMapping))
        session.exec(delete(ApplicationScenario))

        projects = [
            Project(
                id="P001",
                name="Orion X1 Mobile SoC",
                product_family="Premium Smartphone SoC",
                generation="2027-A",
                owner="Mobile Architecture Team",
                phase="Architecture Planning",
                description="Realistic phase-1 demo for a flagship application processor with CPU/GPU/NPU/ISP/video/display/memory/IO/security domains.",
                created_at=created,
                updated_at=created,
            )
        ]
        implOptions = [
            ImplOption(id="S1", project_id="P001", name="Monolithic N3E Baseline", impl_type="1 die", process_combo="N3E monolithic", description="Single large advanced-node die used as a planning baseline.", status="Medium", created_at=created, updated_at=created),
            ImplOption(id="S2", project_id="P001", name="3DIC Performance Option", impl_type="3 tiers W2W", process_combo="N3E logic + N5 SRAM/cache + N6 IO/analog", description="Compute logic on top tier, SRAM/cache and medium logic on middle tier, IO/PHY/always-on/analog on bottom tier.", status="High", created_at=created, updated_at=created),
            ImplOption(id="S3", project_id="P001", name="Cost-Optimized 2.5D Option", impl_type="2 dies on interposer", process_combo="N4P application die + N6 IO/cache die", description="Lower-risk split with a main application die and a companion cache/IO die.", status="Medium", created_at=created, updated_at=created),
        ]
        process_nodes = [
            ProcessNode(id="PN3E", foundry="TSMC", node_name="N3E", logic_density_mtr_per_mm2=235.0, sram_density_mb_per_mm2=1.85, logic_area_scale=1.00, sram_area_scale=1.00, block_area_scale=1.00, voltage_nominal=0.70, cost_factor=1.9, maturity_level="Ramp", description="Advanced high-performance mobile logic process; base area reference for demo logical metrics."),
            ProcessNode(id="PN4P", foundry="TSMC", node_name="N4P", logic_density_mtr_per_mm2=185.0, sram_density_mb_per_mm2=1.45, logic_area_scale=1.27, sram_area_scale=1.28, block_area_scale=1.18, voltage_nominal=0.74, cost_factor=1.45, maturity_level="Mature", description="Cost-optimized advanced mobile logic process."),
            ProcessNode(id="PN5", foundry="TSMC", node_name="N5", logic_density_mtr_per_mm2=171.3, sram_density_mb_per_mm2=1.35, logic_area_scale=1.37, sram_area_scale=1.37, block_area_scale=1.25, voltage_nominal=0.75, cost_factor=1.25, maturity_level="Production", description="Memory/cache-friendly advanced process."),
            ProcessNode(id="PN6", foundry="TSMC", node_name="N6", logic_density_mtr_per_mm2=118.0, sram_density_mb_per_mm2=1.05, logic_area_scale=1.99, sram_area_scale=1.76, block_area_scale=1.45, voltage_nominal=0.80, cost_factor=0.85, maturity_level="Mature", description="Mature companion die process for IO, PHY, always-on, and analog-friendly logic."),
        ]
        module_definitions = [
            ModuleDefinition(id="MD_CPU_P_CORE", name="ARMV9_BIG_CORE", module_type="cpu_core", ip_owner="CPU Team", reuse_class="replicated", description="High-performance Armv9-class CPU core."),
            ModuleDefinition(id="MD_CPU_E_CORE", name="ARMV9_EFF_CORE", module_type="cpu_core", ip_owner="CPU Team", reuse_class="replicated", description="Efficiency CPU core."),
            ModuleDefinition(id="MD_GPU_SHADER", name="IMMORTALIS_SHADER_SLICE", module_type="gpu_slice", ip_owner="GPU Team", reuse_class="replicated", description="Mobile GPU shader slice."),
            ModuleDefinition(id="MD_NPU_TENSOR", name="NPU_TENSOR_TILE", module_type="ai_accelerator", ip_owner="AI Team", reuse_class="replicated", description="INT8/FP16 tensor tile."),
            ModuleDefinition(id="MD_SRAM_BANK", name="SRAM_BANK_8MB", module_type="memory_macro", ip_owner="Memory Team", reuse_class="compiler_macro", description="Compiled SRAM macro bank."),
            ModuleDefinition(id="MD_ISP_PIPE", name="ISP_PIPELINE", module_type="image_pipeline", ip_owner="Camera Team", reuse_class="replicated", description="Image signal processing pipeline."),
            ModuleDefinition(id="MD_VIDEO_CODEC", name="VIDEO_CODEC_ENGINE", module_type="media_engine", ip_owner="Media Team", reuse_class="configurable", description="Encode/decode media engine."),
            ModuleDefinition(id="MD_DISPLAY_PIPE", name="DISPLAY_PIPE", module_type="display_pipeline", ip_owner="Display Team", reuse_class="replicated", description="Display compositor and timing pipe."),
            ModuleDefinition(id="MD_LPDDR_CTRL", name="LPDDR5X_CONTROLLER", module_type="memory_controller", ip_owner="Memory Team", reuse_class="replicated", description="LPDDR5X memory controller channel."),
            ModuleDefinition(id="MD_PHY", name="SERDES_PHY", module_type="phy_analog", ip_owner="PHY Team", reuse_class="fixed_hard_macro", description="Fixed IO/PHY hard macro."),
            ModuleDefinition(id="MD_CRYPTO", name="CRYPTO_ENGINE", module_type="security", ip_owner="Security Team", reuse_class="shared_ip", description="AES/SHA/public-key accelerator."),
            ModuleDefinition(id="MD_NOC", name="COHERENT_NOC", module_type="interconnect", ip_owner="Platform Team", reuse_class="platform_fabric", description="Coherent system fabric and QoS."),
            ModuleDefinition(id="MD_CPU_ALU", name="CPU_EXECUTION_UNIT", module_type="cpu_subblock", ip_owner="CPU Team", reuse_class="replicated", description="CPU execution and ALU datapath."),
            ModuleDefinition(id="MD_L1_CACHE", name="CPU_L1_CACHE", module_type="memory_macro", ip_owner="CPU Team", reuse_class="compiler_macro", description="L1 Instruction and Data cache."),
            ModuleDefinition(id="MD_L2_CACHE", name="CPU_L2_CACHE", module_type="memory_macro", ip_owner="CPU Team", reuse_class="compiler_macro", description="L2 cache macro."),
            ModuleDefinition(id="MD_DISPLAY_PIXEL_PROC", name="DISPLAY_PIXEL_PROCESSOR", module_type="display_subblock", ip_owner="Display Team", reuse_class="replicated", description="Pixel blending and color management pipe unit."),
            ModuleDefinition(id="MD_DISPLAY_PIPE_SRAM", name="DISPLAY_LINE_BUFFER", module_type="memory_macro", ip_owner="Display Team", reuse_class="compiler_macro", description="Display line buffer and FIFO SRAM."),
        ]

        owner_overrides = {
            "B0": "Architecture Team",
            "B_CPU": "CPU Team",
            "B_GPU": "GPU Team",
            "B_NPU": "AI Team",
            "B_ISP": "Camera Team",
            "B_MEDIA": "Media Team",
            "B_DISPLAY": "Display Team",
            "B_MODEM": "Modem Team",
            "B_MEM": "Memory Team",
            "B_NOC": "Platform Team",
            "B_IO": "PHY Team",
            "B_SEC": "Security Team",
            "B_PMU": "Power Team",
        }
        owner_by_component: dict[str, str] = {}

        def logical(row: tuple[str, str | None, str | None, str, str, str, str, int, str]) -> LogicalComponent:
            id, parent_id, module_definition_id, name, instance_type, resource_type, function_domain, count, description = row
            path = name if parent_id is None else f"{logical_paths[parent_id]}/{name}"
            logical_paths[id] = path
            owner_team = owner_overrides.get(id) or (owner_by_component[parent_id] if parent_id else "Architecture Team")
            owner_by_component[id] = owner_team
            visibility_level = "public_summary" if id == "B0" else "team"
            return LogicalComponent(id=id, project_id="P001", parent_id=parent_id, module_definition_id=module_definition_id, name=name, instance_type=instance_type, resource_type=resource_type, function_domain=function_domain, hierarchy_path=path, logical_instance_count=count, owner_team=owner_team, visibility_level=visibility_level, description=description, created_at=created, updated_at=created)

        logical_paths: dict[str, str] = {}
        logical_rows: list[tuple[str, str | None, str | None, str, str, str, str, int, str]] = [
            ("B0", None, None, "SOC_TOP", "top", "mixed", "SoC", 1, "Logical root for the Orion X1 mobile SoC."),
            ("B_CPU", "B0", None, "CPU_CLUSTER", "subsystem", "logic+memory", "Compute", 1, "4P+4E CPU cluster with shared DSU/L3."),
            ("B_CPU_P", "B_CPU", "MD_CPU_P_CORE", "P_CORE", "block", "logic", "CPU", 4, "Four high-performance CPU cores."),
            ("B_CPU_P_ALU", "B_CPU_P", "MD_CPU_ALU", "P_CORE_ALU", "block", "logic", "CPU", 1, "CPU Core execution ALU datapath."),
            ("B_CPU_P_CTRL", "B_CPU_P", None, "P_CORE_CTRL", "block", "logic", "CPU", 1, "CPU Core control logic."),
            ("B_CPU_P_L1", "B_CPU_P", "MD_L1_CACHE", "P_CORE_L1_CACHE", "cache", "logic+memory", "CPU Cache", 1, "L1 Instruction and Data cache."),
            ("B_CPU_E", "B_CPU", "MD_CPU_E_CORE", "E_CORE", "block", "logic", "CPU", 4, "Four efficiency CPU cores."),
            ("B_CPU_E_ALU", "B_CPU_E", "MD_CPU_ALU", "E_CORE_ALU", "block", "logic", "CPU", 1, "Efficiency CPU Core execution ALU datapath."),
            ("B_CPU_E_CTRL", "B_CPU_E", None, "E_CORE_CTRL", "block", "logic", "CPU", 1, "Efficiency CPU Core control logic."),
            ("B_CPU_L3", "B_CPU", "MD_SRAM_BANK", "CPU_DSU_L3", "cache", "logic+memory", "CPU Cache", 1, "Shared DSU and 12 MB L3 cache."),
            ("B_GPU", "B0", None, "GPU_TOP", "subsystem", "logic+memory", "Graphics", 1, "Flagship mobile GPU with shader slices and cache."),
            ("B_GPU_SHADER", "B_GPU", "MD_GPU_SHADER", "GPU_SHADER_SLICE", "block", "logic", "Graphics", 6, "Six shader slices."),
            ("B_GPU_L2", "B_GPU", "MD_SRAM_BANK", "GPU_L2_CACHE", "cache", "memory", "Graphics Memory", 2, "GPU L2/cache SRAM banks."),
            ("B_NPU", "B0", None, "NPU_TOP", "subsystem", "logic+memory", "AI", 1, "AI accelerator subsystem."),
            ("B_NPU_TENSOR", "B_NPU", "MD_NPU_TENSOR", "NPU_TENSOR_TILE", "block", "logic", "AI Compute", 8, "Eight tensor compute tiles."),
            ("B_NPU_SRAM", "B_NPU", "MD_SRAM_BANK", "NPU_LOCAL_SRAM", "macro_group", "memory", "AI Memory", 8, "Local SRAM banks for tensor tiles."),
            ("B_NPU_DMA", "B_NPU", None, "NPU_DMA_QOS", "block", "logic", "AI Data Movement", 1, "DMA, command processor, and QoS logic."),
            ("B_ISP", "B0", None, "ISP_TOP", "subsystem", "logic+memory", "Camera", 1, "Triple-camera ISP complex."),
            ("B_ISP_PIPE", "B_ISP", "MD_ISP_PIPE", "ISP_PIPE", "block", "logic", "Camera", 3, "Three concurrent image pipelines."),
            ("B_CV_DSP", "B_ISP", None, "CV_DSP", "block", "logic+memory", "Camera AI", 1, "Computer-vision DSP for camera features."),
            ("B_MEDIA", "B0", None, "MEDIA_TOP", "subsystem", "logic", "Media", 1, "Video encode/decode subsystem."),
            ("B_VDEC", "B_MEDIA", "MD_VIDEO_CODEC", "VIDEO_DECODER", "block", "logic", "Video Decode", 1, "8K/4K multi-format decode engine."),
            ("B_VENC", "B_MEDIA", "MD_VIDEO_CODEC", "VIDEO_ENCODER", "block", "logic", "Video Encode", 1, "4K/8K encode engine."),
            ("B_DISPLAY", "B0", None, "DISPLAY_TOP", "subsystem", "logic", "Display", 1, "Display compositor and panel interface subsystem."),
            ("B_DPU", "B_DISPLAY", "MD_DISPLAY_PIPE", "DISPLAY_PIPE", "block", "logic", "Display", 2, "Dual display pipelines."),
            ("B_DPU_PIXEL_PROC", "B_DPU", "MD_DISPLAY_PIXEL_PROC", "DISPLAY_PIXEL_PROC", "block", "logic", "Display", 1, "Display compositor pixel processor block."),
            ("B_DPU_SRAM", "B_DPU", "MD_DISPLAY_PIPE_SRAM", "DISPLAY_PIPE_FIFO", "cache", "logic+memory", "Display", 1, "Display line buffer and FIFO SRAM."),
            ("B_MODEM", "B0", None, "5G_MODEM_TOP", "subsystem", "logic+memory", "Cellular", 1, "Integrated 5G baseband and RF digital frontend."),
            ("B_MODEM_DSP", "B_MODEM", None, "BASEBAND_DSP", "block", "logic", "Cellular DSP", 2, "Dual vector DSP baseband engines."),
            ("B_MODEM_SRAM", "B_MODEM", "MD_SRAM_BANK", "BASEBAND_SRAM", "macro_group", "memory", "Cellular Memory", 4, "Baseband SRAM banks."),
            ("B_MODEM_RF", "B_MODEM", None, "RF_DIGITAL_FRONTEND", "block", "logic", "Cellular RF", 1, "Digital RF control and calibration logic."),
            ("B_MEM", "B0", None, "MEMORY_SUBSYSTEM", "subsystem", "logic+memory", "Memory", 1, "LPDDR controllers, system cache, and memory fabric."),
            ("B_SYS_CACHE", "B_MEM", "MD_SRAM_BANK", "SYSTEM_LEVEL_CACHE", "cache", "memory", "System Cache", 1, "32 MB system-level cache."),
            ("B_LPDDR_CTRL", "B_MEM", "MD_LPDDR_CTRL", "LPDDR5X_CONTROLLER", "block", "logic", "Memory Controller", 4, "Four LPDDR5X controller channels."),
            ("B_NOC", "B0", "MD_NOC", "SYSTEM_NOC", "subsystem", "logic", "Interconnect", 1, "Coherent NoC, SMMU, DMA fabric, and QoS."),
            ("B_IO", "B0", None, "IO_SUBSYSTEM", "subsystem", "phy_analog", "External IO", 1, "External memory, storage, USB/PCIe, camera/display PHYs."),
            ("B_DDR_PHY", "B_IO", "MD_PHY", "LPDDR5X_PHY", "phy", "phy_analog", "Memory IO", 4, "Four LPDDR5X PHY slices."),
            ("B_UFS_PHY", "B_IO", "MD_PHY", "UFS_PHY", "phy", "phy_analog", "Storage IO", 1, "UFS 4.x PHY."),
            ("B_USB_PCIE_PHY", "B_IO", "MD_PHY", "USB_PCIE_PHY", "phy", "phy_analog", "High-Speed IO", 1, "Shared USB/PCIe PHY complex."),
            ("B_MIPI_PHY", "B_IO", "MD_PHY", "MIPI_PHY", "phy", "phy_analog", "Camera/Display IO", 6, "MIPI D-PHY/C-PHY slices."),
            ("B_SEC", "B0", None, "SECURE_ISLAND", "subsystem", "logic+memory", "Security", 1, "Root-of-trust and secure enclave."),
            ("B_CRYPTO", "B_SEC", "MD_CRYPTO", "CRYPTO_ENGINE", "block", "logic", "Security", 1, "Crypto accelerator."),
            ("B_PMU", "B0", None, "AON_PMU_SENSOR_HUB", "subsystem", "mixed", "Always On", 1, "Always-on PMU, sensor hub, clock/reset, and low-power control."),
        ]
        logical_components = [logical(row) for row in logical_rows]
        responsibility_roots = [
            ("Architecture Team", "arch_lead", "B0"),
            ("CPU Team", "cpu_owner", "B_CPU"),
            ("GPU Team", "gpu_owner", "B_GPU"),
            ("AI Team", "ai_owner", "B_NPU"),
            ("Camera Team", "camera_owner", "B_ISP"),
            ("Media Team", "media_owner", "B_MEDIA"),
            ("Display Team", "display_owner", "B_DISPLAY"),
            ("Modem Team", "modem_owner", "B_MODEM"),
            ("Memory Team", "memory_owner", "B_MEM"),
            ("Platform Team", "platform_owner", "B_NOC"),
            ("PHY Team", "phy_owner", "B_IO"),
            ("Security Team", "security_owner", "B_SEC"),
            ("Power Team", "power_owner", "B_PMU"),
        ]
        responsibilities = [
            ResponsibilityAssignment(
                id=f"RA_{team.upper().replace(' ', '_')}_{root}",
                project_id="P001",
                impl_option_id="S2",
                user_id=user_id,
                team_name=team,
                logical_component_id=root,
                scope_type="subtree",
                can_read=True,
                can_write=True,
            )
            for team, user_id, root in responsibility_roots
        ]
        tiers = [
            Tier(id="T0", impl_option_id="S2", tier_index=0, name="Compute Logic Tier", process_id="PN3E", role="CPU/GPU/NPU/ISP/media high-performance logic", orientation="Face-down", thickness_um=42, area_limit_mm2=300.0, description="Advanced logic tier with hot compute blocks and fine-pitch hybrid bonding."),
            Tier(id="T1", impl_option_id="S2", tier_index=1, name="SRAM + Cache Tier", process_id="PN5", role="Large SRAM/cache plus medium logic", orientation="Face-up / Face-to-face", thickness_um=48, area_limit_mm2=250.0, description="Cache/SRAM-heavy tier serving CPU/GPU/NPU and modem memories."),
            Tier(id="T2", impl_option_id="S2", tier_index=2, name="IO + Always-On Tier", process_id="PN6", role="LPDDR/PHY/IO/security/AON/analog-friendly logic", orientation="Backside PDN", thickness_um=60, area_limit_mm2=180.0, description="Mature-node companion tier for IO PHYs, PMU, RF digital, and low-power islands."),
        ]

        logical_metric_values = {
            "B0": (2200, 84.0, 168.0, 285.0, 0.0), "B_CPU": (620, 10.0, 28.0, 42.0, 9.8),
            "B_CPU_P": (320, 0.0, 18.4, 20.2, 7.8),
            "B_CPU_P_ALU": (160, 0.0, 10.0, 10.0, 4.0),
            "B_CPU_P_CTRL": (80, 0.0, 5.0, 5.0, 2.0),
            "B_CPU_P_L1": (80, 0.0, 2.0, 3.2, 1.0),
            "B_CPU_E": (220, 0.0, 5.8, 6.5, 2.2),
            "B_CPU_E_ALU": (120, 0.0, 3.5, 4.0, 1.2),
            "B_CPU_E_CTRL": (60, 0.0, 1.5, 1.5, 0.6),
            "B_CPU_L3": (180, 9.6, 2.1, 12.8, 1.6),
            "B_GPU": (740, 10.5, 30.8, 44.0, 9.4), "B_GPU_SHADER": (520, 0.0, 27.6, 30.2, 8.6), "B_GPU_L2": (96, 8.6, 0.8, 9.9, 0.8), "B_NPU": (680, 14.2, 24.7, 43.5, 7.2), "B_NPU_TENSOR": (360, 0.0, 20.8, 22.4, 6.4), "B_NPU_SRAM": (128, 12.8, 0.9, 14.2, 0.7), "B_NPU_DMA": (104, 0.8, 2.5, 3.6, 0.6),
            "B_ISP": (420, 4.0, 13.3, 19.0, 3.6), "B_ISP_PIPE": (240, 1.8, 9.6, 12.0, 2.8), "B_CV_DSP": (150, 1.2, 3.7, 5.2, 0.9), "B_MEDIA": (240, 1.2, 7.8, 10.0, 2.1), "B_VDEC": (120, 0.5, 3.6, 4.6, 0.9), "B_VENC": (130, 0.6, 4.2, 5.3, 1.1), "B_DISPLAY": (150, 0.8, 4.2, 5.6, 1.0),
            "B_DPU": (120, 0.6, 3.6, 4.8, 0.9),
            "B_DPU_PIXEL_PROC": (70, 0.0, 2.0, 3.0, 0.5),
            "B_DPU_SRAM": (40, 0.4, 0.5, 1.0, 0.2),
            "B_MODEM": (520, 7.8, 19.6, 31.0, 5.4), "B_MODEM_DSP": (260, 1.6, 10.2, 12.6, 3.2), "B_MODEM_SRAM": (90, 5.4, 0.7, 6.4, 0.5), "B_MODEM_RF": (110, 0.8, 3.8, 5.1, 0.9), "B_MEM": (260, 32.0, 6.0, 42.0, 2.5), "B_SYS_CACHE": (80, 25.6, 0.9, 28.2, 1.1), "B_LPDDR_CTRL": (180, 1.2, 4.8, 6.1, 1.4), "B_NOC": (300, 0.8, 8.5, 10.5, 2.2),
            "B_IO": (210, 0.5, 18.0, 24.0, 2.0), "B_DDR_PHY": (70, 0.0, 8.8, 9.6, 1.1), "B_UFS_PHY": (24, 0.0, 1.5, 1.8, 0.2), "B_USB_PCIE_PHY": (42, 0.0, 2.8, 3.4, 0.4), "B_MIPI_PHY": (52, 0.0, 3.6, 4.5, 0.3), "B_SEC": (95, 1.2, 2.2, 4.0, 0.6), "B_CRYPTO": (50, 0.2, 1.2, 1.7, 0.4), "B_PMU": (86, 0.8, 2.8, 4.2, 0.3),
        }
        logical_by_id = {component.id: component for component in logical_components}
        children_by_parent: dict[str, list[LogicalComponent]] = {}
        for component in logical_components:
            if component.parent_id:
                children_by_parent.setdefault(component.parent_id, []).append(component)

        def required_categories_for_seed(component_id: str) -> list[str]:
            values = logical_metric_values[component_id]
            child_values = [logical_metric_values[child.id] for child in children_by_parent.get(component_id, []) if child.id in logical_metric_values]
            if child_values:
                sram_area = values[1] - sum(row[1] for row in child_values)
                logic_area = values[2] - sum(row[2] for row in child_values)
                block_area = values[3] - sum(row[3] for row in child_values)
            else:
                sram_area, logic_area, block_area = values[1], values[2], values[3]
            categories = []
            if logic_area > 0.001:
                categories.append("logic")
            if sram_area > 0.001:
                categories.append("sram")
            if block_area > 0.001:
                categories.append("block")
            return categories or ["block"]

        mapping_specs = [
            ("B0", [("T0", 1, 0.45), ("T1", 1, 0.35), ("T2", 1, 0.20)]),
            ("B_CPU", [("T0", 1, 1.00)]),
            ("B_CPU_P", [("T0", 4, 1.00)]),
            ("B_CPU_P_ALU", [("T0", 4, 1.00)]),
            ("B_CPU_P_CTRL", [("T0", 4, 1.00)]),
            ("B_CPU_P_L1", [("T0", 4, 1.00)]),
            ("B_CPU_E", [("T0", 4, 1.00)]),
            ("B_CPU_E_ALU", [("T0", 4, 1.00)]),
            ("B_CPU_E_CTRL", [("T0", 4, 1.00)]),
            ("B_CPU_L3", [("T0", 1, 0.25), ("T1", 1, 0.75)]),
            ("B_GPU_SHADER", [("T0", 6, 1.00)]),
            ("B_GPU_L2", [("T1", 2, 1.00)]),
            ("B_GPU", [("T0", 1, 0.70), ("T1", 1, 0.30)]),
            ("B_NPU", [("T0", 1, 0.60), ("T1", 1, 0.40)]),
            ("B_NPU_TENSOR", [("T0", 4, 1.00), ("T1", 4, 1.00)]),
            ("B_NPU_SRAM", [("T1", 8, 1.00)]),
            ("B_NPU_DMA", [("T0", 1, 0.60), ("T1", 1, 0.40)]),
            ("B_ISP", [("T0", 1, 1.00)]),
            ("B_ISP_PIPE", [("T0", 2, 1.00), ("T1", 1, 1.00)]),
            ("B_CV_DSP", [("T0", 1, 1.00)]),
            ("B_MEDIA", [("T0", 1, 1.00)]),
            ("B_VDEC", [("T0", 1, 1.00)]),
            ("B_VENC", [("T0", 1, 1.00)]),
            ("B_DISPLAY", [("T0", 1, 1.00)]),
            ("B_DPU", [("T0", 2, 1.00)]),
            ("B_DPU_PIXEL_PROC", [("T0", 2, 1.00)]),
            ("B_DPU_SRAM", [("T0", 2, 1.00)]),
            ("B_MODEM", [("T2", 1, 1.00)]),
            ("B_MODEM_DSP", [("T0", 2, 1.00)]),
            ("B_MODEM_SRAM", [("T1", 4, 1.00)]),
            ("B_MODEM_RF", [("T2", 1, 1.00)]),
            ("B_MEM", [("T1", 1, 0.60), ("T2", 1, 0.40)]),
            ("B_SYS_CACHE", [("T1", 1, 1.00)]),
            ("B_LPDDR_CTRL", [("T2", 4, 1.00)]),
            ("B_NOC", [("T0", 1, 0.55), ("T1", 1, 0.30), ("T2", 1, 0.15)]),
            ("B_IO", [("T2", 1, 1.00)]),
            ("B_DDR_PHY", [("T2", 4, 1.00)]),
            ("B_UFS_PHY", [("T2", 1, 1.00)]),
            ("B_USB_PCIE_PHY", [("T2", 1, 1.00)]),
            ("B_MIPI_PHY", [("T2", 6, 1.00)]),
            ("B_CRYPTO", [("T2", 1, 1.00)]),
            ("B_SEC", [("T2", 1, 1.00)]),
            ("B_PMU", [("T2", 1, 1.00)]),
        ]
        partition_rows: list[tuple[str, str, str, str, str, str, int, float]] = []
        for component_id, placements in mapping_specs:
            comp = logical_by_id[component_id]
            component_name = comp.name
            
            # calculate parent multiplier to convert absolute counts in mapping_specs to relative
            curr = comp.parent_id
            parent_mult = 1
            while curr:
                p = logical_by_id[curr]
                parent_mult *= p.logical_instance_count
                curr = p.parent_id
                
            relative_placements = []
            for tier_id, count, share in placements:
                rel_count = count / parent_mult
                if rel_count.is_integer():
                    rel_count = int(rel_count)
                relative_placements.append((tier_id, rel_count, share))
                
            for category in required_categories_for_seed(component_id):
                partial_index = 0
                full_copy_count = sum(c for _, c, share in relative_placements if abs(share - 1.0) < 0.001)
                is_split = not (full_copy_count == comp.logical_instance_count and all(abs(share - 1.0) < 0.001 for _, _, share in relative_placements))
                for tier_id, count, share in relative_placements:
                    partition_type = "partial" if is_split else "full"
                    if partition_type == "partial":
                        partial_index += 1
                    partition_name = canonical_partition_name(component_name, category, tier_id, partition_type, partial_index)
                    partition_rows.append((f"PP_{partition_name}", component_id, tier_id, partition_name, partition_type, category, count, share))
        partitions = [
            PhysicalPartition(id=id, impl_option_id="S2", logical_component_id=logical_id, tier_id=tier_id, partition_name=name, partition_type=ptype, resource_category=category, physical_instance_count=count, partition_ratio=1.0 if ptype == "full" else ratio, content_share=1.0 if ptype == "full" else ratio, description=f"{name} maps {category} content of {logical_id} to {tier_id}.")
            for id, logical_id, tier_id, name, ptype, category, count, ratio in partition_rows
        ]
        def category_area_for_seed(component_id: str, category: str) -> float:
            values = logical_metric_values[component_id]
            child_values = [logical_metric_values[child.id] for child in children_by_parent.get(component_id, []) if child.id in logical_metric_values]
            index = {"sram": 1, "logic": 2, "block": 3}[category]
            value = values[index] - sum(row[index] for row in child_values) if child_values else values[index]
            return max(0.0, value)

        partition_metric_values: dict[str, tuple[float, float, float, float, str]] = {}
        rows_by_component_category: dict[tuple[str, str], list[tuple[str, str, str, str, str, str, int, float]]] = {}
        for row in partition_rows:
            rows_by_component_category.setdefault((row[1], row[5]), []).append(row)
        for (component_id, category), rows in rows_by_component_category.items():
            category_area = category_area_for_seed(component_id, category)
            total_equivalent = sum(count * (1.0 if ptype == "full" else ratio) for _, _, _, _, ptype, _, count, ratio in rows) or 1.0
            component_power = logical_metric_values[component_id][4]
            category_count = len(required_categories_for_seed(component_id)) or 1
            for partition_id, _logical_id, tier_id, _name, ptype, _category, count, ratio in rows:
                equivalent = count * (1.0 if ptype == "full" else ratio)
                share = equivalent / total_equivalent
                logic_area = round(category_area * share, 3) if category == "logic" else 0.0
                sram_area = round(category_area * share, 3) if category == "sram" else 0.0
                block_area = round(category_area * share, 3) if category == "block" else 0.0
                power = round(component_power * share / category_count, 3)
                partition_metric_values[partition_id] = (logic_area, sram_area, block_area, power, f"{category}_{tier_id.lower()}")
        metrics: list[Metric] = []
        for component_id, (signals, sram_area, logic_area, block_area, power) in logical_metric_values.items():
            for name, value, unit, category, workload in [("signal_count_total", signals, "count", "logical", "nominal"), ("logic_area", logic_area, "mm2", "logical_area", "nominal"), ("sram_area", sram_area, "mm2", "logical_area", "nominal"), ("block_area", block_area, "mm2", "logical_area", "nominal"), ("power", power, "W", "power", "peak")]:
                metrics.append(metric(f"M_LOG_{component_id}_{name.upper()}", "S2", "logical_component", component_id, name, value, unit, category, "number", "typical", workload, "review", "Architecture planning estimate for the realistic mobile SoC demo."))
        for partition_id, (logic_area, sram_area, block_area, power, shape_type) in partition_metric_values.items():
            for name, value, unit, category, workload, value_type in [("logic_area", logic_area, "mm2", "implementation_area", "nominal", "number"), ("sram_area", sram_area, "mm2", "implementation_area", "nominal", "number"), ("block_area", block_area, "mm2", "implementation_area", "nominal", "number"), ("power", power, "W", "power", "peak", "number"), ("shape_type", shape_type, "", "physical_shape", "nominal", "text")]:
                metrics.append(metric(f"M_PART_{partition_id}_{name.upper()}", "S2", "physical_partition", partition_id, name, value, unit, category, value_type, "typical", workload, "review", "Physical partition estimate for the realistic mobile SoC demo."))
        metrics.extend([
            metric("M_TIER_T0_POWER", "S2", "tier", "T0", "power", 31.2, "W", "power", "number", "typical", "peak", "review", "Compute-tier peak budget."),
            metric("M_TIER_T1_POWER", "S2", "tier", "T1", "power", 8.7, "W", "power", "number", "typical", "peak", "review", "SRAM/cache-tier peak budget."),
            metric("M_TIER_T2_POWER", "S2", "tier", "T2", "power", 5.4, "W", "power", "number", "typical", "peak", "review", "IO/always-on tier peak budget."),
            metric("M_TIER_T0_UTIL", "S2", "tier", "T0", "utilization", 78, "%", "physical", "number", "typical", "nominal", "review", "Tier utilization target."),
            metric("M_TIER_T1_UTIL", "S2", "tier", "T1", "utilization", 71, "%", "physical", "number", "typical", "nominal", "review", "Tier utilization target."),
            metric("M_TIER_T2_UTIL", "S2", "tier", "T2", "utilization", 63, "%", "physical", "number", "typical", "nominal", "review", "Tier utilization target."),
            metric("M_IMPL_OPTION_S1_AREA", "S1", "impl_option", "S1", "area", 182.0, "mm2", "physical", "number", "typical", "nominal", "review", "Monolithic N3E planning estimate including keepout and IO ring."),
            metric("M_IMPL_OPTION_S1_POWER", "S1", "impl_option", "S1", "power", 49.5, "W", "power", "number", "typical", "peak", "review", "Monolithic peak workload power."),
            metric("M_IMPL_OPTION_S2_AREA", "S2", "impl_option", "S2", "area", 119.0, "mm2", "physical", "number", "typical", "nominal", "review", "Projected exposed package footprint for the 3-tier option."),
            metric("M_IMPL_OPTION_S2_POWER", "S2", "impl_option", "S2", "power", 45.3, "W", "power", "number", "typical", "peak", "review", "3DIC peak workload power with shorter memory paths."),
            metric("M_IMPL_OPTION_S3_AREA", "S3", "impl_option", "S3", "area", 143.0, "mm2", "physical", "number", "typical", "nominal", "review", "2.5D option exposed package footprint estimate."),
            metric("M_IMPL_OPTION_S3_POWER", "S3", "impl_option", "S3", "power", 47.0, "W", "power", "number", "typical", "peak", "review", "2.5D option peak workload power."),
        ])

        app_scenarios = [
            ApplicationScenario(id="AS_MODULE_LIBRARY", project_id="P001", name="Module Use Case Library", category="Internal", description="Internal library bucket for module use case/Profile power values. Application scenarios select from these rows."),
            ApplicationScenario(id="AS_STANDBY_AOD", project_id="P001", name="Standby Always-On Display", category="Everyday", description="Phone in pocket or on desk with AOD, sensor hub, modem paging, and minimal memory retention."),
            ApplicationScenario(id="AS_UI_BROWSING", project_id="P001", name="Interactive UI + Web Browsing", category="Everyday", description="Foreground browser/social feed scrolling with display, modem, CPU bursts, light GPU composition, and memory traffic."),
            ApplicationScenario(id="AS_VIDEO_PLAYBACK", project_id="P001", name="4K HDR Video Playback", category="Multimedia", description="Local or streaming 4K HDR playback using video decode, display pipeline, memory subsystem, and light CPU control."),
            ApplicationScenario(id="AS_CAMERA_4K60", project_id="P001", name="Camera 4K60 Recording", category="Multimedia", description="Sustained 4K 60FPS recording with ISP, video encode, NPU denoise, display preview, IO, and memory traffic."),
            ApplicationScenario(id="AS_GAMING_SUSTAINED", project_id="P001", name="3D Gaming Sustained", category="Gaming", description="Thermal-sustained 3D gaming with GPU rendering, CPU game threads, display refresh, memory bandwidth, and fabric traffic."),
            ApplicationScenario(id="AS_AI_BURST", project_id="P001", name="AI Photo Enhancement Burst", category="AI", description="Short burst of local AI photo enhancement using NPU tensor compute, CPU dispatch, ISP preprocessing, and memory movement."),
            ApplicationScenario(id="AS_5G_VIDEO_CALL", project_id="P001", name="5G Video Call", category="Connectivity", description="Two-way video call with 5G modem, camera pipeline, video codec, display, CPU control, and IO activity."),
        ]
        
        phys_mappings = [
            PhysicalMapping(id="PM_2D_BASE", impl_option_id="S1", name="2D_BASELINE_MAPPING_V01", mapping_version="V01", description="Baseline monolithic 2D die mapping.", mapping_json='{"SOC_TOP": "monolithic"}'),
            PhysicalMapping(id="PM_3DIC_A", impl_option_id="S2", name="3DIC_A_MAPPING_V02", mapping_version="V02", description="3DIC split-die mapping with SRAM cache stacked on Middle Tier.", mapping_json='{"NPU_TOP": "T0/T1 split", "GPU_TOP": "T0/T1 split", "CPU_CLUSTER": "T0"}'),
        ]
        
        op_point_sets = [
            OperatingPointSet(id="OP_DEFAULT", project_id="P001", name="Default", description="Default module Profile. Available to every module; it belongs to a module only after a saved module power value uses it.", op_json="{}"),
            OperatingPointSet(id="OP_STANDBY_AOD", project_id="P001", name="Standby_AOD", description="Low-voltage always-on state for AOD and sensor/modem paging.", op_json='{"SOC_TOP": {"mode": "retention"}, "AON_PMU_SENSOR_HUB": {"voltage_v": 0.55, "frequency_mhz": 26}, "DISPLAY_TOP": {"refresh_hz": 1}}'),
            OperatingPointSet(id="OP_UI_BALANCED", project_id="P001", name="UI_Balanced", description="Balanced interactive profile for browsing and app UI.", op_json='{"CPU_CLUSTER": {"voltage_v": 0.72, "frequency_mhz": 1800}, "GPU_TOP": {"voltage_v": 0.62, "frequency_mhz": 350}, "DISPLAY_TOP": {"refresh_hz": 120}}'),
            OperatingPointSet(id="OP_VIDEO_PLAYBACK", project_id="P001", name="Video_Playback", description="Media playback profile with video decoder and display active.", op_json='{"MEDIA_TOP": {"voltage_v": 0.68, "frequency_mhz": 600}, "DISPLAY_TOP": {"refresh_hz": 60}, "CPU_CLUSTER": {"frequency_mhz": 900}}'),
            OperatingPointSet(id="OP_CAMERA_4K60", project_id="P001", name="Camera_4K60", description="Camera capture profile calibrated for sustained 4K60 recording.", op_json='{"ISP_TOP": {"voltage_v": 0.72, "frequency_mhz": 900}, "VIDEO_ENCODER": {"frequency_mhz": 700}, "NPU_TOP": {"voltage_v": 0.70, "frequency_mhz": 1000}}'),
            OperatingPointSet(id="OP_GAMING_SUSTAINED", project_id="P001", name="Gaming_Sustained", description="Sustained thermal-limit operating points for gaming.", op_json='{"CPU_CLUSTER": {"voltage_v": 0.80, "frequency_mhz": 2200}, "GPU_TOP": {"voltage_v": 0.72, "frequency_mhz": 650}, "DISPLAY_TOP": {"refresh_hz": 120}}'),
            OperatingPointSet(id="OP_AI_BURST", project_id="P001", name="AI_Burst", description="Peak performance burst operating points for short local AI runs.", op_json='{"CPU_CLUSTER": {"voltage_v": 0.86, "frequency_mhz": 2600}, "NPU_TOP": {"voltage_v": 0.82, "frequency_mhz": 1500}, "MEMORY_SUBSYSTEM": {"bandwidth_gbps": 90}}'),
            OperatingPointSet(id="OP_5G_VIDEO_CALL", project_id="P001", name="5G_Video_Call", description="Connectivity and media profile for uplink/downlink video call.", op_json='{"5G_MODEM_TOP": {"mode": "sub6_active"}, "ISP_TOP": {"frequency_mhz": 550}, "VIDEO_ENCODER": {"frequency_mhz": 450}, "DISPLAY_TOP": {"refresh_hz": 60}}'),
        ]
        def safe_key(value: str | None) -> str:
            return (value or "Default").replace(" ", "_").replace("/", "_").replace("-", "_").upper()

        module_usecase_specs = [
            ("B0", "AOD_System_Total", "OP_STANDBY_AOD", 0.145, "review"),
            ("B_PMU", "AOD_SensorHub", "OP_STANDBY_AOD", 0.038, "approved"),
            ("B_DISPLAY", "AOD_1Hz_Panel", "OP_STANDBY_AOD", 0.045, "approved"),
            ("B_MODEM", "Modem_Paging", "OP_STANDBY_AOD", 0.024, "review"),
            ("B_MEM", "Memory_Retention", "OP_STANDBY_AOD", 0.018, "review"),
            ("B_SEC", "Secure_Heartbeat", "OP_STANDBY_AOD", 0.006, "review"),
            ("B_IO", "LowSpeed_IO_Idle", "OP_STANDBY_AOD", 0.010, "review"),

            ("B0", "UI_Browsing_Total", "OP_UI_BALANCED", 2.65, "review"),
            ("B_CPU", "UI_App_Bursts", "OP_UI_BALANCED", 0.78, "approved"),
            ("B_GPU", "UI_Composition", "OP_UI_BALANCED", 0.34, "approved"),
            ("B_DISPLAY", "Display_120Hz_UI", "OP_UI_BALANCED", 0.48, "approved"),
            ("B_MODEM", "Web_Data_Bursty", "OP_UI_BALANCED", 0.26, "review"),
            ("B_MEM", "Browser_Memory_Traffic", "OP_UI_BALANCED", 0.32, "review"),
            ("B_NOC", "Interactive_Fabric", "OP_UI_BALANCED", 0.18, "review"),
            ("B_IO", "Storage_And_MIPI_Active", "OP_UI_BALANCED", 0.11, "review"),
            ("B_PMU", "Active_Power_Control", "OP_UI_BALANCED", 0.055, "review"),
            ("B_SEC", "TLS_Crypto_Assist", "OP_UI_BALANCED", 0.035, "review"),

            ("B0", "Video_Playback_Total", "OP_VIDEO_PLAYBACK", 1.95, "approved"),
            ("B_MEDIA", "Video_Decode_4K_HDR", "OP_VIDEO_PLAYBACK", 0.58, "approved"),
            ("B_DISPLAY", "HDR_Display_60Hz", "OP_VIDEO_PLAYBACK", 0.56, "approved"),
            ("B_CPU", "Playback_Control", "OP_VIDEO_PLAYBACK", 0.18, "approved"),
            ("B_GPU", "Video_Overlay_Composition", "OP_VIDEO_PLAYBACK", 0.08, "review"),
            ("B_MEM", "Video_Buffer_Traffic", "OP_VIDEO_PLAYBACK", 0.28, "review"),
            ("B_NOC", "Media_Fabric", "OP_VIDEO_PLAYBACK", 0.13, "review"),
            ("B_IO", "Display_PHY_Active", "OP_VIDEO_PLAYBACK", 0.09, "review"),
            ("B_PMU", "Playback_Power_Control", "OP_VIDEO_PLAYBACK", 0.045, "review"),

            ("B0", "Camera_4K60_Total", "OP_CAMERA_4K60", 5.35, "review"),
            ("B_ISP", "ISP_4K60_TriplePipe", "OP_CAMERA_4K60", 1.20, "approved"),
            ("B_MEDIA", "Video_Encode_4K60", "OP_CAMERA_4K60", 0.78, "approved"),
            ("B_NPU", "NPU_Denoise_HDR", "OP_CAMERA_4K60", 0.90, "review"),
            ("B_CPU", "Camera_App_Control", "OP_CAMERA_4K60", 0.45, "approved"),
            ("B_GPU", "Preview_Composition", "OP_CAMERA_4K60", 0.25, "review"),
            ("B_DISPLAY", "Camera_Preview_Display", "OP_CAMERA_4K60", 0.28, "approved"),
            ("B_MEM", "Camera_Frame_Buffer", "OP_CAMERA_4K60", 0.58, "review"),
            ("B_NOC", "Camera_Fabric_Traffic", "OP_CAMERA_4K60", 0.32, "review"),
            ("B_IO", "MIPI_CSI_DPHY_Active", "OP_CAMERA_4K60", 0.46, "review"),
            ("B_PMU", "Camera_Power_Control", "OP_CAMERA_4K60", 0.075, "review"),

            ("B0", "Gaming_Sustained_Total", "OP_GAMING_SUSTAINED", 6.95, "review"),
            ("B_GPU", "GPU_3D_Render_120Hz", "OP_GAMING_SUSTAINED", 3.25, "approved"),
            ("B_CPU", "CPU_Game_Threads", "OP_GAMING_SUSTAINED", 1.35, "approved"),
            ("B_DISPLAY", "Gaming_Display_120Hz", "OP_GAMING_SUSTAINED", 0.58, "approved"),
            ("B_MEM", "Gaming_Memory_Bandwidth", "OP_GAMING_SUSTAINED", 0.72, "review"),
            ("B_NOC", "Gaming_Fabric", "OP_GAMING_SUSTAINED", 0.46, "review"),
            ("B_NPU", "Game_AI_Assist_Light", "OP_GAMING_SUSTAINED", 0.16, "review"),
            ("B_IO", "Touch_Audio_IO", "OP_GAMING_SUSTAINED", 0.15, "review"),
            ("B_PMU", "Thermal_Power_Control", "OP_GAMING_SUSTAINED", 0.085, "review"),
            ("B_SEC", "Game_Security_Services", "OP_GAMING_SUSTAINED", 0.025, "review"),

            ("B0", "AI_Photo_Burst_Total", "OP_AI_BURST", 6.35, "measured"),
            ("B_NPU", "NPU_Photo_Enhance", "OP_AI_BURST", 3.75, "measured"),
            ("B_CPU", "AI_Dispatch_PrePost", "OP_AI_BURST", 0.82, "approved"),
            ("B_GPU", "AI_Preview_Composition", "OP_AI_BURST", 0.25, "review"),
            ("B_ISP", "Photo_Preprocess", "OP_AI_BURST", 0.36, "review"),
            ("B_MEM", "AI_Tensor_Buffer_Traffic", "OP_AI_BURST", 0.66, "review"),
            ("B_NOC", "AI_Fabric_Traffic", "OP_AI_BURST", 0.36, "review"),
            ("B_IO", "Image_Storage_IO", "OP_AI_BURST", 0.08, "review"),
            ("B_PMU", "Burst_Power_Control", "OP_AI_BURST", 0.06, "review"),

            ("B0", "5G_Video_Call_Total", "OP_5G_VIDEO_CALL", 3.70, "review"),
            ("B_MODEM", "5G_Sub6_VideoCall", "OP_5G_VIDEO_CALL", 1.20, "approved"),
            ("B_ISP", "FrontCamera_1080p", "OP_5G_VIDEO_CALL", 0.45, "approved"),
            ("B_MEDIA", "VideoCall_Codec", "OP_5G_VIDEO_CALL", 0.32, "approved"),
            ("B_CPU", "VideoCall_App_Stack", "OP_5G_VIDEO_CALL", 0.56, "approved"),
            ("B_DISPLAY", "VideoCall_Display", "OP_5G_VIDEO_CALL", 0.40, "approved"),
            ("B_MEM", "VideoCall_Memory_Traffic", "OP_5G_VIDEO_CALL", 0.28, "review"),
            ("B_NOC", "VideoCall_Fabric", "OP_5G_VIDEO_CALL", 0.23, "review"),
            ("B_IO", "RF_MIPI_USB_IO", "OP_5G_VIDEO_CALL", 0.18, "review"),
            ("B_PMU", "Connectivity_Power_Control", "OP_5G_VIDEO_CALL", 0.065, "review"),
        ]

        scenario_compositions = {
            "AS_STANDBY_AOD": [("B0", "AOD_System_Total", "OP_STANDBY_AOD", False), ("B_PMU", "AOD_SensorHub", "OP_STANDBY_AOD", True), ("B_DISPLAY", "AOD_1Hz_Panel", "OP_STANDBY_AOD", True), ("B_MODEM", "Modem_Paging", "OP_STANDBY_AOD", True), ("B_MEM", "Memory_Retention", "OP_STANDBY_AOD", True), ("B_SEC", "Secure_Heartbeat", "OP_STANDBY_AOD", True), ("B_IO", "LowSpeed_IO_Idle", "OP_STANDBY_AOD", True)],
            "AS_UI_BROWSING": [("B0", "UI_Browsing_Total", "OP_UI_BALANCED", False), ("B_CPU", "UI_App_Bursts", "OP_UI_BALANCED", True), ("B_GPU", "UI_Composition", "OP_UI_BALANCED", True), ("B_DISPLAY", "Display_120Hz_UI", "OP_UI_BALANCED", True), ("B_MODEM", "Web_Data_Bursty", "OP_UI_BALANCED", True), ("B_MEM", "Browser_Memory_Traffic", "OP_UI_BALANCED", True), ("B_NOC", "Interactive_Fabric", "OP_UI_BALANCED", True), ("B_IO", "Storage_And_MIPI_Active", "OP_UI_BALANCED", True), ("B_PMU", "Active_Power_Control", "OP_UI_BALANCED", True), ("B_SEC", "TLS_Crypto_Assist", "OP_UI_BALANCED", True)],
            "AS_VIDEO_PLAYBACK": [("B0", "Video_Playback_Total", "OP_VIDEO_PLAYBACK", False), ("B_MEDIA", "Video_Decode_4K_HDR", "OP_VIDEO_PLAYBACK", True), ("B_DISPLAY", "HDR_Display_60Hz", "OP_VIDEO_PLAYBACK", True), ("B_CPU", "Playback_Control", "OP_VIDEO_PLAYBACK", True), ("B_GPU", "Video_Overlay_Composition", "OP_VIDEO_PLAYBACK", True), ("B_MEM", "Video_Buffer_Traffic", "OP_VIDEO_PLAYBACK", True), ("B_NOC", "Media_Fabric", "OP_VIDEO_PLAYBACK", True), ("B_IO", "Display_PHY_Active", "OP_VIDEO_PLAYBACK", True), ("B_PMU", "Playback_Power_Control", "OP_VIDEO_PLAYBACK", True)],
            "AS_CAMERA_4K60": [("B0", "Camera_4K60_Total", "OP_CAMERA_4K60", False), ("B_ISP", "ISP_4K60_TriplePipe", "OP_CAMERA_4K60", True), ("B_MEDIA", "Video_Encode_4K60", "OP_CAMERA_4K60", True), ("B_NPU", "NPU_Denoise_HDR", "OP_CAMERA_4K60", True), ("B_CPU", "Camera_App_Control", "OP_CAMERA_4K60", True), ("B_GPU", "Preview_Composition", "OP_CAMERA_4K60", True), ("B_DISPLAY", "Camera_Preview_Display", "OP_CAMERA_4K60", True), ("B_MEM", "Camera_Frame_Buffer", "OP_CAMERA_4K60", True), ("B_NOC", "Camera_Fabric_Traffic", "OP_CAMERA_4K60", True), ("B_IO", "MIPI_CSI_DPHY_Active", "OP_CAMERA_4K60", True), ("B_PMU", "Camera_Power_Control", "OP_CAMERA_4K60", True)],
            "AS_GAMING_SUSTAINED": [("B0", "Gaming_Sustained_Total", "OP_GAMING_SUSTAINED", False), ("B_GPU", "GPU_3D_Render_120Hz", "OP_GAMING_SUSTAINED", True), ("B_CPU", "CPU_Game_Threads", "OP_GAMING_SUSTAINED", True), ("B_DISPLAY", "Gaming_Display_120Hz", "OP_GAMING_SUSTAINED", True), ("B_MEM", "Gaming_Memory_Bandwidth", "OP_GAMING_SUSTAINED", True), ("B_NOC", "Gaming_Fabric", "OP_GAMING_SUSTAINED", True), ("B_NPU", "Game_AI_Assist_Light", "OP_GAMING_SUSTAINED", True), ("B_IO", "Touch_Audio_IO", "OP_GAMING_SUSTAINED", True), ("B_PMU", "Thermal_Power_Control", "OP_GAMING_SUSTAINED", True), ("B_SEC", "Game_Security_Services", "OP_GAMING_SUSTAINED", True)],
            "AS_AI_BURST": [("B0", "AI_Photo_Burst_Total", "OP_AI_BURST", False), ("B_NPU", "NPU_Photo_Enhance", "OP_AI_BURST", True), ("B_CPU", "AI_Dispatch_PrePost", "OP_AI_BURST", True), ("B_GPU", "AI_Preview_Composition", "OP_AI_BURST", True), ("B_ISP", "Photo_Preprocess", "OP_AI_BURST", True), ("B_MEM", "AI_Tensor_Buffer_Traffic", "OP_AI_BURST", True), ("B_NOC", "AI_Fabric_Traffic", "OP_AI_BURST", True), ("B_IO", "Image_Storage_IO", "OP_AI_BURST", True), ("B_PMU", "Burst_Power_Control", "OP_AI_BURST", True)],
            "AS_5G_VIDEO_CALL": [("B0", "5G_Video_Call_Total", "OP_5G_VIDEO_CALL", False), ("B_MODEM", "5G_Sub6_VideoCall", "OP_5G_VIDEO_CALL", True), ("B_ISP", "FrontCamera_1080p", "OP_5G_VIDEO_CALL", True), ("B_MEDIA", "VideoCall_Codec", "OP_5G_VIDEO_CALL", True), ("B_CPU", "VideoCall_App_Stack", "OP_5G_VIDEO_CALL", True), ("B_DISPLAY", "VideoCall_Display", "OP_5G_VIDEO_CALL", True), ("B_MEM", "VideoCall_Memory_Traffic", "OP_5G_VIDEO_CALL", True), ("B_NOC", "VideoCall_Fabric", "OP_5G_VIDEO_CALL", True), ("B_IO", "RF_MIPI_USB_IO", "OP_5G_VIDEO_CALL", True), ("B_PMU", "Connectivity_Power_Control", "OP_5G_VIDEO_CALL", True)],
        }

        module_power_obs = [
            PowerObservation(
                id=f"PUC_S2_PM_3DIC_A_{safe_key(component_id)}_{safe_key(use_case)}_{safe_key(op_id)}",
                project_id="P001",
                impl_option_id="S2",
                physical_mapping_id="PM_3DIC_A",
                application_scenario_id="AS_MODULE_LIBRARY",
                operating_point_set_id=op_id,
                scope_type="component",
                scope_id=component_id,
                scope_name=logical_by_id[component_id].name,
                use_case_name=use_case,
                time_window_name="steady_state",
                statistic_type="average",
                power_type="total",
                power_value_w=power_w,
                development_stage="architecture_estimate",
                source_type="module_usecase_seed",
                confidence=confidence,
                is_additive=True,
                note="Realistic mobile SoC demo module use case/Profile power value.",
            )
            for component_id, use_case, op_id, power_w, confidence in module_usecase_specs
        ]
        scenario_selections = [
            ApplicationScenarioSelection(
                id=f"ASC_{safe_key(scenario_id)}_{safe_key(component_id)}_{safe_key(use_case)}_{safe_key(op_id)}",
                project_id="P001",
                impl_option_id="S2",
                physical_mapping_id="PM_3DIC_A",
                application_scenario_id=scenario_id,
                component_id=component_id,
                component_name=logical_by_id[component_id].name,
                use_case_name=use_case,
                operating_point_set_id=op_id,
                included=included,
                note="Seed scenario composition selection. SOC_TOP rows are inactive references for roll-up comparison.",
            )
            for scenario_id, rows in scenario_compositions.items()
            for component_id, use_case, op_id, included in rows
        ]
        power_obs: list[PowerObservation] = []
        
        for row in projects + implOptions + process_nodes + module_definitions + logical_components + tiers + partitions + metrics + responsibilities + app_scenarios + phys_mappings + op_point_sets + power_obs + module_power_obs + scenario_selections:
            session.merge(row)
        session.commit()


def metric(
    id: str,
    impl_option_id: str,
    subject_type: str,
    subject_id: str,
    metric_name: str,
    metric_value: object,
    metric_unit: str,
    metric_category: str,
    value_type: str,
    corner: str,
    workload: str,
    confidence: str,
    source_note: str,
) -> Metric:
    return Metric(
        id=id,
        impl_option_id=impl_option_id,
        subject_type=subject_type,
        subject_id=subject_id,
        metric_name=metric_name,
        metric_value=str(metric_value),
        metric_unit=metric_unit,
        metric_category=metric_category,
        value_type=value_type,
        corner=corner,
        workload=workload,
        confidence=confidence,
        source_note=source_note,
        created_at=now_iso(),
    )


@app.on_event("startup")
def on_startup() -> None:
    DATABASE_DIR.mkdir(parents=True, exist_ok=True)
    create_db_and_tables()
    ensure_sqlite_schema_compatibility()
    if ACTIVE_DATABASE_PATH == DEFAULT_DATABASE_PATH.resolve() and os.getenv("SEED_DEMO", "true").lower() in {"1", "true", "yes", "on"}:
        seed_data()


def database_info(path: Path) -> dict[str, Any]:
    resolved = path.resolve()
    is_active = resolved == ACTIVE_DATABASE_PATH
    project_count = None
    if resolved.exists():
        try:
            temp_engine = create_engine(f"sqlite:///{resolved}", connect_args={"check_same_thread": False})
            with Session(temp_engine) as session:
                SQLModel.metadata.create_all(temp_engine)
                project_count = len(session.exec(select(Project)).all())
            temp_engine.dispose()
        except Exception:
            project_count = None
    return {
        "id": database_id(resolved),
        "name": database_label(resolved),
        "path": str(resolved),
        "active": is_active,
        "is_demo": resolved == DEFAULT_DATABASE_PATH.resolve(),
        "project_count": project_count,
    }


@app.get("/api/databases")
def get_databases() -> dict[str, Any]:
    return {
        "active_id": database_id(ACTIVE_DATABASE_PATH),
        "databases": [database_info(path) for path in database_paths()],
    }


@app.post("/api/databases")
def create_database(payload: DatabaseCreateInput) -> dict[str, Any]:
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Database name is required.")
    safe_name = "".join(ch.lower() if ch.isalnum() else "_" for ch in name).strip("_")
    while "__" in safe_name:
        safe_name = safe_name.replace("__", "_")
    if not safe_name:
        raise HTTPException(status_code=400, detail="Database name must include letters or numbers.")
    path = (DATABASE_DIR / f"{safe_name}.db").resolve()
    if path.exists():
        raise HTTPException(status_code=409, detail=f"Database already exists: {safe_name}")

    previous_path = ACTIVE_DATABASE_PATH
    switch_database(path, create_if_missing=True)
    create_db_and_tables()
    ensure_sqlite_schema_compatibility()
    if payload.seed_demo:
        seed_data()
    info = database_info(ACTIVE_DATABASE_PATH)
    if not payload.seed_demo:
        # Keep the new empty database active so the UI can immediately import into it.
        pass
    if previous_path != ACTIVE_DATABASE_PATH:
        # The newly created database intentionally remains selected.
        pass
    return {"active_id": database_id(ACTIVE_DATABASE_PATH), "database": info, "databases": [database_info(db_path) for db_path in database_paths()]}


@app.post("/api/databases/select")
def select_database(payload: DatabaseSelectInput) -> dict[str, Any]:
    path = database_path_from_id(payload.id)
    switch_database(path)
    create_db_and_tables()
    ensure_sqlite_schema_compatibility()
    return {"active_id": database_id(ACTIVE_DATABASE_PATH), "database": database_info(ACTIVE_DATABASE_PATH), "databases": [database_info(db_path) for db_path in database_paths()]}


def metrics_for(session: Session, impl_option_id: str, subject_type: str, subject_id: str) -> dict[str, Metric]:
    rows = session.exec(
        select(Metric).where(
            Metric.impl_option_id == impl_option_id,
            Metric.subject_type == subject_type,
            Metric.subject_id == subject_id,
        )
    ).all()
    return {row.metric_name: row for row in rows}


def metric_number(metrics: dict[str, Metric], name: str) -> float:
    return number_or_zero(metrics[name].metric_value) if name in metrics else 0


def normalized_content_share(partition_type: str, value: float | None) -> float:
    if partition_type == "full":
        return 1.0
    return float(value if value is not None else 1.0)


def normalized_resource_category(value: str | None) -> str:
    return value if value in ALLOWED_PARTITION_RESOURCE_CATEGORIES else "block"


def partition_equivalent_instances(partition: PhysicalPartition) -> float:
    return partition.physical_instance_count * normalized_content_share(partition.partition_type, partition.content_share)


def canonical_partition_name(component_name: str, category: str, tier_id: str, partition_type: str, partial_index: int = 0) -> str:
    base_name = f"{component_name}_{category}_{tier_id}"
    return f"{base_name}_P{partial_index}" if partition_type == "partial" else base_name


def component_required_resource_categories(session: Session, component: LogicalComponent, impl_option_id: str) -> set[str]:
    metrics = metrics_for(session, impl_option_id, "logical_component", component.id)
    area_summary = logical_area_summary(session, component, impl_option_id)
    metric_names = {
        "logic": "residual_logic_area" if area_summary["has_children"] else "logic_area",
        "sram": "residual_sram_area" if area_summary["has_children"] else "sram_area",
        "block": "residual_block_area" if area_summary["has_children"] else "block_area",
    }
    categories: set[str] = set()
    for category, metric_name in metric_names.items():
        value = area_summary[metric_name] if area_summary["has_children"] else metric_number(metrics, metric_name)
        if value > 0:
            categories.add(category)
    return categories


def is_global_team(team: str | None) -> bool:
    return not team or team in {"Architecture Team", "All", "All Teams"}


def allowed_component_ids_for_team(session: Session, team: str | None, impl_option_id: str = "S2") -> set[str] | None:
    if is_global_team(team):
        return None

    components = session.exec(select(LogicalComponent).order_by(LogicalComponent.hierarchy_path)).all()
    by_id = {component.id: component for component in components}
    assignments = session.exec(
        select(ResponsibilityAssignment).where(
            ResponsibilityAssignment.impl_option_id == impl_option_id,
            ResponsibilityAssignment.team_name == team,
            ResponsibilityAssignment.can_read == True,
        )
    ).all()
    root_ids = [assignment.logical_component_id for assignment in assignments]
    if not root_ids:
        root_ids = [
            component.id
            for component in components
            if component.owner_team == team
            and (not component.parent_id or by_id.get(component.parent_id, component).owner_team != team)
        ]

    allowed: set[str] = set()
    for root_id in root_ids:
        root = by_id.get(root_id)
        if not root:
            continue
        for component in components:
            if component.id == root.id or component.hierarchy_path.startswith(f"{root.hierarchy_path}/"):
                allowed.add(component.id)
    return allowed


def component_rows_for_team(session: Session, team: str | None, impl_option_id: str = "S2") -> tuple[list[LogicalComponent], set[str] | None]:
    rows = session.exec(select(LogicalComponent).order_by(LogicalComponent.hierarchy_path)).all()
    allowed = allowed_component_ids_for_team(session, team, impl_option_id)
    if allowed is None:
        return rows, None
    return [row for row in rows if row.id in allowed], allowed


def component_id_from_name(session: Session, name: str) -> str:
    base = safe_power_id_part(name)
    component_id = f"B_{base}" if not base.startswith("B_") else base
    index = 2
    while session.get(LogicalComponent, component_id):
        component_id = f"B_{base}_{index}" if not base.startswith("B_") else f"{base}_{index}"
        index += 1
    return component_id


def component_path(session: Session, parent_id: str | None, name: str) -> str:
    if not parent_id:
        return name
    parent = session.get(LogicalComponent, parent_id)
    if not parent:
        raise HTTPException(status_code=400, detail=f"Unknown parent_id: {parent_id}")
    return f"{parent.hierarchy_path}/{name}"


def descendant_component_ids(session: Session, component_id: str) -> set[str]:
    rows = session.exec(select(LogicalComponent)).all()
    children: dict[str, list[str]] = {}
    for row in rows:
        if row.parent_id:
            children.setdefault(row.parent_id, []).append(row.id)
    result: set[str] = set()

    def walk(current_id: str) -> None:
        for child_id in children.get(current_id, []):
            result.add(child_id)
            walk(child_id)

    walk(component_id)
    return result


def update_component_subtree_paths(session: Session, component: LogicalComponent) -> None:
    children = session.exec(select(LogicalComponent).where(LogicalComponent.parent_id == component.id)).all()
    for child in children:
        child.hierarchy_path = f"{component.hierarchy_path}/{child.name}"
        child.updated_at = now_iso()
        session.add(child)
        update_component_subtree_paths(session, child)


def ensure_component_write_scope(session: Session, component_id: str | None, team: str | None, impl_option_id: str) -> None:
    if is_global_team(team) or not component_id:
        return
    allowed = allowed_component_ids_for_team(session, team, impl_option_id)
    if allowed is not None and component_id not in allowed:
        raise HTTPException(status_code=403, detail=f"{component_id} is outside team scope {team}")


def scope_component_items(items: list[dict[str, Any]], allowed: set[str] | None) -> list[dict[str, Any]]:
    if allowed is None:
        return items
    return [{**item, "parent": item["parent"] if item["parent"] in allowed else None} for item in items]


def partition_ids_for_components(session: Session, impl_option_id: str, component_ids: set[str]) -> set[str]:
    rows = session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()
    return {row.id for row in rows if row.logical_component_id in component_ids}


def absolute_logical_instance_count(session: Session, component: LogicalComponent) -> int:
    count = component.logical_instance_count
    curr = component
    while curr.parent_id:
        parent = session.get(LogicalComponent, curr.parent_id)
        if not parent:
            break
        count *= parent.logical_instance_count
        curr = parent
    return count


def partition_ui(session: Session, partition: PhysicalPartition) -> dict[str, Any]:
    logical = session.get(LogicalComponent, partition.logical_component_id)
    metrics = metrics_for(session, partition.impl_option_id, "physical_partition", partition.id)
    logical_count = logical.logical_instance_count if logical else 0
    content_share = normalized_content_share(partition.partition_type, partition.content_share)
    return {
        "id": partition.id,
        "impl_option_id": partition.impl_option_id,
        "logical_component_id": partition.logical_component_id,
        "logical_component_name": logical.name if logical else partition.logical_component_id,
        "tier_id": partition.tier_id,
        "partition_name": partition.partition_name,
        "partition_type": partition.partition_type,
        "resource_category": normalized_resource_category(partition.resource_category),
        "physical_instance_count": partition.physical_instance_count,
        "content_share": content_share,
        "instance_share": round(partition.physical_instance_count / logical_count, 4) if logical_count else 0,
        "partition_ratio": content_share,
        "logic_area": metric_number(metrics, "logic_area"),
        "sram_area": metric_number(metrics, "sram_area"),
        "block_area": metric_number(metrics, "block_area"),
        "power": metric_number(metrics, "power"),
        "shape_type": metrics["shape_type"].metric_value if "shape_type" in metrics else "",
        "description": partition.description,
    }


def logical_area_summary(session: Session, component: LogicalComponent, impl_option_id: str) -> dict[str, Any]:
    metrics = metrics_for(session, impl_option_id, "logical_component", component.id)
    total = {
        "logic_area": metric_number(metrics, "logic_area"),
        "sram_area": metric_number(metrics, "sram_area"),
        "block_area": metric_number(metrics, "block_area"),
    }
    child_rows = session.exec(select(LogicalComponent).where(LogicalComponent.parent_id == component.id)).all()
    child_sum = {"logic_area": 0.0, "sram_area": 0.0, "block_area": 0.0}
    for child in child_rows:
        child_metrics = metrics_for(session, impl_option_id, "logical_component", child.id)
        for metric_name in child_sum:
            child_sum[metric_name] += metric_number(child_metrics, metric_name)
    residual = {metric_name: round(total[metric_name] - child_sum[metric_name], 4) for metric_name in total}
    return {
        "has_children": bool(child_rows),
        "child_logic_area": round(child_sum["logic_area"], 4),
        "child_sram_area": round(child_sum["sram_area"], 4),
        "child_block_area": round(child_sum["block_area"], 4),
        "residual_logic_area": residual["logic_area"],
        "residual_sram_area": residual["sram_area"],
        "residual_block_area": residual["block_area"],
    }


def descendant_component_ids(session: Session, component_id: str) -> set[str]:
    rows = session.exec(select(LogicalComponent)).all()
    children_by_parent: dict[str, list[str]] = {}
    for row in rows:
        if row.parent_id:
            children_by_parent.setdefault(row.parent_id, []).append(row.id)
    ids = {component_id}
    stack = [component_id]
    while stack:
        parent_id = stack.pop()
        child_ids = children_by_parent.get(parent_id, [])
        ids.update(child_ids)
        stack.extend(child_ids)
    return ids


def process_scale_for_category(process: ProcessNode | None, category: str) -> float:
    if not process:
        return 1
    if category == "logic":
        return process.logic_area_scale
    if category == "sram":
        return process.sram_area_scale
    return process.block_area_scale


def partition_base_area_for_category(partition_row: dict[str, Any]) -> float:
    category = partition_row["resource_category"]
    if category == "logic":
        return partition_row["logic_area"]
    if category == "sram":
        return partition_row["sram_area"]
    if category == "block":
        return partition_row["block_area"]
    return partition_row["logic_area"] + partition_row["sram_area"] + partition_row["block_area"]


def component_tier_area_distribution(session: Session, component: LogicalComponent, impl_option_id: str) -> list[dict[str, Any]]:
    component_ids = descendant_component_ids(session, component.id)
    partitions = session.exec(
        select(PhysicalPartition).where(
            PhysicalPartition.impl_option_id == impl_option_id,
            PhysicalPartition.logical_component_id.in_(component_ids),
        )
    ).all()
    tiers = {tier.id: tier for tier in session.exec(select(Tier).where(Tier.impl_option_id == impl_option_id)).all()}
    processes = {process.id: process for process in session.exec(select(ProcessNode)).all()}
    rows_by_tier: dict[str, dict[str, Any]] = {}

    for partition in partitions:
        partition_row = partition_ui(session, partition)
        tier = tiers.get(partition.tier_id)
        process = processes.get(tier.process_id) if tier else None
        category = partition_row["resource_category"]
        scale = process_scale_for_category(process, category)
        base_area = partition_base_area_for_category(partition_row)
        scaled_area = base_area * scale
        row = rows_by_tier.setdefault(
            partition.tier_id,
            {
                "tier_id": partition.tier_id,
                "tier_name": tier.name if tier else partition.tier_id,
                "process_id": tier.process_id if tier else "",
                "process": f"{process.foundry} {process.node_name}" if process else "",
                "base_logic_area": 0.0,
                "base_sram_area": 0.0,
                "base_block_area": 0.0,
                "base_total_area": 0.0,
                "logic_area": 0.0,
                "sram_area": 0.0,
                "block_area": 0.0,
                "total_area": 0.0,
                "partition_count": 0,
            },
        )
        row[f"base_{category}_area"] += base_area
        row["base_total_area"] += base_area
        row[f"{category}_area"] += scaled_area
        row["total_area"] += scaled_area
        row["partition_count"] += 1

    tier_order = {tier.id: tier.tier_index for tier in tiers.values()}
    return [
        {
            **row,
            "base_logic_area": round(row["base_logic_area"], 4),
            "base_sram_area": round(row["base_sram_area"], 4),
            "base_block_area": round(row["base_block_area"], 4),
            "base_total_area": round(row["base_total_area"], 4),
            "logic_area": round(row["logic_area"], 4),
            "sram_area": round(row["sram_area"], 4),
            "block_area": round(row["block_area"], 4),
            "total_area": round(row["total_area"], 4),
        }
        for row in sorted(rows_by_tier.values(), key=lambda item: tier_order.get(item["tier_id"], 999))
    ]


def component_ui(session: Session, component: LogicalComponent, impl_option_id: str = "S2") -> dict[str, Any]:
    metrics = metrics_for(session, impl_option_id, "logical_component", component.id)
    partitions = session.exec(
        select(PhysicalPartition).where(
            PhysicalPartition.impl_option_id == impl_option_id,
            PhysicalPartition.logical_component_id == component.id,
        )
    ).all()
    tier_ids = sorted({partition.tier_id for partition in partitions})
    confidence_order = {"approved": 0, "review": 1, "draft": 2}
    confidence = min((metric.confidence for metric in metrics.values()), key=lambda item: confidence_order.get(item, 9), default="draft")
    partition_rows = [partition_ui(session, partition) for partition in partitions]
    equivalent_by_category = {
        category: round(sum(row["physical_instance_count"] * row["content_share"] for row in partition_rows if row["resource_category"] == category), 4)
        for category in sorted(ALLOWED_PARTITION_RESOURCE_CATEGORIES)
    }
    physical_instance_count = max(equivalent_by_category.values(), default=0)
    abs_logical_count = absolute_logical_instance_count(session, component)
    instance_share = round(physical_instance_count / component.logical_instance_count, 4) if component.logical_instance_count else 0
    block_area = metric_number(metrics, "block_area")
    if not block_area:
        block_area = sum(row["logic_area"] + row["sram_area"] + row["block_area"] for row in partition_rows)
    area_summary = logical_area_summary(session, component, impl_option_id)
    
    # compute own_mapping_closed
    own_closed = True
    self_area = {
        "logic": area_summary["residual_logic_area"] if area_summary["has_children"] else metric_number(metrics, "logic_area"),
        "sram": area_summary["residual_sram_area"] if area_summary["has_children"] else metric_number(metrics, "sram_area"),
        "block": area_summary["residual_block_area"] if area_summary["has_children"] else metric_number(metrics, "block_area"),
    }
    
    for category in ALLOWED_PARTITION_RESOURCE_CATEGORIES:
        category_partitions = [p for p in partitions if normalized_resource_category(p.resource_category) == category]
        expected_area = self_area[category]
        
        if len(category_partitions) == 0:
            if expected_area > 0.01:
                own_closed = False
            continue
            
        equiv = sum(partition_equivalent_instances(p) for p in category_partitions)
        mapped_area = sum(
            metric_number(metrics_for(session, impl_option_id, "physical_partition", p.id), f"{category}_area")
            for p in category_partitions
        )
        
        if abs(equiv - component.logical_instance_count) > 0.001 or abs(mapped_area - expected_area) > 0.01:
            own_closed = False

    # compute subtree_mapping_closed recursively
    child_rows = session.exec(select(LogicalComponent).where(LogicalComponent.parent_id == component.id)).all()
    if not own_closed:
        subtree_closed = False
    else:
        def check_descendant_closed(c: LogicalComponent) -> bool:
            c_metrics = metrics_for(session, impl_option_id, "logical_component", c.id)
            c_partitions = session.exec(
                select(PhysicalPartition).where(
                    PhysicalPartition.impl_option_id == impl_option_id,
                    PhysicalPartition.logical_component_id == c.id,
                )
            ).all()
            c_area_summary = logical_area_summary(session, c, impl_option_id)
            c_self_area = {
                "logic": c_area_summary["residual_logic_area"] if c_area_summary["has_children"] else metric_number(c_metrics, "logic_area"),
                "sram": c_area_summary["residual_sram_area"] if c_area_summary["has_children"] else metric_number(c_metrics, "sram_area"),
                "block": c_area_summary["residual_block_area"] if c_area_summary["has_children"] else metric_number(c_metrics, "block_area"),
            }
            
            for cat in ALLOWED_PARTITION_RESOURCE_CATEGORIES:
                cat_parts = [p for p in c_partitions if normalized_resource_category(p.resource_category) == cat]
                exp_area = c_self_area[cat]
                if len(cat_parts) == 0:
                    if exp_area > 0.01:
                        return False
                    continue
                eq = sum(partition_equivalent_instances(p) for p in cat_parts)
                ma = sum(
                    metric_number(metrics_for(session, impl_option_id, "physical_partition", p.id), f"{cat}_area")
                    for p in cat_parts
                )
                if abs(eq - c.logical_instance_count) > 0.001 or abs(ma - exp_area) > 0.01:
                    return False
            
            c_children = session.exec(select(LogicalComponent).where(LogicalComponent.parent_id == c.id)).all()
            for child in c_children:
                if not check_descendant_closed(child):
                    return False
            return True

        subtree_closed = True
        for child in child_rows:
            if not check_descendant_closed(child):
                subtree_closed = False
                break

    return {
        "id": component.id,
        "project_id": component.project_id,
        "parent": component.parent_id,
        "name": component.name,
        "type": component.instance_type,
        "domain": component.function_domain,
        "resource": component.resource_type,
        "hierarchy_path": component.hierarchy_path,
        "logical_instance_count": component.logical_instance_count,
        "absolute_logical_instance_count": abs_logical_count,
        "owner_team": component.owner_team,
        "visibility_level": component.visibility_level,
        "physical_instance_count": physical_instance_count,
        "equivalent_instances_by_category": equivalent_by_category,
        "instance_share": instance_share,
        "partition_ratio": instance_share,
        "signal_count_total": metric_number(metrics, "signal_count_total"),
        "logic_area": metric_number(metrics, "logic_area"),
        "sram_area": metric_number(metrics, "sram_area"),
        "block_area": block_area,
        "area": block_area,
        "power": metric_number(metrics, "power") + sum(row["power"] for row in partition_rows),
        "tier": "/".join(tier_ids) if tier_ids else "-",
        "confidence": confidence,
        "partitions": partition_rows,
        "tier_area_distribution": component_tier_area_distribution(session, component, impl_option_id),
        "description": component.description,
        "own_mapping_closed": own_closed,
        "subtree_mapping_closed": subtree_closed,
        **area_summary,
    }


def build_component_tree(items: list[dict[str, Any]], parent: str | None = None) -> list[dict[str, Any]]:
    return [{**item, "children": build_component_tree(items, item["id"])} for item in items if item["parent"] == parent]


def impl_option_ui(session: Session, impl_option: ImplOption) -> dict[str, Any]:
    metrics = metrics_for(session, impl_option.id, "impl_option", impl_option.id)
    area = metric_number(metrics, "area")
    power = metric_number(metrics, "power")
    return {
        "id": impl_option.id,
        "project_id": impl_option.project_id,
        "name": impl_option.name,
        "process": impl_option.process_combo,
        "process_combo": impl_option.process_combo,
        "die": impl_option.impl_type,
        "impl_type": impl_option.impl_type,
        "area": area,
        "power": power,
        "risk": impl_option.status,
        "cost": "High" if impl_option.id == "S2" else "Medium",
        "thermal": "High" if impl_option.id == "S2" else "Medium",
        "description": impl_option.description,
        "status": impl_option.status,
        "created_at": impl_option.created_at,
        "updated_at": impl_option.updated_at,
    }


def impl_option_detail_ui(session: Session, impl_option_id: str) -> dict[str, Any]:
    implementation = session.get(ImplOptionDetail, impl_option_id)
    tiers = session.exec(select(ImplementationTier).where(ImplementationTier.impl_option_id == impl_option_id).order_by(ImplementationTier.tier_index)).all()
    interfaces = session.exec(select(ImplementationInterface).where(ImplementationInterface.impl_option_id == impl_option_id)).all()
    package_escape = session.get(ImplementationPackageEscape, impl_option_id)
    return {
        "exists": implementation is not None,
        "impl_option_id": impl_option_id,
        "implementation_type": implementation.implementation_type if implementation else "",
        "status": implementation.status if implementation else "draft",
        "version": implementation.version if implementation else 0,
        "updated_at": implementation.updated_at if implementation else "",
        "tiers": [
            {
                "id": tier.tier_id,
                "name": tier.name,
                "process": tier.process,
                "role": tier.role,
                "thickness_um": tier.thickness_um,
            }
            for tier in tiers
        ] if tiers else [
            {
                "id": tier.id,
                "name": tier.name,
                "process": tier.process_id,
                "role": tier.role,
                "thickness_um": tier.thickness_um,
            }
            for tier in session.exec(select(Tier).where(Tier.impl_option_id == impl_option_id).order_by(Tier.tier_index)).all()
        ],
        "interfaces": [
            {
                "id": row.id.removeprefix(f"{impl_option_id}:"),
                "from_tier_id": row.from_tier_id,
                "to_tier_id": row.to_tier_id,
                "orientation": row.orientation,
                "interconnect": row.interconnect,
                "hb_pitch_um": row.hb_pitch_um,
                "upper_tsv_pitch_um": row.upper_tsv_pitch_um,
                "upper_tsv_keepout_um": row.upper_tsv_keepout_um,
                "lower_tsv_pitch_um": row.lower_tsv_pitch_um,
                "lower_tsv_keepout_um": row.lower_tsv_keepout_um,
                "description": row.description,
            }
            for row in interfaces
        ],
        "package_escape": {
            "bottom_tier_id": package_escape.bottom_tier_id if package_escape else "",
            "requires_tsv": package_escape.requires_tsv if package_escape else False,
            "pitch_um": package_escape.pitch_um if package_escape else 0,
            "keepout_um": package_escape.keepout_um if package_escape else 0,
            "description": package_escape.description if package_escape else "",
        },
    }


def impl_option_detail_impact_errors(session: Session, impl_option_id: str, payload: ImplOptionDetailUpdate) -> list[str]:
    errors: list[str] = []
    new_tier_ids = [tier.id for tier in payload.tiers]
    if not new_tier_ids:
        errors.append("At least one implementation tier is required.")
    if len(new_tier_ids) != len(set(new_tier_ids)):
        errors.append("Tier ids must be unique within an implementation.")

    partition_rows = session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()
    partition_usage: dict[str, int] = {}
    for row in partition_rows:
        partition_usage[row.tier_id] = partition_usage.get(row.tier_id, 0) + 1

    new_tier_set = set(new_tier_ids)
    for tier_id in sorted(tier_id for tier_id in partition_usage if tier_id not in new_tier_set):
        errors.append(f"Tier {tier_id} is used by {partition_usage[tier_id]} physical partitions and cannot be removed or renamed.")

    existing_tiers = session.exec(select(ImplementationTier).where(ImplementationTier.impl_option_id == impl_option_id)).all()
    existing_index = {row.tier_id: row.tier_index for row in existing_tiers}
    for index, tier_id in enumerate(new_tier_ids):
        if tier_id in partition_usage and tier_id in existing_index and existing_index[tier_id] != index:
            errors.append(f"Tier {tier_id} is used by {partition_usage[tier_id]} physical partitions and cannot be reordered.")

    for row in payload.interfaces:
        if row.from_tier_id not in new_tier_set or row.to_tier_id not in new_tier_set:
            errors.append(f"Interface {row.id} references tiers outside this implementation.")
    if payload.package_escape.bottom_tier_id and payload.package_escape.bottom_tier_id not in new_tier_set:
        errors.append(f"Package escape bottom_tier_id {payload.package_escape.bottom_tier_id} is not in this implementation.")
    return errors


def make_quality_issue(
    severity: str,
    title: str,
    detail: str,
    action: str,
    entity_type: str,
    entity_id: str,
) -> dict[str, str]:
    return {
        "id": f"{entity_type}:{entity_id}:{title}".replace(" ", "_").lower(),
        "severity": severity,
        "title": title,
        "detail": detail,
        "action": action,
        "entity_type": entity_type,
        "entity_id": entity_id,
    }


def quality_issues_for(session: Session, impl_option_id: str = "S2", team: str | None = None) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    components, allowed_component_ids = component_rows_for_team(session, team, impl_option_id)
    partitions = session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()
    metrics = session.exec(select(Metric).where(Metric.impl_option_id == impl_option_id)).all()
    if allowed_component_ids is not None:
        partitions = [row for row in partitions if row.logical_component_id in allowed_component_ids]
        allowed_partition_ids = {row.id for row in partitions}
        metrics = [
            row
            for row in metrics
            if (row.subject_type == "logical_component" and row.subject_id in allowed_component_ids)
            or (row.subject_type == "physical_partition" and row.subject_id in allowed_partition_ids)
        ]
    all_components = session.exec(select(LogicalComponent)).all()
    by_id = {c.id: c for c in all_components}
    abs_counts: dict[str, int] = {}
    def get_abs_count(cid: str) -> int:
        if cid in abs_counts:
            return abs_counts[cid]
        c = by_id.get(cid)
        if not c:
            return 1
        if not c.parent_id:
            abs_counts[cid] = c.logical_instance_count
        else:
            abs_counts[cid] = c.logical_instance_count * get_abs_count(c.parent_id)
        return abs_counts[cid]

    partitions_by_component: dict[str, list[PhysicalPartition]] = {}
    metrics_by_subject: dict[tuple[str, str], dict[str, Metric]] = {}
    children_by_parent: dict[str, list[LogicalComponent]] = {}

    for partition in partitions:
        partitions_by_component.setdefault(partition.logical_component_id, []).append(partition)
    for row in metrics:
        metrics_by_subject.setdefault((row.subject_type, row.subject_id), {})[row.metric_name] = row
    for component in components:
        if component.parent_id:
            children_by_parent.setdefault(component.parent_id, []).append(component)

    def self_area_by_category(component: LogicalComponent) -> dict[str, float]:
        available = metrics_by_subject.get(("logical_component", component.id), {})
        child_rows = children_by_parent.get(component.id, [])
        return {
            "logic": max(
                0.0,
                metric_number(available, "logic_area")
                - sum(metric_number(metrics_by_subject.get(("logical_component", child.id), {}), "logic_area") for child in child_rows),
            ),
            "sram": max(
                0.0,
                metric_number(available, "sram_area")
                - sum(metric_number(metrics_by_subject.get(("logical_component", child.id), {}), "sram_area") for child in child_rows),
            ),
            "block": max(
                0.0,
                metric_number(available, "block_area")
                - sum(metric_number(metrics_by_subject.get(("logical_component", child.id), {}), "block_area") for child in child_rows),
            ),
        }

    def partition_area_by_category(partition: PhysicalPartition, category: str) -> float:
        partition_metrics = metrics_by_subject.get(("physical_partition", partition.id), {})
        return metric_number(partition_metrics, f"{category}_area")

    own_mapping_closed: dict[str, bool] = {}
    area_epsilon = 0.01

    for component in components:
        component_partitions = partitions_by_component.get(component.id, [])
        self_area = self_area_by_category(component)
        component_closed = True

        for category in sorted(ALLOWED_PARTITION_RESOURCE_CATEGORIES):
            category_partitions = [partition for partition in component_partitions if normalized_resource_category(partition.resource_category) == category]
            expected_area = self_area[category]
            if len(category_partitions) == 0:
                if expected_area > area_epsilon:
                    component_closed = False
                    issues.append(
                        make_quality_issue(
                            "High",
                            f"{category.upper()} implementation coverage not closed",
                            f"{component.name} self/residual {category} maps to 0.000 equivalent instances, expected {component.logical_instance_count}.",
                            "Add physical partition rows for this resource category so count * content_share closes to the logical instance count.",
                            "logical_component",
                            component.id,
                        )
                    )
                continue
                
            equivalent_instances = sum(partition_equivalent_instances(partition) for partition in category_partitions)
            mapped_area = sum(partition_area_by_category(partition, category) for partition in category_partitions)

            if abs(equivalent_instances - component.logical_instance_count) > 0.001:
                component_closed = False
                issues.append(
                    make_quality_issue(
                        "High",
                        f"{category.upper()} implementation coverage not closed",
                        f"{component.name} self/residual {category} maps to {equivalent_instances:.3f} equivalent instances, expected {component.logical_instance_count}.",
                        "Adjust physical_instance_count and content_share for this resource category so count * content_share closes to the logical instance count.",
                        "logical_component",
                        component.id,
                    )
                )
            if abs(mapped_area - expected_area) > area_epsilon:
                component_closed = False
                issues.append(
                    make_quality_issue(
                        "High",
                        f"{category.upper()} area mapping not closed",
                        f"{component.name} self/residual {category} area maps to {mapped_area:.3f} mm2, expected {expected_area:.3f} mm2.",
                        "Adjust physical partition metrics for this resource category so direct partition base area equals the component self/residual area.",
                        "logical_component",
                        component.id,
                    )
                )

        own_mapping_closed[component.id] = component_closed

        for partition in component_partitions:
            if partition.resource_category not in ALLOWED_PARTITION_RESOURCE_CATEGORIES:
                issues.append(
                    make_quality_issue(
                        "Medium",
                        "Unsupported partition resource category",
                        f"{partition.id} uses resource_category={partition.resource_category}.",
                        "Use logic, sram, or block.",
                        "physical_partition",
                        partition.id,
                    )
                )
            if partition.partition_type == "full" and abs(partition.content_share - 1.0) > 0.001:
                issues.append(
                    make_quality_issue(
                        "Medium",
                        "Full partition content_share must be 1",
                        f"{partition.id} is full but content_share={partition.content_share}.",
                        "Set full partitions to content_share=1 or change the partition_type to partial.",
                        "physical_partition",
                        partition.id,
                    )
                )

    subtree_mapping_closed: dict[str, bool] = {}

    def is_subtree_mapping_closed(component: LogicalComponent) -> bool:
        if component.id in subtree_mapping_closed:
            return subtree_mapping_closed[component.id]
        child_rows = children_by_parent.get(component.id, [])
        child_status = all(is_subtree_mapping_closed(child) for child in child_rows)
        subtree_mapping_closed[component.id] = own_mapping_closed.get(component.id, False) and child_status
        return subtree_mapping_closed[component.id]

    for component in components:
        child_rows = children_by_parent.get(component.id, [])
        if not child_rows:
            continue
        open_children = [child.name for child in child_rows if not is_subtree_mapping_closed(child)]
        if open_children:
            issues.append(
                make_quality_issue(
                    "High",
                    "Subtree implementation mapping not closed",
                    f"{component.name} is not fully mapped because child subtree(s) are open: {', '.join(open_children)}.",
                    "Close every child subtree and this component's own residual/self mapping for each non-zero resource category.",
                    "logical_component",
                    component.id,
                )
            )

    required_logical_metrics = {"signal_count_total", "logic_area", "sram_area", "block_area"}
    parent_ids = {row.parent_id for row in components if row.parent_id}
    leaf_components = [row for row in components if row.id not in parent_ids]

    for component in components:
        child_rows = children_by_parent.get(component.id, [])
        if not child_rows:
            continue
        available = metrics_by_subject.get(("logical_component", component.id), {})
        missing_area = sorted({"logic_area", "sram_area", "block_area"} - set(available))
        if missing_area:
            issues.append(
                make_quality_issue(
                    "Medium",
                    "Parent total area metrics missing",
                    f"{component.name} needs total area metrics to compute residual area: {', '.join(missing_area)}.",
                    "Fill parent total logic_area, sram_area, and block_area; residual area is derived automatically.",
                    "logical_component",
                    component.id,
                )
            )
            continue
        for metric_name in ["logic_area", "sram_area", "block_area"]:
            parent_value = metric_number(available, metric_name)
            child_value = sum(metric_number(metrics_by_subject.get(("logical_component", child.id), {}), metric_name) for child in child_rows)
            if parent_value + 0.001 < child_value:
                issues.append(
                    make_quality_issue(
                        "High",
                        "Parent area smaller than child sum",
                        f"{component.name} {metric_name}={parent_value:.3f}, but direct children sum to {child_value:.3f}.",
                        "Parent total area should include child modules; residual area is computed as parent total minus direct child total.",
                        "logical_component",
                        component.id,
                    )
                )

    for component in leaf_components:
        available = metrics_by_subject.get(("logical_component", component.id), {})
        missing = sorted(required_logical_metrics - set(available))
        if missing:
            issues.append(
                make_quality_issue(
                    "Medium",
                    "Logical metrics missing",
                    f"{component.name} is missing logical metrics: {', '.join(missing)}.",
                    "Add the missing metric rows with subject_type=logical_component.",
                    "logical_component",
                    component.id,
                )
            )

    valid_subject_ids = {
        "logical_component": {row.id for row in components},
        "physical_partition": {row.id for row in partitions},
        "tier": {row.id for row in session.exec(select(Tier).where(Tier.impl_option_id == impl_option_id)).all()},
        "impl_option": {impl_option_id},
    }
    for row in metrics:
        if row.subject_id not in valid_subject_ids.get(row.subject_type, set()):
            issues.append(
                make_quality_issue(
                    "High",
                    "Metric subject missing",
                    f"{row.id} references missing {row.subject_type} subject_id={row.subject_id}.",
                    "Fix subject_type / subject_id or import the referenced entity first.",
                    "metric",
                    row.id,
                )
            )
        if row.value_type == "number":
            try:
                float(row.metric_value)
            except ValueError:
                issues.append(
                    make_quality_issue(
                        "High",
                        "Metric value is not numeric",
                        f"{row.id} declares value_type=number but metric_value={row.metric_value!r}.",
                        "Replace metric_value with a numeric value or change value_type.",
                        "metric",
                        row.id,
                    )
                )

    # Check tier area limits after process scaling
    tiers_in_impl_option = session.exec(select(Tier).where(Tier.impl_option_id == impl_option_id)).all()
    processes = {p.id: p for p in session.exec(select(ProcessNode)).all()}
    for tier in tiers_in_impl_option:
        process = processes.get(tier.process_id)
        tier_partitions = [p for p in partitions if p.tier_id == tier.id]
        total_scaled_area = 0.0
        for partition in tier_partitions:
            partition_metrics = metrics_by_subject.get(("physical_partition", partition.id), {})
            p_logic = metric_number(partition_metrics, "logic_area")
            p_sram = metric_number(partition_metrics, "sram_area")
            p_block = metric_number(partition_metrics, "block_area")
            
            scaled_logic = p_logic * process_scale_for_category(process, "logic")
            scaled_sram = p_sram * process_scale_for_category(process, "sram")
            scaled_block = p_block * process_scale_for_category(process, "block")
            
            total_scaled_area += (scaled_logic + scaled_sram + scaled_block)
            
        if tier.area_limit_mm2 > 0 and total_scaled_area > tier.area_limit_mm2:
            issues.append(
                make_quality_issue(
                    "Medium",
                    "Tier physical area limit exceeded",
                    f"Tier {tier.id} ({tier.name}) computed area {total_scaled_area:.3f} mm² (after process scaling) exceeds its limit of {tier.area_limit_mm2:.3f} mm².",
                    "Optimize partition mappings, move blocks to other tiers, or use a more advanced process node with better area scaling.",
                    "tier",
                    tier.id,
                )
            )

    return issues


@app.get("/api/application-scenarios")
def get_application_scenarios() -> list[ApplicationScenario]:
    with Session(engine) as session:
        return list(session.exec(select(ApplicationScenario).where(ApplicationScenario.id != "AS_MODULE_LIBRARY")).all())


@app.post("/api/application-scenarios")
def create_application_scenario(payload: ApplicationScenarioInput) -> ApplicationScenario:
    with Session(engine) as session:
        if not session.get(Project, payload.project_id):
            raise HTTPException(status_code=400, detail=f"Unknown project_id: {payload.project_id}")
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Application scenario name is required.")
        base_id = f"AS_{safe_power_id_part(name)}"
        scenario_id = base_id
        index = 2
        while session.get(ApplicationScenario, scenario_id):
            scenario_id = f"{base_id}_{index}"
            index += 1
        row = ApplicationScenario(
            id=scenario_id,
            project_id=payload.project_id,
            name=name,
            category=(payload.category or "Custom").strip() or "Custom",
            description=payload.description or "",
        )
        session.add(row)
        session.commit()
        session.refresh(row)
        return row


@app.put("/api/application-scenarios/{scenario_id}")
def update_application_scenario(scenario_id: str, payload: ApplicationScenarioInput) -> ApplicationScenario:
    if scenario_id == "AS_MODULE_LIBRARY":
        raise HTTPException(status_code=400, detail="Internal module library scenario cannot be edited.")
    with Session(engine) as session:
        row = session.get(ApplicationScenario, scenario_id)
        if not row:
            raise HTTPException(status_code=404, detail=f"Application scenario not found: {scenario_id}")
        if not session.get(Project, payload.project_id):
            raise HTTPException(status_code=400, detail=f"Unknown project_id: {payload.project_id}")
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Application scenario name is required.")
        row.project_id = payload.project_id
        row.name = name
        row.category = (payload.category or "Custom").strip() or "Custom"
        row.description = payload.description or ""
        session.add(row)
        session.commit()
        session.refresh(row)
        return row


@app.delete("/api/application-scenarios/{scenario_id}")
def delete_application_scenario(scenario_id: str) -> dict[str, Any]:
    if scenario_id == "AS_MODULE_LIBRARY":
        raise HTTPException(status_code=400, detail="Internal module library scenario cannot be deleted.")
    with Session(engine) as session:
        row = session.get(ApplicationScenario, scenario_id)
        if not row:
            raise HTTPException(status_code=404, detail=f"Application scenario not found: {scenario_id}")
        selections = list(
            session.exec(
                select(ApplicationScenarioSelection).where(
                    ApplicationScenarioSelection.application_scenario_id == scenario_id,
                )
            ).all()
        )
        observations = list(
            session.exec(
                select(PowerObservation).where(
                    PowerObservation.application_scenario_id == scenario_id,
                )
            ).all()
        )
        for selection in selections:
            session.delete(selection)
        for observation in observations:
            session.delete(observation)
        session.delete(row)
        session.commit()
        return {
            "success": True,
            "deleted_id": scenario_id,
            "deleted_selection_count": len(selections),
            "deleted_observation_count": len(observations),
        }


@app.get("/api/physical-mappings")
def get_physical_mappings(impl_option_id: str | None = None) -> list[PhysicalMapping]:
    with Session(engine) as session:
        stmt = select(PhysicalMapping)
        if impl_option_id:
            stmt = stmt.where(PhysicalMapping.impl_option_id == impl_option_id)
        return list(session.exec(stmt).all())


@app.get("/api/operating-point-sets")
def get_operating_point_sets() -> list[OperatingPointSet]:
    with Session(engine) as session:
        return list(session.exec(select(OperatingPointSet)).all())


def safe_power_id_part(value: str | None) -> str:
    return (value or "Default").replace(" ", "_").replace("/", "_").replace("-", "_").replace(".", "_").upper()


def profile_id_from_name(name: str | None) -> str:
    cleaned = (name or "Default").strip() or "Default"
    if cleaned.lower() == "default":
        return "OP_DEFAULT"
    return f"OP_{safe_power_id_part(cleaned)}"


def module_power_observation_id(impl_option_id: str, physical_mapping_id: str, component_id: str, use_case_name: str, operating_point_set_id: str) -> str:
    return (
        f"PUC_{safe_power_id_part(impl_option_id)}_{safe_power_id_part(physical_mapping_id)}_"
        f"{safe_power_id_part(component_id)}_{safe_power_id_part(use_case_name)}_{safe_power_id_part(operating_point_set_id)}"
    )


def app_selection_id(application_scenario_id: str, component_id: str, use_case_name: str, operating_point_set_id: str) -> str:
    return (
        f"ASC_{safe_power_id_part(application_scenario_id)}_{safe_power_id_part(component_id)}_"
        f"{safe_power_id_part(use_case_name)}_{safe_power_id_part(operating_point_set_id)}"
    )


def validate_power_context(
    session: Session,
    project_id: str,
    impl_option_id: str,
    physical_mapping_id: str,
    application_scenario_id: str | None = None,
    operating_point_set_id: str | None = None,
    component_id: str | None = None,
) -> None:
    if not session.get(Project, project_id):
        raise HTTPException(status_code=400, detail=f"Unknown project_id: {project_id}")
    if not session.get(ImplOption, impl_option_id):
        raise HTTPException(status_code=400, detail=f"Unknown impl_option_id: {impl_option_id}")
    mapping = session.get(PhysicalMapping, physical_mapping_id)
    if not mapping:
        raise HTTPException(status_code=400, detail=f"Unknown physical_mapping_id: {physical_mapping_id}")
    if mapping.impl_option_id != impl_option_id:
        raise HTTPException(status_code=400, detail=f"physical_mapping_id {physical_mapping_id} does not belong to impl_option_id {impl_option_id}")
    if application_scenario_id and not session.get(ApplicationScenario, application_scenario_id):
        raise HTTPException(status_code=400, detail=f"Unknown application_scenario_id: {application_scenario_id}")
    if operating_point_set_id and not session.get(OperatingPointSet, operating_point_set_id):
        raise HTTPException(status_code=400, detail=f"Unknown operating_point_set_id: {operating_point_set_id}")
    if component_id and not session.get(LogicalComponent, component_id):
        raise HTTPException(status_code=400, detail=f"Unknown component_id: {component_id}")


def module_power_rows(session: Session, impl_option_id: str, physical_mapping_id: str) -> list[PowerObservation]:
    return list(
        session.exec(
            select(PowerObservation).where(
                PowerObservation.impl_option_id == impl_option_id,
                PowerObservation.physical_mapping_id == physical_mapping_id,
                PowerObservation.application_scenario_id == "AS_MODULE_LIBRARY",
                PowerObservation.scope_type == "component",
            )
        ).all()
    )


@app.get("/api/module-power-usecases")
def get_module_power_usecases(impl_option_id: str, physical_mapping_id: str) -> list[dict[str, Any]]:
    with Session(engine) as session:
        rows = module_power_rows(session, impl_option_id, physical_mapping_id)
        op_names = {row.id: row.name for row in session.exec(select(OperatingPointSet)).all()}
        return [
            {
                "id": row.id,
                "project_id": row.project_id,
                "impl_option_id": row.impl_option_id,
                "physical_mapping_id": row.physical_mapping_id,
                "component_id": row.scope_id,
                "component_name": row.scope_name,
                "use_case_name": row.use_case_name or "Default",
                "operating_point_set_id": row.operating_point_set_id,
                "operating_point_set_name": op_names.get(row.operating_point_set_id, row.operating_point_set_id),
                "power_value_w": row.power_value_w,
                "confidence": row.confidence,
                "note": row.note,
            }
            for row in rows
        ]


@app.post("/api/module-power-usecases")
def upsert_module_power_usecase(payload: ModulePowerUseCaseInput) -> dict[str, Any]:
    with Session(engine) as session:
        validate_power_context(
            session,
            payload.project_id,
            payload.impl_option_id,
            payload.physical_mapping_id,
            component_id=payload.component_id,
        )
        if payload.power_value_w < 0:
            raise HTTPException(status_code=400, detail="power_value_w must be >= 0")
        use_case_name = payload.use_case_name.strip() or "Default"
        op_name = (payload.operating_point_set_name or "").strip()
        op_id = (payload.operating_point_set_id or "").strip() or profile_id_from_name(op_name)
        if not op_id:
            op_id = "OP_DEFAULT"
        if op_id == "OP_DEFAULT":
            op_name = "Default"
        elif not op_name:
            op_name = op_id.removeprefix("OP_").replace("_", " ").title()
        if not session.get(OperatingPointSet, op_id):
            session.add(
                OperatingPointSet(
                    id=op_id,
                    project_id=payload.project_id,
                    name=op_name,
                    description=f"Module Profile created from the application power editor for {payload.component_name}.",
                    op_json="{}",
                )
            )
        row = PowerObservation(
            id=module_power_observation_id(payload.impl_option_id, payload.physical_mapping_id, payload.component_id, use_case_name, op_id),
            project_id=payload.project_id,
            impl_option_id=payload.impl_option_id,
            physical_mapping_id=payload.physical_mapping_id,
            application_scenario_id="AS_MODULE_LIBRARY",
            operating_point_set_id=op_id,
            scope_type="component",
            scope_id=payload.component_id,
            scope_name=payload.component_name,
            use_case_name=use_case_name,
            time_window_name="steady_state",
            statistic_type="average",
            power_type="total",
            power_value_w=payload.power_value_w,
            development_stage="architecture_estimate",
            source_type="web_ui",
            confidence=payload.confidence or "draft",
            is_additive=True,
            note=payload.note,
        )
        session.merge(row)
        session.commit()
        return get_module_power_usecases(payload.impl_option_id, payload.physical_mapping_id)[0] if False else {
            "id": row.id,
            "project_id": row.project_id,
            "impl_option_id": row.impl_option_id,
            "physical_mapping_id": row.physical_mapping_id,
            "component_id": row.scope_id,
            "component_name": row.scope_name,
            "use_case_name": row.use_case_name or "Default",
            "operating_point_set_id": row.operating_point_set_id,
            "operating_point_set_name": op_name,
            "power_value_w": row.power_value_w,
            "confidence": row.confidence,
            "note": row.note,
        }


@app.delete("/api/module-power-usecases/{usecase_id}")
def delete_module_power_usecase(usecase_id: str) -> dict[str, Any]:
    with Session(engine) as session:
        row = session.get(PowerObservation, usecase_id)
        if not row or row.application_scenario_id != "AS_MODULE_LIBRARY" or row.scope_type != "component":
            raise HTTPException(status_code=404, detail=f"Module power use case not found: {usecase_id}")
        matching_selections = list(
            session.exec(
                select(ApplicationScenarioSelection).where(
                    ApplicationScenarioSelection.impl_option_id == row.impl_option_id,
                    ApplicationScenarioSelection.physical_mapping_id == row.physical_mapping_id,
                    ApplicationScenarioSelection.component_id == row.scope_id,
                    ApplicationScenarioSelection.use_case_name == (row.use_case_name or "Default"),
                    ApplicationScenarioSelection.operating_point_set_id == row.operating_point_set_id,
                )
            ).all()
        )
        for selection in matching_selections:
            session.delete(selection)
        session.delete(row)
        session.commit()
        return {
            "success": True,
            "deleted_id": usecase_id,
            "deleted_selection_count": len(matching_selections),
        }


@app.get("/api/application-scenario-composition")
def get_application_scenario_composition(impl_option_id: str, physical_mapping_id: str, application_scenario_id: str) -> list[ApplicationScenarioSelection]:
    with Session(engine) as session:
        return list(
            session.exec(
                select(ApplicationScenarioSelection).where(
                    ApplicationScenarioSelection.impl_option_id == impl_option_id,
                    ApplicationScenarioSelection.physical_mapping_id == physical_mapping_id,
                    ApplicationScenarioSelection.application_scenario_id == application_scenario_id,
                )
            ).all()
        )


POWER_ROLLUP_ABS_TOLERANCE_W = 0.001
POWER_ROLLUP_REL_TOLERANCE = 0.01


def power_rollup_tolerance(parent_power: float | None, child_sum: float) -> float:
    baseline = max(abs(parent_power or 0.0), abs(child_sum), 0.0)
    return max(POWER_ROLLUP_ABS_TOLERANCE_W, baseline * POWER_ROLLUP_REL_TOLERANCE)


def component_hierarchy_maps(session: Session) -> tuple[dict[str, LogicalComponent], dict[str, str | None], dict[str, list[str]], dict[str, set[str]], dict[str, set[str]]]:
    components = list(session.exec(select(LogicalComponent)).all())
    by_id = {row.id: row for row in components}
    parent_by_id = {row.id: row.parent_id for row in components}
    children_by_parent: dict[str, list[str]] = {}
    for row in components:
        if row.parent_id:
            children_by_parent.setdefault(row.parent_id, []).append(row.id)

    ancestors_by_id: dict[str, set[str]] = {}
    for component_id, parent_id in parent_by_id.items():
        ancestors: set[str] = set()
        current = parent_id
        while current:
            ancestors.add(current)
            current = parent_by_id.get(current)
        ancestors_by_id[component_id] = ancestors

    descendants_by_id: dict[str, set[str]] = {}

    def descendants(component_id: str) -> set[str]:
        if component_id in descendants_by_id:
            return descendants_by_id[component_id]
        result: set[str] = set()
        for child_id in children_by_parent.get(component_id, []):
            result.add(child_id)
            result.update(descendants(child_id))
        descendants_by_id[component_id] = result
        return result

    for component_id in by_id:
        descendants(component_id)

    return by_id, parent_by_id, children_by_parent, ancestors_by_id, descendants_by_id


def selection_power(selection: ApplicationScenarioSelection, library: dict[tuple[str | None, str, str], PowerObservation]) -> float | None:
    observation = library.get((selection.component_id, selection.use_case_name or "Default", selection.operating_point_set_id))
    return observation.power_value_w if observation else None


def validate_active_power_hierarchy(
    selections: list[ApplicationScenarioSelection],
    ancestors_by_id: dict[str, set[str]],
) -> None:
    active_ids = {row.component_id for row in selections if row.included}
    for row in selections:
        if not row.included:
            continue
        active_ancestors = sorted(ancestors_by_id.get(row.component_id, set()) & active_ids)
        if active_ancestors:
            raise HTTPException(
                status_code=400,
                detail=f"Power composition double-count risk: {row.component_name} and ancestor {active_ancestors[0]} are both included.",
            )


def build_power_hierarchy_rollups(
    session: Session,
    selections: list[ApplicationScenarioSelection],
    library: dict[tuple[str | None, str, str], PowerObservation],
) -> list[dict[str, Any]]:
    components_by_id, _parent_by_id, children_by_parent, _ancestors_by_id, descendants_by_id = component_hierarchy_maps(session)
    selection_by_component = {row.component_id: row for row in selections}
    rollups: list[dict[str, Any]] = []
    for parent_id, child_ids in children_by_parent.items():
        parent_selection = selection_by_component.get(parent_id)
        child_selections = [selection_by_component[child_id] for child_id in child_ids if child_id in selection_by_component]
        if not parent_selection and not child_selections:
            continue
        parent_power = selection_power(parent_selection, library) if parent_selection else None
        child_sum = 0.0
        missing_child_count = 0
        assigned_child_count = 0
        for child_selection in child_selections:
            assigned_child_count += 1
            child_power = selection_power(child_selection, library)
            if child_power is None:
                missing_child_count += 1
            else:
                child_sum += child_power

        residual = None
        status = "incomplete"
        if parent_power is None:
            status = "incomplete"
        else:
            residual = round(parent_power - child_sum, 4)
            tolerance = power_rollup_tolerance(parent_power, child_sum)
            if missing_child_count > 0:
                status = "incomplete"
            elif child_sum - parent_power > tolerance:
                status = "over_specified"
            elif abs(parent_power - child_sum) <= tolerance:
                status = "closed"
            else:
                status = "residual"

        parent = components_by_id.get(parent_id)
        rollups.append(
            {
                "parent_component_id": parent_id,
                "parent_component_name": parent.name if parent else parent_id,
                "parent_included": bool(parent_selection.included) if parent_selection else False,
                "parent_power_value_w": parent_power,
                "assigned_child_count": assigned_child_count,
                "missing_child_count": missing_child_count,
                "child_sum_power_w": round(child_sum, 4),
                "residual_power_w": residual,
                "status": status,
                "covered_descendant_ids": sorted(descendants_by_id.get(parent_id, set())),
            }
        )
    return rollups


def validate_power_rollups(rollups: list[dict[str, Any]]) -> None:
    for rollup in rollups:
        if rollup["parent_included"] and rollup["status"] == "over_specified":
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Power roll-up over-specified for {rollup['parent_component_name']}: "
                    f"children sum {rollup['child_sum_power_w']:.4f} W exceeds parent inclusive "
                    f"{rollup['parent_power_value_w']:.4f} W."
                ),
            )


@app.put("/api/application-scenario-composition")
def update_application_scenario_composition(payload: ApplicationScenarioCompositionUpdate) -> dict[str, Any]:
    with Session(engine) as session:
        validate_power_context(
            session,
            payload.project_id,
            payload.impl_option_id,
            payload.physical_mapping_id,
            application_scenario_id=payload.application_scenario_id,
        )
        existing = list(
            session.exec(
                select(ApplicationScenarioSelection).where(
                    ApplicationScenarioSelection.impl_option_id == payload.impl_option_id,
                    ApplicationScenarioSelection.physical_mapping_id == payload.physical_mapping_id,
                    ApplicationScenarioSelection.application_scenario_id == payload.application_scenario_id,
                )
            ).all()
        )
        for row in existing:
            session.delete(row)
        pending_selections: list[ApplicationScenarioSelection] = []
        library = {
            (row.scope_id, row.use_case_name or "Default", row.operating_point_set_id): row
            for row in module_power_rows(session, payload.impl_option_id, payload.physical_mapping_id)
        }
        _components_by_id, _parent_by_id, _children_by_parent, ancestors_by_id, _descendants_by_id = component_hierarchy_maps(session)
        for item in payload.selections:
            validate_power_context(
                session,
                payload.project_id,
                payload.impl_option_id,
                payload.physical_mapping_id,
                application_scenario_id=payload.application_scenario_id,
                operating_point_set_id=item.operating_point_set_id,
                component_id=item.component_id,
            )
            use_case_name = item.use_case_name.strip() or "Default"
            library_row = library.get((item.component_id, use_case_name, item.operating_point_set_id))
            if item.included and not library_row:
                raise HTTPException(
                    status_code=400,
                    detail=f"{item.component_name} / {use_case_name} / {item.operating_point_set_id} has no saved module power value.",
                )
            pending_selections.append(
                ApplicationScenarioSelection(
                    id=app_selection_id(payload.application_scenario_id, item.component_id, use_case_name, item.operating_point_set_id),
                    project_id=payload.project_id,
                    impl_option_id=payload.impl_option_id,
                    physical_mapping_id=payload.physical_mapping_id,
                    application_scenario_id=payload.application_scenario_id,
                    component_id=item.component_id,
                    component_name=item.component_name,
                    use_case_name=use_case_name,
                    operating_point_set_id=item.operating_point_set_id,
                    included=item.included,
                    note=item.note or "",
                )
            )
        validate_active_power_hierarchy(pending_selections, ancestors_by_id)
        rollups = build_power_hierarchy_rollups(session, pending_selections, library)
        validate_power_rollups(rollups)
        for row in pending_selections:
            session.add(row)
        session.commit()
        return {
            "selections": get_application_scenario_composition(payload.impl_option_id, payload.physical_mapping_id, payload.application_scenario_id),
            "summary": application_power_summary(session, payload.impl_option_id, payload.physical_mapping_id, payload.application_scenario_id),
        }


def application_power_summary(session: Session, impl_option_id: str, physical_mapping_id: str, application_scenario_id: str) -> dict[str, Any]:
    selections = list(
        session.exec(
            select(ApplicationScenarioSelection).where(
                ApplicationScenarioSelection.impl_option_id == impl_option_id,
                ApplicationScenarioSelection.physical_mapping_id == physical_mapping_id,
                ApplicationScenarioSelection.application_scenario_id == application_scenario_id,
            )
        ).all()
    )
    library = {
        (row.scope_id, row.use_case_name or "Default", row.operating_point_set_id): row
        for row in module_power_rows(session, impl_option_id, physical_mapping_id)
    }
    hierarchy_rollups = build_power_hierarchy_rollups(session, selections, library)
    rows: list[dict[str, Any]] = []
    total = 0.0
    missing = 0
    for selection in selections:
        observation = library.get((selection.component_id, selection.use_case_name, selection.operating_point_set_id))
        power = observation.power_value_w if observation else None
        if selection.included and power is not None:
            total += power
        if selection.included and power is None:
            missing += 1
        rows.append(
            {
                "id": selection.id,
                "component_id": selection.component_id,
                "component_name": selection.component_name,
                "use_case_name": selection.use_case_name,
                "operating_point_set_id": selection.operating_point_set_id,
                "included": selection.included,
                "power_value_w": power,
                "confidence": observation.confidence if observation else None,
                "note": selection.note,
            }
        )
    return {
        "filters": {
            "impl_option_id": impl_option_id,
            "physical_mapping_id": physical_mapping_id,
            "application_scenario_id": application_scenario_id,
        },
        "total_additive_power_w": round(total, 4),
        "non_additive_reference_power_w": None,
        "residual_power_w": None,
        "selected_count": sum(1 for row in selections if row.included),
        "missing_count": missing,
        "selections": rows,
        "hierarchy_rollups": hierarchy_rollups,
        "by_component": {row["component_name"]: row["power_value_w"] for row in rows if row["included"] and row["power_value_w"] is not None},
    }


@app.get("/api/application-power-summary")
def get_application_power_summary(impl_option_id: str, physical_mapping_id: str, application_scenario_id: str) -> dict[str, Any]:
    with Session(engine) as session:
        return application_power_summary(session, impl_option_id, physical_mapping_id, application_scenario_id)


@app.get("/api/power-observations")
def get_power_observations(
    project_id: str | None = None,
    impl_option_id: str | None = None,
    physical_mapping_id: str | None = None,
    application_scenario_id: str | None = None,
    operating_point_set_id: str | None = None,
    scope_type: str | None = None,
    statistic_type: str | None = None,
    power_type: str | None = None,
    development_stage: str | None = None,
    confidence: str | None = None,
    is_additive: bool | None = None,
) -> list[PowerObservation]:
    with Session(engine) as session:
        stmt = select(PowerObservation)
        if project_id:
            stmt = stmt.where(PowerObservation.project_id == project_id)
        if impl_option_id:
            stmt = stmt.where(PowerObservation.impl_option_id == impl_option_id)
        if physical_mapping_id:
            stmt = stmt.where(PowerObservation.physical_mapping_id == physical_mapping_id)
        if application_scenario_id:
            stmt = stmt.where(PowerObservation.application_scenario_id == application_scenario_id)
        if operating_point_set_id:
            stmt = stmt.where(PowerObservation.operating_point_set_id == operating_point_set_id)
        if scope_type:
            stmt = stmt.where(PowerObservation.scope_type == scope_type)
        if statistic_type:
            stmt = stmt.where(PowerObservation.statistic_type == statistic_type)
        if power_type:
            stmt = stmt.where(PowerObservation.power_type == power_type)
        if development_stage:
            stmt = stmt.where(PowerObservation.development_stage == development_stage)
        if confidence:
            stmt = stmt.where(PowerObservation.confidence == confidence)
        if is_additive is not None:
            stmt = stmt.where(PowerObservation.is_additive == is_additive)
        return list(session.exec(stmt).all())


@app.get("/api/power-summary")
def get_power_summary(
    impl_option_id: str,
    physical_mapping_id: str,
    application_scenario_id: str,
    operating_point_set_id: str,
    statistic_type: str = "average",
    power_type: str = "total",
    time_window_name: str | None = None,
    development_stage: str | None = None,
) -> dict[str, Any]:
    with Session(engine) as session:
        stmt = select(PowerObservation).where(
            PowerObservation.impl_option_id == impl_option_id,
            PowerObservation.physical_mapping_id == physical_mapping_id,
            PowerObservation.application_scenario_id == application_scenario_id,
            PowerObservation.operating_point_set_id == operating_point_set_id,
            PowerObservation.statistic_type == statistic_type,
            PowerObservation.power_type == power_type,
        )
        if time_window_name:
            stmt = stmt.where(PowerObservation.time_window_name == time_window_name)
        if development_stage:
            stmt = stmt.where(PowerObservation.development_stage == development_stage)
            
        observations = list(session.exec(stmt).all())
        
        additive_obs = [o for o in observations if o.is_additive]
        total_additive_power_w = sum(o.power_value_w for o in additive_obs)
        
        non_additive_obs = [o for o in observations if not o.is_additive]
        soc_ref = next((o for o in non_additive_obs if o.scope_type == "soc"), None)
        non_additive_reference_power_w = soc_ref.power_value_w if soc_ref else None
        
        residual_power_w = None
        if non_additive_reference_power_w is not None:
            residual_power_w = round(non_additive_reference_power_w - total_additive_power_w, 4)
            
        by_scope_type = {}
        for o in additive_obs:
            by_scope_type[o.scope_type] = round(by_scope_type.get(o.scope_type, 0.0) + o.power_value_w, 4)
            
        by_component = {}
        for o in additive_obs:
            if o.scope_type == "component":
                by_component[o.scope_name] = round(by_component.get(o.scope_name, 0.0) + o.power_value_w, 4)
                
        by_stage = {}
        for o in observations:
            stage = o.development_stage or "unknown"
            by_stage[stage] = by_stage.get(stage, 0) + 1
            
        return {
            "filters": {
                "impl_option_id": impl_option_id,
                "physical_mapping_id": physical_mapping_id,
                "application_scenario_id": application_scenario_id,
                "operating_point_set_id": operating_point_set_id,
                "statistic_type": statistic_type,
                "power_type": power_type,
                "time_window_name": time_window_name,
                "development_stage": development_stage,
            },
            "total_additive_power_w": round(total_additive_power_w, 4),
            "non_additive_reference_power_w": non_additive_reference_power_w,
            "residual_power_w": residual_power_w,
            "by_scope_type": by_scope_type,
            "by_component": by_component,
            "by_stage": by_stage,
            "non_additive_references": [
                {
                    "scope_type": o.scope_type,
                    "scope_name": o.scope_name,
                    "power_value_w": o.power_value_w,
                    "development_stage": o.development_stage,
                }
                for o in non_additive_obs
            ],
            "observations": [o.dict() for o in observations],
        }


@app.post("/api/power-observations")
def create_power_observation(payload: PowerObservationCreate) -> PowerObservation:
    import uuid
    with Session(engine) as session:
        validate_power_context(
            session,
            payload.project_id,
            payload.impl_option_id,
            payload.physical_mapping_id,
            application_scenario_id=payload.application_scenario_id,
            operating_point_set_id=payload.operating_point_set_id,
            component_id=payload.scope_id if payload.scope_type == "component" else None,
        )
        if payload.power_value_w < 0:
            raise HTTPException(status_code=400, detail="power_value_w must be >= 0")
        
        obs_id = f"PO_{uuid.uuid4().hex[:8].upper()}"
        obs = PowerObservation(
            id=obs_id,
            project_id=payload.project_id,
            impl_option_id=payload.impl_option_id,
            physical_mapping_id=payload.physical_mapping_id,
            application_scenario_id=payload.application_scenario_id,
            operating_point_set_id=payload.operating_point_set_id,
            scope_type=payload.scope_type,
            scope_id=payload.scope_id,
            scope_name=payload.scope_name,
            use_case_name=payload.use_case_name,
            time_window_name=payload.time_window_name,
            statistic_type=payload.statistic_type,
            power_type=payload.power_type,
            power_value_w=payload.power_value_w,
            development_stage=payload.development_stage,
            source_type="web_ui",
            confidence=payload.confidence or "draft",
            is_additive=payload.is_additive,
            note=payload.note
        )
        session.add(obs)
        session.commit()
        session.refresh(obs)
        return obs


@app.put("/api/power-observations/{observation_id}")
def update_power_observation(observation_id: str, payload: PowerObservationCreate) -> PowerObservation:
    with Session(engine) as session:
        obs = session.get(PowerObservation, observation_id)
        if not obs:
            raise HTTPException(status_code=404, detail=f"Unknown power observation: {observation_id}")
        validate_power_context(
            session,
            payload.project_id,
            payload.impl_option_id,
            payload.physical_mapping_id,
            application_scenario_id=payload.application_scenario_id,
            operating_point_set_id=payload.operating_point_set_id,
            component_id=payload.scope_id if payload.scope_type == "component" else None,
        )
        if payload.power_value_w < 0:
            raise HTTPException(status_code=400, detail="power_value_w must be >= 0")
        
        # Update fields
        obs.scope_type = payload.scope_type
        obs.scope_id = payload.scope_id
        obs.scope_name = payload.scope_name
        obs.use_case_name = payload.use_case_name
        obs.time_window_name = payload.time_window_name
        obs.statistic_type = payload.statistic_type
        obs.power_type = payload.power_type
        obs.power_value_w = payload.power_value_w
        obs.development_stage = payload.development_stage
        obs.confidence = payload.confidence
        obs.is_additive = payload.is_additive
        obs.note = payload.note
        
        session.add(obs)
        session.commit()
        session.refresh(obs)
        return obs


@app.delete("/api/power-observations/{observation_id}")
def delete_power_observation(observation_id: str) -> dict[str, Any]:
    with Session(engine) as session:
        obs = session.get(PowerObservation, observation_id)
        if not obs:
            raise HTTPException(status_code=404, detail=f"Unknown power observation: {observation_id}")
        session.delete(obs)
        session.commit()
        return {"success": True, "deleted_id": observation_id}


@app.get("/api/design-options")
def get_design_options() -> list[dict[str, Any]]:
    return get_impl_options()


@app.get("/api/projects")
def get_projects() -> list[Project]:
    with Session(engine) as session:
        return list(session.exec(select(Project)).all())


@app.get("/api/impl-options")
def get_implOptions() -> list[dict[str, Any]]:
    with Session(engine) as session:
        return [impl_option_ui(session, impl_option) for impl_option in session.exec(select(ImplOption)).all()]


@app.get("/api/impl-options/{impl_option_id}/detail")
def get_impl_option_detail(impl_option_id: str) -> dict[str, Any]:
    with Session(engine) as session:
        impl_option = session.get(ImplOption, impl_option_id)
        if not impl_option:
            raise HTTPException(status_code=404, detail=f"Unknown impl_option_id: {impl_option_id}")
        return impl_option_detail_ui(session, impl_option_id)


@app.put("/api/impl-options/{impl_option_id}/detail")
def update_impl_option_detail(impl_option_id: str, payload: ImplOptionDetailUpdate) -> dict[str, Any]:
    with Session(engine) as session:
        impl_option = session.get(ImplOption, impl_option_id)
        if not impl_option:
            raise HTTPException(status_code=404, detail=f"Unknown impl_option_id: {impl_option_id}")

        errors = impl_option_detail_impact_errors(session, impl_option_id, payload)
        if errors:
            raise HTTPException(status_code=409, detail={"errors": errors})

        previous = session.get(ImplOptionDetail, impl_option_id)
        session.merge(
            ImplOptionDetail(
                impl_option_id=impl_option_id,
                implementation_type=payload.implementation_type,
                status=payload.status,
                version=(previous.version + 1) if previous else 1,
                updated_at=now_iso(),
            )
        )
        session.exec(delete(ImplementationTier).where(ImplementationTier.impl_option_id == impl_option_id))
        session.exec(delete(ImplementationInterface).where(ImplementationInterface.impl_option_id == impl_option_id))
        existing_escape = session.get(ImplementationPackageEscape, impl_option_id)
        if existing_escape:
            session.delete(existing_escape)

        for index, tier in enumerate(payload.tiers):
            session.add(
                ImplementationTier(
                    id=f"{impl_option_id}:{tier.id}",
                    impl_option_id=impl_option_id,
                    tier_id=tier.id,
                    tier_index=index,
                    name=tier.name,
                    process=tier.process,
                    role=tier.role,
                    thickness_um=tier.thickness_um,
                )
            )
        for row in payload.interfaces:
            session.add(
                ImplementationInterface(
                    id=f"{impl_option_id}:{row.id}",
                    impl_option_id=impl_option_id,
                    from_tier_id=row.from_tier_id,
                    to_tier_id=row.to_tier_id,
                    orientation=row.orientation,
                    interconnect=row.interconnect,
                    hb_pitch_um=row.hb_pitch_um,
                    upper_tsv_pitch_um=row.upper_tsv_pitch_um,
                    upper_tsv_keepout_um=row.upper_tsv_keepout_um,
                    lower_tsv_pitch_um=row.lower_tsv_pitch_um,
                    lower_tsv_keepout_um=row.lower_tsv_keepout_um,
                    description=row.description,
                )
            )
        session.add(
            ImplementationPackageEscape(
                impl_option_id=impl_option_id,
                bottom_tier_id=payload.package_escape.bottom_tier_id,
                requires_tsv=payload.package_escape.requires_tsv,
                pitch_um=payload.package_escape.pitch_um,
                keepout_um=payload.package_escape.keepout_um,
                description=payload.package_escape.description,
            )
        )
        session.commit()
        return {"implementation": impl_option_detail_ui(session, impl_option_id), "impact": {"blocked": False, "errors": []}}


@app.get("/api/module-definitions")
def get_module_definitions() -> list[ModuleDefinition]:
    with Session(engine) as session:
        return list(session.exec(select(ModuleDefinition)).all())


@app.post("/api/components")
def create_logical_component(payload: LogicalComponentInput) -> dict[str, Any]:
    with Session(engine) as session:
        if not session.get(Project, payload.project_id):
            raise HTTPException(status_code=400, detail=f"Unknown project_id: {payload.project_id}")
        if payload.parent_id:
            ensure_component_write_scope(session, payload.parent_id, payload.team, payload.impl_option_id)
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Component name is required.")
        if payload.logical_instance_count < 0:
            raise HTTPException(status_code=400, detail="logical_instance_count must be non-negative")
        component_id = (payload.id or "").strip() or component_id_from_name(session, name)
        if session.get(LogicalComponent, component_id):
            raise HTTPException(status_code=409, detail=f"Logical component already exists: {component_id}")
        if payload.module_definition_id and not session.get(ModuleDefinition, payload.module_definition_id):
            raise HTTPException(status_code=400, detail=f"Unknown module_definition_id: {payload.module_definition_id}")
        row = LogicalComponent(
            id=component_id,
            project_id=payload.project_id,
            parent_id=payload.parent_id or None,
            module_definition_id=payload.module_definition_id or None,
            name=name,
            instance_type=payload.instance_type.strip() or "block",
            resource_type=payload.resource_type.strip() or "logic",
            function_domain=payload.function_domain.strip() or "General",
            hierarchy_path=component_path(session, payload.parent_id, name),
            logical_instance_count=payload.logical_instance_count,
            owner_team=payload.owner_team.strip() or "Architecture Team",
            visibility_level=payload.visibility_level.strip() or "team",
            description=payload.description or "",
            created_at=now_iso(),
            updated_at=now_iso(),
        )
        session.add(row)
        session.commit()
        session.refresh(row)
        return component_ui(session, row, payload.impl_option_id)


@app.put("/api/components/{component_id}")
def update_logical_component(component_id: str, payload: LogicalComponentInput) -> dict[str, Any]:
    with Session(engine) as session:
        row = session.get(LogicalComponent, component_id)
        if not row:
            raise HTTPException(status_code=404, detail=f"Unknown logical component: {component_id}")
        ensure_component_write_scope(session, component_id, payload.team, payload.impl_option_id)
        if payload.parent_id:
            ensure_component_write_scope(session, payload.parent_id, payload.team, payload.impl_option_id)
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Component name is required.")
        if payload.logical_instance_count < 0:
            raise HTTPException(status_code=400, detail="logical_instance_count must be non-negative")
        if payload.parent_id == component_id:
            raise HTTPException(status_code=400, detail="A component cannot be its own parent.")
        if payload.parent_id and payload.parent_id in descendant_component_ids(session, component_id):
            raise HTTPException(status_code=400, detail="A component cannot be moved under its own descendant.")
        if payload.parent_id and not session.get(LogicalComponent, payload.parent_id):
            raise HTTPException(status_code=400, detail=f"Unknown parent_id: {payload.parent_id}")
        if payload.module_definition_id and not session.get(ModuleDefinition, payload.module_definition_id):
            raise HTTPException(status_code=400, detail=f"Unknown module_definition_id: {payload.module_definition_id}")

        row.project_id = payload.project_id
        row.parent_id = payload.parent_id or None
        row.module_definition_id = payload.module_definition_id or None
        row.name = name
        row.instance_type = payload.instance_type.strip() or "block"
        row.resource_type = payload.resource_type.strip() or "logic"
        row.function_domain = payload.function_domain.strip() or "General"
        row.logical_instance_count = payload.logical_instance_count
        row.owner_team = payload.owner_team.strip() or "Architecture Team"
        row.visibility_level = payload.visibility_level.strip() or "team"
        row.description = payload.description or ""
        row.hierarchy_path = component_path(session, row.parent_id, row.name)
        row.updated_at = now_iso()
        session.add(row)
        update_component_subtree_paths(session, row)
        session.commit()
        session.refresh(row)
        return component_ui(session, row, payload.impl_option_id)


@app.delete("/api/components/{component_id}")
def delete_logical_component(component_id: str, payload: LogicalComponentDeleteInput) -> dict[str, Any]:
    with Session(engine) as session:
        row = session.get(LogicalComponent, component_id)
        if not row:
            raise HTTPException(status_code=404, detail=f"Unknown logical component: {component_id}")
        ensure_component_write_scope(session, component_id, payload.team, payload.impl_option_id)
        ids = {component_id, *descendant_component_ids(session, component_id)}
        if not payload.cascade and len(ids) > 1:
            raise HTTPException(status_code=409, detail="Component has children. Enable cascade to delete the subtree.")
        partitions = session.exec(select(PhysicalPartition).where(PhysicalPartition.logical_component_id.in_(ids))).all()
        partition_ids = {partition.id for partition in partitions}
        if partition_ids:
            session.exec(delete(Metric).where(Metric.subject_type == "physical_partition", Metric.subject_id.in_(partition_ids)))
        session.exec(delete(PhysicalPartition).where(PhysicalPartition.logical_component_id.in_(ids)))
        session.exec(delete(Metric).where(Metric.subject_type == "logical_component", Metric.subject_id.in_(ids)))
        session.exec(delete(PowerObservation).where(PowerObservation.scope_type == "component", PowerObservation.scope_id.in_(ids)))
        session.exec(delete(ApplicationScenarioSelection).where(ApplicationScenarioSelection.component_id.in_(ids)))
        for logical_id in sorted(ids, key=lambda value: session.get(LogicalComponent, value).hierarchy_path if session.get(LogicalComponent, value) else "", reverse=True):
            item = session.get(LogicalComponent, logical_id)
            if item:
                session.delete(item)
        session.commit()
        return {"deleted_component_ids": sorted(ids), "deleted_partition_ids": sorted(partition_ids)}


@app.get("/api/components")
def get_components(impl_option_id: str = "S2", team: str | None = None) -> list[dict[str, Any]]:
    with Session(engine) as session:
        rows, allowed = component_rows_for_team(session, team, impl_option_id)
        return scope_component_items([component_ui(session, row, impl_option_id) for row in rows], allowed)


@app.get("/api/components/tree")
def get_component_tree(impl_option_id: str = "S2", team: str | None = None) -> list[dict[str, Any]]:
    with Session(engine) as session:
        rows, allowed = component_rows_for_team(session, team, impl_option_id)
        return build_component_tree(scope_component_items([component_ui(session, row, impl_option_id) for row in rows], allowed))


@app.get("/api/physical-partitions")
def get_physical_partitions(impl_option_id: str = "S2", team: str | None = None) -> list[dict[str, Any]]:
    with Session(engine) as session:
        allowed = allowed_component_ids_for_team(session, team, impl_option_id)
        rows = session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()
        if allowed is not None:
            rows = [row for row in rows if row.logical_component_id in allowed]
        return [partition_ui(session, row) for row in rows]



def recalculate_component_partitions(session: Session, impl_option_id: str, component_id: str):
    partitions = session.exec(
        select(PhysicalPartition).where(
            PhysicalPartition.impl_option_id == impl_option_id,
            PhysicalPartition.logical_component_id == component_id,
        )
    ).all()
    if not partitions:
        return

    logical_metrics = metrics_for(session, impl_option_id, "logical_component", component_id)
    current_logic_area = metric_number(logical_metrics, "logic_area")
    current_sram_area = metric_number(logical_metrics, "sram_area")
    current_block_area = metric_number(logical_metrics, "block_area")
    current_power = metric_number(logical_metrics, "power")

    child_rows = session.exec(select(LogicalComponent).where(LogicalComponent.parent_id == component_id)).all()
    child_sum = {"logic_area": 0.0, "sram_area": 0.0, "block_area": 0.0}
    for child in child_rows:
        child_metrics = metrics_for(session, impl_option_id, "logical_component", child.id)
        for m_name in child_sum:
            child_sum[m_name] += metric_number(child_metrics, m_name)

    self_area = {
        "logic": max(0.0, current_logic_area - child_sum["logic_area"]),
        "sram": max(0.0, current_sram_area - child_sum["sram_area"]),
        "block": max(0.0, current_block_area - child_sum["block_area"]),
    }

    required_cats = {cat for cat, val in self_area.items() if val > 0.01}
    if not required_cats:
        required_cats = {"block"}
    category_count = len(required_cats)

    partitions_by_cat = {}
    for p in partitions:
        partitions_by_cat.setdefault(normalized_resource_category(p.resource_category), []).append(p)

    for cat in ["logic", "sram", "block"]:
        cat_items = partitions_by_cat.get(cat, [])
        cat_area = self_area[cat]

        total_equiv = 0.0
        for p in cat_items:
            c_share = normalized_content_share(p.partition_type, p.content_share if p.content_share is not None else p.partition_ratio)
            total_equiv += p.physical_instance_count * c_share

        if total_equiv <= 0.001:
            total_equiv = 1.0

        for p in cat_items:
            c_share = normalized_content_share(p.partition_type, p.content_share if p.content_share is not None else p.partition_ratio)
            equiv = p.physical_instance_count * c_share
            share = equiv / total_equiv

            p_logic_val = round(cat_area * share, 3) if cat == "logic" else 0.0
            p_sram_val = round(cat_area * share, 3) if cat == "sram" else 0.0
            p_block_val = round(cat_area * share, 3) if cat == "block" else 0.0
            p_power_val = round(current_power * share / category_count, 3)
            p_shape = f"{cat}_{p.tier_id.lower()}"

            p_metric_configs = [
                ("logic_area", p_logic_val, "mm2", "implementation_area", "number", "typical", "nominal"),
                ("sram_area", p_sram_val, "mm2", "implementation_area", "number", "typical", "nominal"),
                ("block_area", p_block_val, "mm2", "implementation_area", "number", "typical", "nominal"),
                ("power", p_power_val, "W", "power", "number", "typical", "peak"),
                ("shape_type", p_shape, "", "physical_shape", "text", "typical", "nominal"),
            ]
            for name, val, unit, category_type, val_type, corner, workload in p_metric_configs:
                metric_id = f"M_PART_{p.id}_{name.upper()}"
                existing_metric = session.get(Metric, metric_id)
                if existing_metric:
                    existing_metric.metric_value = str(val)
                    session.add(existing_metric)
                else:
                    new_metric = Metric(
                        id=metric_id,
                        impl_option_id=impl_option_id,
                        subject_type="physical_partition",
                        subject_id=p.id,
                        metric_name=name,
                        metric_value=str(val),
                        metric_unit=unit,
                        metric_category=category_type,
                        value_type=val_type,
                        corner=corner,
                        workload=workload,
                        confidence="review",
                        source_note="Recalculated on child area change",
                        created_at=now_iso(),
                    )
                    session.add(new_metric)


@app.put("/api/components/{component_id}/detail")
def update_component_detail(component_id: str, payload: ComponentDetailUpdate) -> dict[str, Any]:
    with Session(engine) as session:
        component = session.get(LogicalComponent, component_id)
        if not component:
            raise HTTPException(status_code=404, detail=f"Unknown logical component: {component_id}")

        allowed = allowed_component_ids_for_team(session, payload.team, payload.impl_option_id)
        if allowed is not None and component_id not in allowed:
            raise HTTPException(status_code=403, detail=f"{component_id} is outside team scope {payload.team}")

        impl_option = session.get(ImplOption, payload.impl_option_id)
        if not impl_option:
            raise HTTPException(status_code=400, detail=f"Unknown impl_option_id: {payload.impl_option_id}")
        tier_ids = {row.id for row in session.exec(select(Tier).where(Tier.impl_option_id == payload.impl_option_id)).all()}
        if payload.logical_instance_count < 0:
            raise HTTPException(status_code=400, detail="logical_instance_count must be non-negative")

        canonical_partitions: list[tuple[PartitionInput, str, str, str]] = []
        partial_counters: dict[tuple[str, str], int] = {}
        for partition in payload.partitions:
            category = normalized_resource_category(partition.resource_category)
            partial_index = 0
            if partition.partition_type == "partial":
                counter_key = (category, partition.tier_id)
                partial_index = partial_counters.get(counter_key, 0) + 1
                partial_counters[counter_key] = partial_index
            partition_name = canonical_partition_name(component.name, category, partition.tier_id, partition.partition_type, partial_index)
            canonical_partitions.append((partition, category, f"PP_{partition_name}", partition_name))

        seen_partition_ids: set[str] = set()
        for partition, category, partition_id, _partition_name in canonical_partitions:
            if partition_id in seen_partition_ids:
                raise HTTPException(status_code=400, detail=f"Duplicate generated partition id: {partition_id}")
            seen_partition_ids.add(partition_id)
            if partition.tier_id not in tier_ids:
                raise HTTPException(status_code=400, detail=f"Unknown tier_id for impl_option {payload.impl_option_id}: {partition.tier_id}")
            if partition.partition_type not in ALLOWED_PARTITION_TYPES:
                raise HTTPException(status_code=400, detail=f"Unsupported partition_type: {partition.partition_type}")
            if category not in ALLOWED_PARTITION_RESOURCE_CATEGORIES:
                raise HTTPException(status_code=400, detail=f"Unsupported resource_category: {partition.resource_category}")
            if partition.physical_instance_count < 0:
                raise HTTPException(status_code=400, detail=f"{partition.id} has negative physical_instance_count")
            content_share = normalized_content_share(partition.partition_type, partition.content_share if partition.content_share is not None else partition.partition_ratio)
            if content_share < 0:
                raise HTTPException(status_code=400, detail=f"{partition.id} has negative content_share")
            if partition.partition_type == "full" and abs(content_share - 1.0) > 0.001:
                raise HTTPException(status_code=400, detail=f"{partition.id} is full, so content_share must be 1")

        component.logical_instance_count = payload.logical_instance_count
        component.updated_at = now_iso()
        session.add(component)

        # Update logical component metrics if provided
        metric_configs = [
            ("signal_count_total", payload.signal_count_total, "count", "logical", "number", "typical", "nominal"),
            ("logic_area", payload.logic_area, "mm2", "logical_area", "number", "typical", "nominal"),
            ("sram_area", payload.sram_area, "mm2", "logical_area", "number", "typical", "nominal"),
            ("block_area", payload.block_area, "mm2", "logical_area", "number", "typical", "nominal"),
            ("power", payload.power, "W", "power", "number", "typical", "peak"),
        ]
        
        for name, val, unit, category, value_type, corner, workload in metric_configs:
            if val is not None:
                metric_id = f"M_LOG_{component_id}_{name.upper()}"
                existing_metric = session.get(Metric, metric_id)
                if existing_metric:
                    existing_metric.metric_value = str(val)
                    session.add(existing_metric)
                else:
                    new_metric = Metric(
                        id=metric_id,
                        impl_option_id=payload.impl_option_id,
                        subject_type="logical_component",
                        subject_id=component_id,
                        metric_name=name,
                        metric_value=str(val),
                        metric_unit=unit,
                        metric_category=category,
                        value_type=value_type,
                        corner=corner,
                        workload=workload,
                        confidence="review",
                        source_note="Updated via web interface editor",
                        created_at=now_iso(),
                    )
                    session.add(new_metric)

        existing = session.exec(
            select(PhysicalPartition).where(
                PhysicalPartition.impl_option_id == payload.impl_option_id,
                PhysicalPartition.logical_component_id == component_id,
            )
        ).all()
        for partition in existing:
            if partition.id not in seen_partition_ids:
                session.exec(
                    delete(Metric).where(
                        Metric.impl_option_id == payload.impl_option_id,
                        Metric.subject_type == "physical_partition",
                        Metric.subject_id == partition.id,
                    )
                )
                session.delete(partition)

        for partition, category, partition_id, partition_name in canonical_partitions:
            content_share = normalized_content_share(partition.partition_type, partition.content_share if partition.content_share is not None else partition.partition_ratio)
            row = PhysicalPartition(
                id=partition_id,
                impl_option_id=payload.impl_option_id,
                logical_component_id=component_id,
                tier_id=partition.tier_id,
                partition_name=partition_name,
                partition_type=partition.partition_type,
                resource_category=category,
                physical_instance_count=partition.physical_instance_count,
                partition_ratio=content_share,
                content_share=content_share,
                description=partition.description,
            )
            session.merge(row)

        # Recalculate physical partition metrics to maintain data consistency
        logical_metrics = metrics_for(session, payload.impl_option_id, "logical_component", component_id)
        current_logic_area = payload.logic_area if payload.logic_area is not None else metric_number(logical_metrics, "logic_area")
        current_sram_area = payload.sram_area if payload.sram_area is not None else metric_number(logical_metrics, "sram_area")
        current_block_area = payload.block_area if payload.block_area is not None else metric_number(logical_metrics, "block_area")
        current_power = payload.power if payload.power is not None else metric_number(logical_metrics, "power")

        child_rows = session.exec(select(LogicalComponent).where(LogicalComponent.parent_id == component_id)).all()
        child_sum = {"logic_area": 0.0, "sram_area": 0.0, "block_area": 0.0}
        for child in child_rows:
            child_metrics = metrics_for(session, payload.impl_option_id, "logical_component", child.id)
            for m_name in child_sum:
                child_sum[m_name] += metric_number(child_metrics, m_name)

        self_area = {
            "logic": max(0.0, current_logic_area - child_sum["logic_area"]),
            "sram": max(0.0, current_sram_area - child_sum["sram_area"]),
            "block": max(0.0, current_block_area - child_sum["block_area"]),
        }

        required_cats = {cat for cat, val in self_area.items() if val > 0.01}
        if not required_cats:
            required_cats = {"block"}
        category_count = len(required_cats)

        partitions_by_cat = {}
        for item in canonical_partitions:
            partitions_by_cat.setdefault(item[1], []).append(item)

        for cat in ["logic", "sram", "block"]:
            cat_items = partitions_by_cat.get(cat, [])
            cat_area = self_area[cat]
            
            total_equiv = 0.0
            for p_in, _, _, _ in cat_items:
                c_share = normalized_content_share(p_in.partition_type, p_in.content_share if p_in.content_share is not None else p_in.partition_ratio)
                total_equiv += p_in.physical_instance_count * c_share
                
            if total_equiv <= 0.001:
                total_equiv = 1.0
                
            for p_in, _, p_id, _ in cat_items:
                c_share = normalized_content_share(p_in.partition_type, p_in.content_share if p_in.content_share is not None else p_in.partition_ratio)
                equiv = p_in.physical_instance_count * c_share
                share = equiv / total_equiv
                
                p_logic_val = round(cat_area * share, 3) if cat == "logic" else 0.0
                p_sram_val = round(cat_area * share, 3) if cat == "sram" else 0.0
                p_block_val = round(cat_area * share, 3) if cat == "block" else 0.0
                p_power_val = round(current_power * share / category_count, 3)
                p_shape = f"{cat}_{p_in.tier_id.lower()}"
                
                p_metric_configs = [
                    ("logic_area", p_logic_val, "mm2", "implementation_area", "number", "typical", "nominal"),
                    ("sram_area", p_sram_val, "mm2", "implementation_area", "number", "typical", "nominal"),
                    ("block_area", p_block_val, "mm2", "implementation_area", "number", "typical", "nominal"),
                    ("power", p_power_val, "W", "power", "number", "typical", "peak"),
                    ("shape_type", p_shape, "", "physical_shape", "text", "typical", "nominal"),
                ]
                for name, val, unit, category_type, val_type, corner, workload in p_metric_configs:
                    metric_id = f"M_PART_{p_id}_{name.upper()}"
                    existing_metric = session.get(Metric, metric_id)
                    if existing_metric:
                        existing_metric.metric_value = str(val)
                        session.add(existing_metric)
                    else:
                        new_metric = Metric(
                            id=metric_id,
                            impl_option_id=payload.impl_option_id,
                            subject_type="physical_partition",
                            subject_id=p_id,
                            metric_name=name,
                            metric_value=str(val),
                            metric_unit=unit,
                            metric_category=category_type,
                            value_type=val_type,
                            corner=corner,
                            workload=workload,
                            confidence="review",
                            source_note="Recalculated on component detail save",
                            created_at=now_iso(),
                        )
                        session.add(new_metric)

        if component.parent_id:
            recalculate_component_partitions(session, payload.impl_option_id, component.parent_id)

        session.commit()
        session.refresh(component)
        return {
            "component": component_ui(session, component, payload.impl_option_id),
            "quality_issues": quality_issues_for(session, payload.impl_option_id, payload.team),
        }


@app.get("/api/tiers")
def get_tiers(impl_option_id: str = "S2") -> list[dict[str, Any]]:
    with Session(engine) as session:
        tiers = session.exec(select(Tier).where(Tier.impl_option_id == impl_option_id).order_by(Tier.tier_index)).all()
        process_nodes = {node.id: node for node in session.exec(select(ProcessNode)).all()}
        return [
            {
                "id": tier.id,
                "impl_option_id": tier.impl_option_id,
                "tier_index": tier.tier_index,
                "name": tier.name,
                "process_id": tier.process_id,
                "process": f"{process_nodes[tier.process_id].foundry} {process_nodes[tier.process_id].node_name}" if tier.process_id in process_nodes else tier.process_id,
                "role": tier.role,
                "orientation": tier.orientation,
                "thickness_um": tier.thickness_um,
                "area": tier.area_limit_mm2,
                "area_limit_mm2": tier.area_limit_mm2,
                "power": metric_number(metrics_for(session, impl_option_id, "tier", tier.id), "power"),
                "utilization": metric_number(metrics_for(session, impl_option_id, "tier", tier.id), "utilization"),
                "interconnect": "HB < 1um" if tier.id == "T0" else "HB + TSV" if tier.id == "T1" else "TSV < 5um",
                "description": tier.description,
            }
            for tier in tiers
        ]


@app.get("/api/metrics")
def get_metrics(impl_option_id: str | None = None, team: str | None = None) -> list[Metric]:
    with Session(engine) as session:
        statement = select(Metric)
        if impl_option_id:
            statement = statement.where(Metric.impl_option_id == impl_option_id)
        rows = list(session.exec(statement).all())
        if is_global_team(team):
            return rows
        scoped_impl_option_id = impl_option_id or "S2"
        allowed_component_ids = allowed_component_ids_for_team(session, team, scoped_impl_option_id) or set()
        allowed_partition_ids = partition_ids_for_components(session, scoped_impl_option_id, allowed_component_ids)
        return [
            row
            for row in rows
            if row.impl_option_id == scoped_impl_option_id
            and (
                (row.subject_type == "logical_component" and row.subject_id in allowed_component_ids)
                or (row.subject_type == "physical_partition" and row.subject_id in allowed_partition_ids)
            )
        ]


@app.get("/api/quality/issues")
def get_quality_issues(impl_option_id: str = "S2", team: str | None = None) -> list[dict[str, str]]:
    with Session(engine) as session:
        return quality_issues_for(session, impl_option_id, team)


@app.get("/api/responsibilities/teams")
def get_responsibility_teams(impl_option_id: str = "S2") -> list[str]:
    with Session(engine) as session:
        assignments = session.exec(
            select(ResponsibilityAssignment).where(ResponsibilityAssignment.impl_option_id == impl_option_id)
        ).all()
        teams = {assignment.team_name for assignment in assignments}
        teams.update(row.owner_team for row in session.exec(select(LogicalComponent)).all() if row.owner_team)
        return ["Architecture Team"] + sorted(team for team in teams if team != "Architecture Team")


@app.get("/api/dashboard")
def get_dashboard(impl_option_id: str = "S2") -> dict[str, Any]:
    with Session(engine) as session:
        implOptions = [impl_option_ui(session, row) for row in session.exec(select(ImplOption)).all()]
        component_rows = session.exec(select(LogicalComponent).order_by(LogicalComponent.hierarchy_path)).all()
        components = [component_ui(session, row, impl_option_id) for row in component_rows]
        partitions = [partition_ui(session, row) for row in session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()]
        target = next((item for item in implOptions if item["id"] == impl_option_id), implOptions[0])
        parent_ids = {row.parent_id for row in component_rows if row.parent_id}
        leaf_ids = {row.id for row in component_rows if row.id not in parent_ids}
        leaf_components = [item for item in components if item["id"] in leaf_ids]
        sram_area = sum(item["sram_area"] for item in leaf_components)
        phy_area = sum(item["block_area"] for item in leaf_components if "phy" in item["resource"])
        resource_area = {"Logic + mixed": 0.0, "SRAM / memory": 0.0, "PHY / Analog": 0.0}
        for item in leaf_components:
            if "phy" in item["resource"]:
                resource_area["PHY / Analog"] += item["block_area"]
            elif "memory" in item["resource"] and "logic" not in item["resource"]:
                resource_area["SRAM / memory"] += item["block_area"]
            else:
                resource_area["Logic + mixed"] += item["block_area"]
        total = sum(resource_area.values()) or 1
        return {
            "target_impl_option": target,
            "metrics": {
                "total_area": target["area"],
                "total_power": target["power"],
                "total_sram_area": round(sram_area, 2),
                "phy_area": round(phy_area, 1),
                "partition_count": len(partitions),
            },
            "resource_mix": [
                {"label": label, "value": round(value / total * 100), "tone": tone}
                for (label, value), tone in zip(resource_area.items(), ["bg-slate-900", "bg-slate-500", "bg-slate-300"])
            ],
            "projects": list(session.exec(select(Project)).all()),
            "implOptions": implOptions,
        }


IMPORT_SHEETS: dict[str, tuple[type[SQLModel], list[str], set[str]]] = {
    "module_definitions": (
        ModuleDefinition,
        ["id", "name", "module_type", "ip_owner", "reuse_class", "description"],
        {"id", "name", "module_type"},
    ),
    "projects": (
        Project,
        ["id", "name", "product_family", "generation", "owner", "phase"],
        {"id", "name", "product_family", "generation", "owner", "phase"},
    ),
    "implOptions": (
        ImplOption,
        ["id", "project_id", "name", "impl_type", "process_combo", "status"],
        {"id", "project_id", "name", "impl_type", "process_combo", "status"},
    ),
    "tiers": (
        Tier,
        ["id", "impl_option_id", "tier_index", "name", "process_id", "role", "orientation", "area_limit_mm2"],
        {"id", "impl_option_id", "tier_index", "name", "process_id", "role"},
    ),
    "logical_components": (
        LogicalComponent,
        ["id", "project_id", "parent_id", "module_definition_id", "name", "instance_type", "resource_type", "function_domain", "hierarchy_path", "logical_instance_count", "description"],
        {"id", "project_id", "name", "instance_type", "resource_type", "function_domain", "hierarchy_path", "logical_instance_count"},
    ),
    "physical_partitions": (
        PhysicalPartition,
        ["id", "impl_option_id", "logical_component_id", "tier_id", "partition_name", "resource_category", "partition_type", "physical_instance_count", "content_share", "description"],
        {"id", "impl_option_id", "logical_component_id", "tier_id", "partition_name", "partition_type", "physical_instance_count"},
    ),
    "metrics": (
        Metric,
        ["id", "impl_option_id", "subject_type", "subject_id", "metric_name", "metric_value", "metric_unit", "metric_category", "value_type", "corner", "workload", "confidence", "source_note", "created_at"],
        {"impl_option_id", "subject_type", "subject_id", "metric_name", "metric_value", "value_type", "corner", "workload", "confidence"},
    ),
}

ALLOWED_SUBJECT_TYPES = {"logical_component", "physical_partition", "tier", "impl_option"}
ALLOWED_VALUE_TYPES = {"number", "text", "boolean"}
ALLOWED_CONFIDENCE = {"approved", "review", "draft"}
ALLOWED_PARTITION_TYPES = {"full", "partial"}
ALLOWED_PARTITION_RESOURCE_CATEGORIES = {"logic", "sram", "block"}


def normalize_cell(value: Any) -> Any:
    if value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def read_sheet_rows(workbook_file: SpooledTemporaryFile[bytes], sheet_name: str, expected_columns: list[str], required_columns: set[str]) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_file, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Missing sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    header = [str(cell.value or "").strip() for cell in sheet[1]]
    aliases = {"content_share": "partition_ratio"} if sheet_name == "physical_partitions" else {}
    missing = [column for column in required_columns if column not in header and aliases.get(column) not in header]
    if missing:
        raise HTTPException(status_code=400, detail=f"Sheet {sheet_name} is missing columns: {', '.join(missing)}")
    indexes = {
        column: header.index(column if column in header else aliases[column])
        for column in expected_columns
        if column in header or aliases.get(column) in header
    }
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {column: normalize_cell(row[indexes[column]]) if column in indexes else None for column in expected_columns}
        if all(value is None for value in record.values()):
            continue
        missing_required = [column for column in required_columns if record.get(column) is None]
        if missing_required:
            raise HTTPException(status_code=400, detail=f"Sheet {sheet_name} row {row_index} missing required columns: {', '.join(missing_required)}")
        rows.append(record)
    return rows


def prepare_import_rows(all_rows: dict[str, list[dict[str, Any]]]) -> None:
    created = now_iso()
    for row in all_rows["projects"]:
        row.setdefault("description", "")
        row["created_at"] = row.get("created_at") or created
        row["updated_at"] = row.get("updated_at") or created
    for row in all_rows["implOptions"]:
        row.setdefault("description", "")
        row["created_at"] = row.get("created_at") or created
        row["updated_at"] = row.get("updated_at") or created
    for row in all_rows["tiers"]:
        row["thickness_um"] = row.get("thickness_um") or 0
        row.setdefault("description", "")
    for row in all_rows["logical_components"]:
        row["parent_id"] = row.get("parent_id") or None
        row["module_definition_id"] = row.get("module_definition_id") or None
        row["logical_instance_count"] = int(row["logical_instance_count"])
        row["owner_team"] = row.get("owner_team") or "Architecture Team"
        row["visibility_level"] = row.get("visibility_level") or "team"
        row["created_at"] = row.get("created_at") or created
        row["updated_at"] = row.get("updated_at") or created
    for row in all_rows["physical_partitions"]:
        row["physical_instance_count"] = int(row["physical_instance_count"])
        row["resource_category"] = normalized_resource_category(row.get("resource_category"))
        raw_content_share = row.get("content_share")
        row["content_share"] = normalized_content_share(row["partition_type"], float(raw_content_share) if raw_content_share is not None else None)
        row["partition_ratio"] = row["content_share"]
    for row in all_rows["metrics"]:
        row["metric_value"] = str(row["metric_value"])
        row["metric_unit"] = row.get("metric_unit") or ""
        row["metric_category"] = row.get("metric_category") or ""
        row["source_note"] = row.get("source_note") or ""
        row["created_at"] = row.get("created_at") or created
        if not row.get("id"):
            row["id"] = metric_id(row)


def validate_import_rows(all_rows: dict[str, list[dict[str, Any]]], existing_refs: dict[str, Any] | None = None) -> list[str]:
    errors: list[str] = []
    existing_refs = existing_refs or {}
    project_ids = {row["id"] for row in all_rows["projects"]} | existing_refs.get("projects", set())
    module_definition_ids = {row["id"] for row in all_rows["module_definitions"]} | existing_refs.get("module_definitions", set())
    impl_option_ids = {row["id"] for row in all_rows["implOptions"]} | existing_refs.get("implOptions", set())
    tier_ids = {row["id"] for row in all_rows["tiers"]} | existing_refs.get("tiers", set())
    tier_impl_option_ids = {row["id"]: row["impl_option_id"] for row in all_rows["tiers"]}
    tier_impl_option_ids.update(existing_refs.get("tier_implOptions", {}))
    component_ids = {row["id"] for row in all_rows["logical_components"]} | existing_refs.get("logical_components", set())
    partition_ids = {row["id"] for row in all_rows["physical_partitions"]} | existing_refs.get("physical_partitions", set())

    for row in all_rows["implOptions"]:
        if row["project_id"] not in project_ids:
            errors.append(f"impl_option {row['id']} references missing project_id {row['project_id']}")
    for row in all_rows["logical_components"]:
        if row["instance_type"] == "parent_residual":
            errors.append(f"logical_component {row['id']} uses parent_residual; residual/self area is computed from parent total metrics minus child metrics")
        if row["project_id"] not in project_ids:
            errors.append(f"logical_component {row['id']} references missing project_id {row['project_id']}")
        if row.get("parent_id") and row["parent_id"] not in component_ids:
            errors.append(f"logical_component {row['id']} references missing parent_id {row['parent_id']}")
        if row.get("module_definition_id") and row["module_definition_id"] not in module_definition_ids:
            errors.append(f"logical_component {row['id']} references missing module_definition_id {row['module_definition_id']}")
    for row in all_rows["tiers"]:
        if row["impl_option_id"] not in impl_option_ids:
            errors.append(f"tier {row['id']} references missing impl_option_id {row['impl_option_id']}")
    for row in all_rows["physical_partitions"]:
        if row["impl_option_id"] not in impl_option_ids:
            errors.append(f"physical_partition {row['id']} references missing impl_option_id {row['impl_option_id']}")
        if row["logical_component_id"] not in component_ids:
            errors.append(f"physical_partition {row['id']} references missing logical_component_id {row['logical_component_id']}")
        if row["tier_id"] not in tier_ids:
            errors.append(f"physical_partition {row['id']} references missing tier_id {row['tier_id']}")
        elif tier_impl_option_ids.get(row["tier_id"]) != row["impl_option_id"]:
            errors.append(f"physical_partition {row['id']} tier_id {row['tier_id']} belongs to impl_option {tier_impl_option_ids.get(row['tier_id'])}, not {row['impl_option_id']}")
        if row["partition_type"] not in ALLOWED_PARTITION_TYPES:
            errors.append(f"physical_partition {row['id']} uses unsupported partition_type {row['partition_type']}")
        if row["resource_category"] not in ALLOWED_PARTITION_RESOURCE_CATEGORIES:
            errors.append(f"physical_partition {row['id']} uses unsupported resource_category {row['resource_category']}")
        if row["physical_instance_count"] < 0:
            errors.append(f"physical_partition {row['id']} has negative physical_instance_count")
        if row["content_share"] < 0:
            errors.append(f"physical_partition {row['id']} has negative content_share")
    for row in all_rows["metrics"]:
        if row["impl_option_id"] not in impl_option_ids:
            errors.append(f"metric {row['id']} references missing impl_option_id {row['impl_option_id']}")
        if row["subject_type"] not in ALLOWED_SUBJECT_TYPES:
            errors.append(f"metric {row['id']} uses unsupported subject_type {row['subject_type']}")
        if row["subject_type"] == "logical_component" and row["subject_id"] not in component_ids:
            errors.append(f"metric {row['id']} references missing logical_component subject_id {row['subject_id']}")
        if row["subject_type"] == "physical_partition" and row["subject_id"] not in partition_ids:
            errors.append(f"metric {row['id']} references missing physical_partition subject_id {row['subject_id']}")
        if row["subject_type"] == "tier" and row["subject_id"] not in tier_ids:
            errors.append(f"metric {row['id']} references missing tier subject_id {row['subject_id']}")
        if row["subject_type"] == "impl_option" and row["subject_id"] not in impl_option_ids:
            errors.append(f"metric {row['id']} references missing impl_option subject_id {row['subject_id']}")
        if row["value_type"] not in ALLOWED_VALUE_TYPES:
            errors.append(f"metric {row['id']} uses unsupported value_type {row['value_type']}")
        if row["confidence"] not in ALLOWED_CONFIDENCE:
            errors.append(f"metric {row['id']} uses unsupported confidence {row['confidence']}")
        if row["value_type"] == "number":
            try:
                float(row["metric_value"])
            except (TypeError, ValueError):
                errors.append(f"metric {row['id']} has non-numeric metric_value {row['metric_value']}")
    return errors


def existing_reference_ids(session: Session) -> dict[str, Any]:
    tiers = session.exec(select(Tier)).all()
    return {
        "projects": {row.id for row in session.exec(select(Project)).all()},
        "module_definitions": {row.id for row in session.exec(select(ModuleDefinition)).all()},
        "implOptions": {row.id for row in session.exec(select(ImplOption)).all()},
        "tiers": {row.id for row in tiers},
        "tier_implOptions": {row.id: row.impl_option_id for row in tiers},
        "logical_components": {row.id for row in session.exec(select(LogicalComponent)).all()},
        "physical_partitions": {row.id for row in session.exec(select(PhysicalPartition)).all()},
    }


def validate_team_import_scope(all_rows: dict[str, list[dict[str, Any]]], session: Session, team: str | None, impl_option_id: str = "S2") -> list[str]:
    if is_global_team(team):
        return []

    errors: list[str] = []
    allowed_component_ids = allowed_component_ids_for_team(session, team, impl_option_id) or set()
    if not allowed_component_ids:
        return [f"team {team} has no assigned component scope in impl_option {impl_option_id}"]

    existing_partitions = session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()
    allowed_partition_ids = {row.id for row in existing_partitions if row.logical_component_id in allowed_component_ids}
    workbook_partition_ids = {row["id"] for row in all_rows["physical_partitions"] if row["logical_component_id"] in allowed_component_ids}
    allowed_partition_ids |= workbook_partition_ids

    immutable_logical_fields = {
        "project_id",
        "parent_id",
        "module_definition_id",
        "name",
        "instance_type",
        "resource_type",
        "function_domain",
        "hierarchy_path",
    }
    for row in all_rows["logical_components"]:
        if row["id"] not in allowed_component_ids:
            errors.append(f"logical_component {row['id']} is outside team scope {team}")
            continue
        existing = session.get(LogicalComponent, row["id"])
        if existing:
            for field in immutable_logical_fields:
                if (row.get(field) or None) != (getattr(existing, field) or None):
                    errors.append(f"logical_component {row['id']} cannot change structural field {field} in a team workbook")

    for row in all_rows["physical_partitions"]:
        if row["impl_option_id"] != impl_option_id:
            errors.append(f"physical_partition {row['id']} uses impl_option_id {row['impl_option_id']}, expected {impl_option_id}")
        if row["logical_component_id"] not in allowed_component_ids:
            errors.append(f"physical_partition {row['id']} maps outside team scope {team}")

    for row in all_rows["metrics"]:
        if row["impl_option_id"] != impl_option_id:
            errors.append(f"metric {row['id']} uses impl_option_id {row['impl_option_id']}, expected {impl_option_id}")
        if row["subject_type"] == "logical_component" and row["subject_id"] not in allowed_component_ids:
            errors.append(f"metric {row['id']} references logical_component outside team scope {team}")
        elif row["subject_type"] == "physical_partition" and row["subject_id"] not in allowed_partition_ids:
            errors.append(f"metric {row['id']} references physical_partition outside team scope {team}")
        elif row["subject_type"] in {"tier", "impl_option"}:
            errors.append(f"metric {row['id']} uses shared subject_type {row['subject_type']}; team workbooks may only update logical_component or physical_partition metrics")

    return errors


def row_dict(row: SQLModel, columns: list[str]) -> dict[str, Any]:
    return {column: getattr(row, column, None) for column in columns}


def write_import_sheet(workbook: Workbook, sheet_name: str, columns: list[str], rows: list[dict[str, Any]], editable: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column) for column in columns])

    header_fill = PatternFill("solid", fgColor="0F172A" if editable else "334155")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        max_len = max([len(str(column))] + [len(str(row.get(column) or "")) for row in rows[:100]])
        sheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), 36)

    if sheet_name == "metrics":
        validations = {
            "subject_type": ["logical_component", "physical_partition"],
            "value_type": sorted(ALLOWED_VALUE_TYPES),
            "corner": ["typical", "best", "worst"],
            "workload": ["nominal", "peak", "idle"],
            "confidence": sorted(ALLOWED_CONFIDENCE),
        }
        for column_name, values in validations.items():
            if column_name not in columns:
                continue
            column_letter = get_column_letter(columns.index(column_name) + 1)
            validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}500")

    if sheet_name == "physical_partitions":
        validations = {
            "resource_category": sorted(ALLOWED_PARTITION_RESOURCE_CATEGORIES),
            "partition_type": sorted(ALLOWED_PARTITION_TYPES),
        }
        for column_name, values in validations.items():
            if column_name not in columns:
                continue
            column_letter = get_column_letter(columns.index(column_name) + 1)
            validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}500")


def build_team_import_workbook(session: Session, team: str, impl_option_id: str = "S2") -> str:
    allowed_component_ids = allowed_component_ids_for_team(session, team, impl_option_id)
    if allowed_component_ids is None:
        raise HTTPException(status_code=400, detail="Team template is only generated for scoped teams. Use the full template for Architecture Team.")
    if not allowed_component_ids:
        raise HTTPException(status_code=404, detail=f"No component scope found for team {team}.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = f"SoC team import workbook - {team}"

    scope_sheet = workbook.create_sheet("responsibility_scope")
    scope_sheet.append(["field", "value"])
    scope_sheet.append(["team", team])
    scope_sheet.append(["impl_option_id", impl_option_id])
    scope_sheet.append(["editable_sheets", "logical_components, physical_partitions, metrics"])
    scope_sheet.append(["rule", "Do not edit rows outside this workbook. Shared reference sheets are context only."])
    for cell in scope_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(color="FFFFFF", bold=True)
    scope_sheet.column_dimensions["A"].width = 22
    scope_sheet.column_dimensions["B"].width = 90

    projects = [row_dict(row, IMPORT_SHEETS["projects"][1]) for row in session.exec(select(Project)).all()]
    implOptions = [row_dict(row, IMPORT_SHEETS["implOptions"][1]) for row in session.exec(select(ImplOption)).all()]
    tiers = [row_dict(row, IMPORT_SHEETS["tiers"][1]) for row in session.exec(select(Tier).where(Tier.impl_option_id == impl_option_id).order_by(Tier.tier_index)).all()]

    components = session.exec(select(LogicalComponent).order_by(LogicalComponent.hierarchy_path)).all()
    scoped_components = [row for row in components if row.id in allowed_component_ids]
    module_definition_ids = {row.module_definition_id for row in scoped_components if row.module_definition_id}
    module_definitions = [
        row_dict(row, IMPORT_SHEETS["module_definitions"][1])
        for row in session.exec(select(ModuleDefinition)).all()
        if row.id in module_definition_ids
    ]
    logical_components = [row_dict(row, IMPORT_SHEETS["logical_components"][1]) for row in scoped_components]

    partitions = [
        row
        for row in session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()
        if row.logical_component_id in allowed_component_ids
    ]
    physical_partitions = [row_dict(row, IMPORT_SHEETS["physical_partitions"][1]) for row in partitions]
    partition_ids = {row.id for row in partitions}

    metrics = [
        row
        for row in session.exec(select(Metric).where(Metric.impl_option_id == impl_option_id)).all()
        if (row.subject_type == "logical_component" and row.subject_id in allowed_component_ids)
        or (row.subject_type == "physical_partition" and row.subject_id in partition_ids)
    ]
    metric_rows = [row_dict(row, IMPORT_SHEETS["metrics"][1]) for row in metrics]

    sheet_rows = {
        "module_definitions": module_definitions,
        "projects": projects,
        "implOptions": implOptions,
        "tiers": tiers,
        "logical_components": logical_components,
        "physical_partitions": physical_partitions,
        "metrics": metric_rows,
    }
    editable_sheets = {"logical_components", "physical_partitions", "metrics"}
    for sheet_name, (_, columns, _) in IMPORT_SHEETS.items():
        write_import_sheet(workbook, sheet_name, columns, sheet_rows[sheet_name], sheet_name in editable_sheets)

    temp_file = NamedTemporaryFile(prefix=f"soc_{team.lower().replace(' ', '_')}_", suffix=".xlsx", delete=False)
    temp_file.close()
    workbook.save(temp_file.name)
    return temp_file.name


@app.post("/api/import/excel")
async def import_excel(file: UploadFile = File(...), team: str | None = None, impl_option_id: str = "S2") -> dict[str, Any]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    workbook_bytes = await file.read()
    all_rows: dict[str, list[dict[str, Any]]] = {}
    for sheet_name, (_, columns, required) in IMPORT_SHEETS.items():
        with SpooledTemporaryFile() as temp_file:
            temp_file.write(workbook_bytes)
            temp_file.seek(0)
            all_rows[sheet_name] = read_sheet_rows(temp_file, sheet_name, columns, required)
    prepare_import_rows(all_rows)

    imported: dict[str, int] = {}
    with Session(engine) as session:
        existing_refs = existing_reference_ids(session) if not is_global_team(team) else None
        errors = validate_import_rows(all_rows, existing_refs)
        errors.extend(validate_team_import_scope(all_rows, session, team, impl_option_id))
        if errors:
            raise HTTPException(status_code=400, detail={"errors": errors})

        editable_sheets = set(IMPORT_SHEETS) if is_global_team(team) else {"logical_components", "physical_partitions", "metrics"}
        for sheet_name, (model, _, _) in IMPORT_SHEETS.items():
            count = 0
            if sheet_name not in editable_sheets:
                imported[sheet_name] = count
                continue
            for row in all_rows[sheet_name]:
                session.merge(model(**row))
                count += 1
            imported[sheet_name] = count
        session.commit()
    return {"filename": file.filename, "imported": imported, "errors": []}


@app.get("/api/import/template")
def get_import_template(background_tasks: BackgroundTasks, team: str | None = None, impl_option_id: str = "S2") -> FileResponse:
    if not is_global_team(team):
        with Session(engine) as session:
            path = build_team_import_workbook(session, team or "", impl_option_id)
        background_tasks.add_task(os.remove, path)
        safe_team = (team or "team").lower().replace(" ", "_")
        return FileResponse(
            path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"soc_team_import_{safe_team}_{impl_option_id}.xlsx",
        )

    if not TEMPLATE_PATH.exists():
        raise HTTPException(status_code=404, detail="Import template has not been generated.")
    return FileResponse(
        TEMPLATE_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="soc_import_template.xlsx",
    )
