from __future__ import annotations

from datetime import datetime
import os
from tempfile import NamedTemporaryFile, SpooledTemporaryFile
from typing import Any

from fastapi import BackgroundTasks, FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlmodel import SQLModel, Session, select

from backend import db
from backend.models import (
    ImplOption,
    LogicalComponent,
    Metric,
    ModuleDefinition,
    PhysicalPartition,
    Project,
    ResponsibilityAssignment,
    Tier,
)


def metric_id(row: dict[str, Any]) -> str:
    return (
        f"{row['impl_option_id']}-{row['subject_type']}-{row['subject_id']}-"
        f"{row['metric_name']}-{row['corner']}-{row['workload']}"
    )


def normalized_content_share(partition_type: str, value: float | None) -> float:
    if partition_type == "full":
        return 1.0
    return float(value if value is not None else 1.0)


def normalized_resource_category(value: str | None) -> str:
    return value if value in ALLOWED_PARTITION_RESOURCE_CATEGORIES else "block"


def is_global_team(team: str | None) -> bool:
    return team in {None, "", "Architecture Team", "All Teams"}


def descendant_component_ids(session: Session, component_id: str) -> set[str]:
    components = session.exec(select(LogicalComponent)).all()
    children: dict[str | None, list[str]] = {}
    for component in components:
        children.setdefault(component.parent_id, []).append(component.id)
    result: set[str] = set()

    def walk(current_id: str) -> None:
        for child_id in children.get(current_id, []):
            if child_id in result:
                continue
            result.add(child_id)
            walk(child_id)

    walk(component_id)
    return result


def allowed_component_ids_for_team(session: Session, team: str | None, impl_option_id: str = "S2") -> set[str] | None:
    if is_global_team(team):
        return None
    assignments = session.exec(
        select(ResponsibilityAssignment).where(
            ResponsibilityAssignment.team_name == team,
            ResponsibilityAssignment.impl_option_id == impl_option_id,
            ResponsibilityAssignment.can_read == True,
        )
    ).all()
    allowed: set[str] = set()
    for assignment in assignments:
        allowed.add(assignment.logical_component_id)
        if assignment.scope_type == "subtree":
            allowed.update(descendant_component_ids(session, assignment.logical_component_id))
    return allowed


IMPORT_SHEETS: dict[str, tuple[type[SQLModel], list[str], set[str]]] = {
    "module_definitions": (
        ModuleDefinition,
        ["id", "name", "module_type", "ip_owner", "reuse_class", "description"],
        {"id", "name", "module_type"},
    ),
    "projects": (
        Project,
        ["id", "name", "product_family", "generation", "owner", "phase"],
        {"id", "name", "product_family", "generation", "owner", "phase"},
    ),
    "implOptions": (
        ImplOption,
        ["id", "project_id", "name", "impl_type", "process_combo", "status"],
        {"id", "project_id", "name", "impl_type", "process_combo", "status"},
    ),
    "tiers": (
        Tier,
        ["id", "impl_option_id", "tier_index", "name", "process_id", "role", "orientation", "area_limit_mm2"],
        {"id", "impl_option_id", "tier_index", "name", "process_id", "role"},
    ),
    "logical_components": (
        LogicalComponent,
        ["id", "project_id", "parent_id", "module_definition_id", "name", "instance_type", "resource_type", "function_domain", "hierarchy_path", "logical_instance_count", "description"],
        {"id", "project_id", "name", "instance_type", "resource_type", "function_domain", "hierarchy_path", "logical_instance_count"},
    ),
    "physical_partitions": (
        PhysicalPartition,
        ["id", "impl_option_id", "logical_component_id", "tier_id", "partition_name", "resource_category", "partition_type", "physical_instance_count", "content_share", "description"],
        {"id", "impl_option_id", "logical_component_id", "tier_id", "partition_name", "partition_type", "physical_instance_count"},
    ),
    "metrics": (
        Metric,
        ["id", "impl_option_id", "subject_type", "subject_id", "metric_name", "metric_value", "metric_unit", "metric_category", "value_type", "corner", "workload", "confidence", "source_note", "created_at"],
        {"impl_option_id", "subject_type", "subject_id", "metric_name", "metric_value", "value_type", "corner", "workload", "confidence"},
    ),
}

ALLOWED_SUBJECT_TYPES = {"logical_component", "physical_partition", "tier", "impl_option"}
ALLOWED_VALUE_TYPES = {"number", "text", "boolean"}
ALLOWED_CONFIDENCE = {"approved", "review", "draft"}
ALLOWED_PARTITION_TYPES = {"full", "partial"}
ALLOWED_PARTITION_RESOURCE_CATEGORIES = {"logic", "sram", "block"}
METRIC_IDENTITY_FIELDS = ("impl_option_id", "subject_type", "subject_id", "metric_name", "corner", "workload")
LEGACY_REDUNDANT_METRIC_IDS = {
    "M_IMPL_OPTION_AREA",
    "M_PART_GPU_LOGIC_AREA_TOP",
    "M_PART_GPU_LOGIC_AREA_MID",
}


def normalize_cell(value: Any) -> Any:
    if value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def read_sheet_rows(workbook_file: SpooledTemporaryFile[bytes], sheet_name: str, expected_columns: list[str], required_columns: set[str]) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_file, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Missing sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    header = [str(cell.value or "").strip() for cell in sheet[1]]
    aliases = {"content_share": "partition_ratio"} if sheet_name == "physical_partitions" else {}
    missing = [column for column in required_columns if column not in header and aliases.get(column) not in header]
    if missing:
        raise HTTPException(status_code=400, detail=f"Sheet {sheet_name} is missing columns: {', '.join(missing)}")
    indexes = {
        column: header.index(column if column in header else aliases[column])
        for column in expected_columns
        if column in header or aliases.get(column) in header
    }
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {column: normalize_cell(row[indexes[column]]) if column in indexes else None for column in expected_columns}
        if all(value is None for value in record.values()):
            continue
        missing_required = [column for column in required_columns if record.get(column) is None]
        if missing_required:
            raise HTTPException(status_code=400, detail=f"Sheet {sheet_name} row {row_index} missing required columns: {', '.join(missing_required)}")
        rows.append(record)
    return rows


def prepare_import_rows(all_rows: dict[str, list[dict[str, Any]]]) -> None:
    created = db.now_iso()
    for row in all_rows["projects"]:
        row.setdefault("description", "")
        row["created_at"] = row.get("created_at") or created
        row["updated_at"] = row.get("updated_at") or created
    for row in all_rows["implOptions"]:
        row.setdefault("description", "")
        row["created_at"] = row.get("created_at") or created
        row["updated_at"] = row.get("updated_at") or created
    for row in all_rows["tiers"]:
        row["thickness_um"] = row.get("thickness_um") or 0
        row.setdefault("description", "")
    for row in all_rows["logical_components"]:
        row["parent_id"] = row.get("parent_id") or None
        row["module_definition_id"] = row.get("module_definition_id") or None
        row["logical_instance_count"] = int(row["logical_instance_count"])
        row["owner_team"] = row.get("owner_team") or "Architecture Team"
        row["visibility_level"] = row.get("visibility_level") or "team"
        row["created_at"] = row.get("created_at") or created
        row["updated_at"] = row.get("updated_at") or created
    for row in all_rows["physical_partitions"]:
        row["physical_instance_count"] = int(row["physical_instance_count"])
        row["resource_category"] = normalized_resource_category(row.get("resource_category"))
        raw_content_share = row.get("content_share")
        row["content_share"] = normalized_content_share(row["partition_type"], float(raw_content_share) if raw_content_share is not None else None)
        row["partition_ratio"] = row["content_share"]
    for row in all_rows["metrics"]:
        row["metric_value"] = str(row["metric_value"])
        row["metric_unit"] = row.get("metric_unit") or ""
        row["metric_category"] = row.get("metric_category") or ""
        row["source_note"] = row.get("source_note") or ""
        row["created_at"] = row.get("created_at") or created
        if not row.get("id"):
            row["id"] = metric_id(row)


def drop_redundant_legacy_metric_rows(all_rows: dict[str, list[dict[str, Any]]], existing_refs: dict[str, Any]) -> None:
    existing_metric_identities = existing_refs.get("metric_identities", {})
    workbook_canonical_ids: dict[tuple[str, str, str, str, str, str], set[str]] = {}
    for row in all_rows["metrics"]:
        identity = tuple(str(row[field]) for field in METRIC_IDENTITY_FIELDS)
        if row["id"] not in LEGACY_REDUNDANT_METRIC_IDS:
            workbook_canonical_ids.setdefault(identity, set()).add(row["id"])

    filtered_metrics: list[dict[str, Any]] = []
    for row in all_rows["metrics"]:
        if row["id"] not in LEGACY_REDUNDANT_METRIC_IDS:
            filtered_metrics.append(row)
            continue

        identity = tuple(str(row[field]) for field in METRIC_IDENTITY_FIELDS)
        existing_id = existing_metric_identities.get(identity)
        has_canonical = (existing_id and existing_id != row["id"]) or bool(workbook_canonical_ids.get(identity))
        if not has_canonical:
            filtered_metrics.append(row)

    all_rows["metrics"] = filtered_metrics


def validate_import_rows(all_rows: dict[str, list[dict[str, Any]]], existing_refs: dict[str, Any] | None = None) -> list[str]:
    errors: list[str] = []
    existing_refs = existing_refs or {}
    project_ids = {row["id"] for row in all_rows["projects"]} | existing_refs.get("projects", set())
    module_definition_ids = {row["id"] for row in all_rows["module_definitions"]} | existing_refs.get("module_definitions", set())
    impl_option_ids = {row["id"] for row in all_rows["implOptions"]} | existing_refs.get("implOptions", set())
    tier_ids = {row["id"] for row in all_rows["tiers"]} | existing_refs.get("tiers", set())
    tier_impl_option_ids = {row["id"]: row["impl_option_id"] for row in all_rows["tiers"]}
    tier_impl_option_ids.update(existing_refs.get("tier_implOptions", {}))
    component_ids = {row["id"] for row in all_rows["logical_components"]} | existing_refs.get("logical_components", set())
    partition_ids = {row["id"] for row in all_rows["physical_partitions"]} | existing_refs.get("physical_partitions", set())
    existing_component_paths = existing_refs.get("logical_component_paths", {})
    existing_metric_identities = existing_refs.get("metric_identities", {})
    workbook_component_paths: dict[tuple[str, str], list[str]] = {}
    for row in all_rows["logical_components"]:
        workbook_component_paths.setdefault((row["project_id"], row["hierarchy_path"]), []).append(row["id"])
    for (project_id, hierarchy_path), ids in workbook_component_paths.items():
        if len(set(ids)) > 1:
            errors.append(
                f"logical_components {', '.join(ids)} duplicate hierarchy_path {hierarchy_path} in project {project_id}; use logical_instance_count for repeated instances"
            )
        existing_id = existing_component_paths.get((project_id, hierarchy_path))
        if existing_id and existing_id not in ids:
            errors.append(
                f"logical_component hierarchy_path {hierarchy_path} in project {project_id} already belongs to {existing_id}; cannot import duplicate path for {', '.join(ids)}"
            )

    workbook_metric_identities: dict[tuple[str, str, str, str, str, str], list[str]] = {}
    for row in all_rows["metrics"]:
        identity = tuple(str(row[field]) for field in METRIC_IDENTITY_FIELDS)
        workbook_metric_identities.setdefault(identity, []).append(row["id"])
    for identity, ids in workbook_metric_identities.items():
        distinct_ids = sorted(set(ids))
        if len(distinct_ids) > 1:
            errors.append(
                f"metrics {', '.join(distinct_ids)} duplicate metric identity {'/'.join(identity)}; keep one row per impl_option, subject, metric_name, corner, and workload"
            )
        existing_id = existing_metric_identities.get(identity)
        if existing_id and existing_id not in distinct_ids:
            errors.append(
                f"metric identity {'/'.join(identity)} already belongs to {existing_id}; cannot import duplicate identity for {', '.join(distinct_ids)}"
            )

    for row in all_rows["implOptions"]:
        if row["project_id"] not in project_ids:
            errors.append(f"impl_option {row['id']} references missing project_id {row['project_id']}")
    for row in all_rows["logical_components"]:
        if row["instance_type"] == "parent_residual":
            errors.append(f"logical_component {row['id']} uses parent_residual; residual/self area is computed from parent total metrics minus child metrics")
        if row["project_id"] not in project_ids:
            errors.append(f"logical_component {row['id']} references missing project_id {row['project_id']}")
        if row.get("parent_id") and row["parent_id"] not in component_ids:
            errors.append(f"logical_component {row['id']} references missing parent_id {row['parent_id']}")
        if row.get("module_definition_id") and row["module_definition_id"] not in module_definition_ids:
            errors.append(f"logical_component {row['id']} references missing module_definition_id {row['module_definition_id']}")
    for row in all_rows["tiers"]:
        if row["impl_option_id"] not in impl_option_ids:
            errors.append(f"tier {row['id']} references missing impl_option_id {row['impl_option_id']}")
    for row in all_rows["physical_partitions"]:
        if row["impl_option_id"] not in impl_option_ids:
            errors.append(f"physical_partition {row['id']} references missing impl_option_id {row['impl_option_id']}")
        if row["logical_component_id"] not in component_ids:
            errors.append(f"physical_partition {row['id']} references missing logical_component_id {row['logical_component_id']}")
        if row["tier_id"] not in tier_ids:
            errors.append(f"physical_partition {row['id']} references missing tier_id {row['tier_id']}")
        elif tier_impl_option_ids.get(row["tier_id"]) != row["impl_option_id"]:
            errors.append(f"physical_partition {row['id']} tier_id {row['tier_id']} belongs to impl_option {tier_impl_option_ids.get(row['tier_id'])}, not {row['impl_option_id']}")
        if row["partition_type"] not in ALLOWED_PARTITION_TYPES:
            errors.append(f"physical_partition {row['id']} uses unsupported partition_type {row['partition_type']}")
        if row["resource_category"] not in ALLOWED_PARTITION_RESOURCE_CATEGORIES:
            errors.append(f"physical_partition {row['id']} uses unsupported resource_category {row['resource_category']}")
        if row["physical_instance_count"] < 0:
            errors.append(f"physical_partition {row['id']} has negative physical_instance_count")
        if row["content_share"] < 0:
            errors.append(f"physical_partition {row['id']} has negative content_share")
    for row in all_rows["metrics"]:
        if row["impl_option_id"] not in impl_option_ids:
            errors.append(f"metric {row['id']} references missing impl_option_id {row['impl_option_id']}")
        if row["subject_type"] not in ALLOWED_SUBJECT_TYPES:
            errors.append(f"metric {row['id']} uses unsupported subject_type {row['subject_type']}")
        if row["subject_type"] == "logical_component" and row["subject_id"] not in component_ids:
            errors.append(f"metric {row['id']} references missing logical_component subject_id {row['subject_id']}")
        if row["subject_type"] == "physical_partition" and row["subject_id"] not in partition_ids:
            errors.append(f"metric {row['id']} references missing physical_partition subject_id {row['subject_id']}")
        if row["subject_type"] == "tier" and row["subject_id"] not in tier_ids:
            errors.append(f"metric {row['id']} references missing tier subject_id {row['subject_id']}")
        if row["subject_type"] == "impl_option" and row["subject_id"] not in impl_option_ids:
            errors.append(f"metric {row['id']} references missing impl_option subject_id {row['subject_id']}")
        if row["value_type"] not in ALLOWED_VALUE_TYPES:
            errors.append(f"metric {row['id']} uses unsupported value_type {row['value_type']}")
        if row["confidence"] not in ALLOWED_CONFIDENCE:
            errors.append(f"metric {row['id']} uses unsupported confidence {row['confidence']}")
        if row["value_type"] == "number":
            try:
                float(row["metric_value"])
            except (TypeError, ValueError):
                errors.append(f"metric {row['id']} has non-numeric metric_value {row['metric_value']}")
    return errors


def existing_reference_ids(session: Session) -> dict[str, Any]:
    tiers = session.exec(select(Tier)).all()
    components = session.exec(select(LogicalComponent)).all()
    metrics = session.exec(select(Metric)).all()
    return {
        "projects": {row.id for row in session.exec(select(Project)).all()},
        "module_definitions": {row.id for row in session.exec(select(ModuleDefinition)).all()},
        "implOptions": {row.id for row in session.exec(select(ImplOption)).all()},
        "tiers": {row.id for row in tiers},
        "tier_implOptions": {row.id: row.impl_option_id for row in tiers},
        "logical_components": {row.id for row in components},
        "logical_component_paths": {(row.project_id, row.hierarchy_path): row.id for row in components},
        "metric_identities": {
            (row.impl_option_id, row.subject_type, row.subject_id, row.metric_name, row.corner, row.workload): row.id for row in metrics
        },
        "physical_partitions": {row.id for row in session.exec(select(PhysicalPartition)).all()},
    }


def validate_team_import_scope(all_rows: dict[str, list[dict[str, Any]]], session: Session, team: str | None, impl_option_id: str = "S2") -> list[str]:
    if is_global_team(team):
        return []

    errors: list[str] = []
    allowed_component_ids = allowed_component_ids_for_team(session, team, impl_option_id) or set()
    if not allowed_component_ids:
        return [f"team {team} has no assigned component scope in impl_option {impl_option_id}"]

    existing_partitions = session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()
    allowed_partition_ids = {row.id for row in existing_partitions if row.logical_component_id in allowed_component_ids}
    workbook_partition_ids = {row["id"] for row in all_rows["physical_partitions"] if row["logical_component_id"] in allowed_component_ids}
    allowed_partition_ids |= workbook_partition_ids

    immutable_logical_fields = {
        "project_id",
        "parent_id",
        "module_definition_id",
        "name",
        "instance_type",
        "resource_type",
        "function_domain",
        "hierarchy_path",
    }
    for row in all_rows["logical_components"]:
        if row["id"] not in allowed_component_ids:
            errors.append(f"logical_component {row['id']} is outside team scope {team}")
            continue
        existing = session.get(LogicalComponent, row["id"])
        if existing:
            for field in immutable_logical_fields:
                if (row.get(field) or None) != (getattr(existing, field) or None):
                    errors.append(f"logical_component {row['id']} cannot change structural field {field} in a team workbook")

    for row in all_rows["physical_partitions"]:
        if row["impl_option_id"] != impl_option_id:
            errors.append(f"physical_partition {row['id']} uses impl_option_id {row['impl_option_id']}, expected {impl_option_id}")
        if row["logical_component_id"] not in allowed_component_ids:
            errors.append(f"physical_partition {row['id']} maps outside team scope {team}")

    for row in all_rows["metrics"]:
        if row["impl_option_id"] != impl_option_id:
            errors.append(f"metric {row['id']} uses impl_option_id {row['impl_option_id']}, expected {impl_option_id}")
        if row["subject_type"] == "logical_component" and row["subject_id"] not in allowed_component_ids:
            errors.append(f"metric {row['id']} references logical_component outside team scope {team}")
        elif row["subject_type"] == "physical_partition" and row["subject_id"] not in allowed_partition_ids:
            errors.append(f"metric {row['id']} references physical_partition outside team scope {team}")
        elif row["subject_type"] in {"tier", "impl_option"}:
            errors.append(f"metric {row['id']} uses shared subject_type {row['subject_type']}; team workbooks may only update logical_component or physical_partition metrics")

    return errors


def row_dict(row: SQLModel, columns: list[str]) -> dict[str, Any]:
    return {column: getattr(row, column, None) for column in columns}


def write_import_sheet(workbook: Workbook, sheet_name: str, columns: list[str], rows: list[dict[str, Any]], editable: bool = True) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column) for column in columns])

    header_fill = PatternFill("solid", fgColor="0F172A" if editable else "334155")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        max_len = max([len(str(column))] + [len(str(row.get(column) or "")) for row in rows[:100]])
        sheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), 36)

    if sheet_name == "metrics":
        validations = {
            "subject_type": ["logical_component", "physical_partition"],
            "value_type": sorted(ALLOWED_VALUE_TYPES),
            "corner": ["typical", "best", "worst"],
            "workload": ["nominal", "peak", "idle"],
            "confidence": sorted(ALLOWED_CONFIDENCE),
        }
        for column_name, values in validations.items():
            if column_name not in columns:
                continue
            column_letter = get_column_letter(columns.index(column_name) + 1)
            validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}500")

    if sheet_name == "physical_partitions":
        validations = {
            "resource_category": sorted(ALLOWED_PARTITION_RESOURCE_CATEGORIES),
            "partition_type": sorted(ALLOWED_PARTITION_TYPES),
        }
        for column_name, values in validations.items():
            if column_name not in columns:
                continue
            column_letter = get_column_letter(columns.index(column_name) + 1)
            validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}500")


def build_team_import_workbook(session: Session, team: str, impl_option_id: str = "S2") -> str:
    allowed_component_ids = allowed_component_ids_for_team(session, team, impl_option_id)
    if allowed_component_ids is None:
        raise HTTPException(status_code=400, detail="Team template is only generated for scoped teams. Use the full template for Architecture Team.")
    if not allowed_component_ids:
        raise HTTPException(status_code=404, detail=f"No component scope found for team {team}.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = f"SoC team import workbook - {team}"

    scope_sheet = workbook.create_sheet("responsibility_scope")
    scope_sheet.append(["field", "value"])
    scope_sheet.append(["team", team])
    scope_sheet.append(["impl_option_id", impl_option_id])
    scope_sheet.append(["editable_sheets", "logical_components, physical_partitions, metrics"])
    scope_sheet.append(["rule", "Do not edit rows outside this workbook. Shared reference sheets are context only."])
    for cell in scope_sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.font = Font(color="FFFFFF", bold=True)
    scope_sheet.column_dimensions["A"].width = 22
    scope_sheet.column_dimensions["B"].width = 90

    projects = [row_dict(row, IMPORT_SHEETS["projects"][1]) for row in session.exec(select(Project)).all()]
    implOptions = [row_dict(row, IMPORT_SHEETS["implOptions"][1]) for row in session.exec(select(ImplOption)).all()]
    tiers = [row_dict(row, IMPORT_SHEETS["tiers"][1]) for row in session.exec(select(Tier).where(Tier.impl_option_id == impl_option_id).order_by(Tier.tier_index)).all()]

    components = session.exec(select(LogicalComponent).order_by(LogicalComponent.hierarchy_path)).all()
    scoped_components = [row for row in components if row.id in allowed_component_ids]
    module_definition_ids = {row.module_definition_id for row in scoped_components if row.module_definition_id}
    module_definitions = [
        row_dict(row, IMPORT_SHEETS["module_definitions"][1])
        for row in session.exec(select(ModuleDefinition)).all()
        if row.id in module_definition_ids
    ]
    logical_components = [row_dict(row, IMPORT_SHEETS["logical_components"][1]) for row in scoped_components]

    partitions = [
        row
        for row in session.exec(select(PhysicalPartition).where(PhysicalPartition.impl_option_id == impl_option_id)).all()
        if row.logical_component_id in allowed_component_ids
    ]
    physical_partitions = [row_dict(row, IMPORT_SHEETS["physical_partitions"][1]) for row in partitions]
    partition_ids = {row.id for row in partitions}

    metrics = [
        row
        for row in session.exec(select(Metric).where(Metric.impl_option_id == impl_option_id)).all()
        if (row.subject_type == "logical_component" and row.subject_id in allowed_component_ids)
        or (row.subject_type == "physical_partition" and row.subject_id in partition_ids)
    ]
    metric_rows = [row_dict(row, IMPORT_SHEETS["metrics"][1]) for row in metrics]

    sheet_rows = {
        "module_definitions": module_definitions,
        "projects": projects,
        "implOptions": implOptions,
        "tiers": tiers,
        "logical_components": logical_components,
        "physical_partitions": physical_partitions,
        "metrics": metric_rows,
    }
    editable_sheets = {"logical_components", "physical_partitions", "metrics"}
    for sheet_name, (_, columns, _) in IMPORT_SHEETS.items():
        write_import_sheet(workbook, sheet_name, columns, sheet_rows[sheet_name], sheet_name in editable_sheets)

    temp_file = NamedTemporaryFile(prefix=f"soc_{team.lower().replace(' ', '_')}_", suffix=".xlsx", delete=False)
    temp_file.close()
    workbook.save(temp_file.name)
    return temp_file.name


def register_import_routes(app: FastAPI) -> None:
    @app.post("/api/import/excel")
    async def import_excel(file: UploadFile = File(...), team: str | None = None, impl_option_id: str = "S2") -> dict[str, Any]:
        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

        workbook_bytes = await file.read()
        all_rows: dict[str, list[dict[str, Any]]] = {}
        for sheet_name, (_, columns, required) in IMPORT_SHEETS.items():
            with SpooledTemporaryFile() as temp_file:
                temp_file.write(workbook_bytes)
                temp_file.seek(0)
                all_rows[sheet_name] = read_sheet_rows(temp_file, sheet_name, columns, required)
        prepare_import_rows(all_rows)

        imported: dict[str, int] = {}
        with Session(db.engine) as session:
            existing_refs = existing_reference_ids(session)
            drop_redundant_legacy_metric_rows(all_rows, existing_refs)
            errors = validate_import_rows(all_rows, existing_refs)
            errors.extend(validate_team_import_scope(all_rows, session, team, impl_option_id))
            if errors:
                raise HTTPException(status_code=400, detail={"errors": errors})

            editable_sheets = set(IMPORT_SHEETS) if is_global_team(team) else {"logical_components", "physical_partitions", "metrics"}
            for sheet_name, (model, _, _) in IMPORT_SHEETS.items():
                count = 0
                if sheet_name not in editable_sheets:
                    imported[sheet_name] = count
                    continue
                for row in all_rows[sheet_name]:
                    session.merge(model(**row))
                    count += 1
                imported[sheet_name] = count
            session.commit()
        return {"filename": file.filename, "imported": imported, "errors": []}


    @app.get("/api/import/template")
    def get_import_template(background_tasks: BackgroundTasks, team: str | None = None, impl_option_id: str = "S2") -> FileResponse:
        if not is_global_team(team):
            with Session(db.engine) as session:
                path = build_team_import_workbook(session, team or "", impl_option_id)
            background_tasks.add_task(os.remove, path)
            safe_team = (team or "team").lower().replace(" ", "_")
            return FileResponse(
                path,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=f"soc_team_import_{safe_team}_{impl_option_id}.xlsx",
            )

        if not db.TEMPLATE_PATH.exists():
            raise HTTPException(status_code=404, detail="Import template has not been generated.")
        return FileResponse(
            db.TEMPLATE_PATH,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="soc_import_template.xlsx",
        )
