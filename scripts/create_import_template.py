from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "templates" / "soc_import_demo.xlsx"

METRIC_DICTIONARY = [
    ["metric_name", "default_unit", "metric_category", "recommended_workload", "description"],
    ["logicMTr", "MTr", "scale", "nominal", "Logic transistor scale in million transistors."],
    ["memoryMb", "Mb", "memory", "nominal", "Modeled SRAM/cache/memory macro capacity."],
    ["area", "mm2", "physical", "nominal", "Physical area estimate for the selected scenario."],
    ["power", "W", "power", "peak", "Power estimate for the selected workload."],
    ["frequency", "GHz", "performance", "peak", "Target or achieved operating frequency."],
    ["utilization", "%", "physical", "nominal", "Floorplan or tier utilization percentage."],
]

CONFIDENCE_VALUES = ["approved", "review", "draft"]
CORNER_VALUES = ["typical", "slow", "fast", "ss", "ff", "tt"]
WORKLOAD_VALUES = sorted({row[3] for row in METRIC_DICTIONARY[1:]} | {"idle", "nominal", "peak"})


def metric_rows() -> list[list[object]]:
    rows: list[list[object]] = [
        [
            "id",
            "scenario_id",
            "instance_id",
            "metric_name",
            "metric_value",
            "metric_unit",
            "metric_category",
            "corner",
            "workload",
            "confidence",
            "created_at",
        ]
    ]
    metric_values = {
        "B0": (0, 0, 74.6, 15.2),
        "B1": (1850, 24, 12.8, 4.1),
        "B2": (3100, 18, 18.7, 5.5),
        "B3": (2400, 64, 21.2, 3.8),
        "B4": (980, 12, 7.4, 1.4),
        "B5": (0, 0, 4.8, 0.9),
        "B6": (0, 0, 3.1, 0.6),
        "B7": (420, 2, 2.7, 0.95),
        "B8": (0, 48, 8.2, 0.8),
    }
    for instance_id, (logic, memory, area, power) in sorted(metric_values.items()):
        for name, value, unit, category, workload in [
            ("logicMTr", logic, "MTr", "scale", "nominal"),
            ("memoryMb", memory, "Mb", "memory", "nominal"),
            ("area", area, "mm2", "physical", "nominal"),
            ("power", power, "W", "power", "peak"),
        ]:
            excel_row = len(rows) + 1
            rows.append(
                [
                    f'=IF(OR(B{excel_row}="",C{excel_row}="",D{excel_row}="",H{excel_row}="",I{excel_row}=""),"",B{excel_row}&"-"&C{excel_row}&"-"&D{excel_row}&"-"&H{excel_row}&"-"&I{excel_row})',
                    "S2",
                    instance_id,
                    name,
                    value,
                    f'=IFERROR(VLOOKUP(D{excel_row},metric_dictionary!$A$2:$D$200,2,FALSE),"{unit}")',
                    f'=IFERROR(VLOOKUP(D{excel_row},metric_dictionary!$A$2:$D$200,3,FALSE),"{category}")',
                    "typical",
                    f'=IFERROR(VLOOKUP(D{excel_row},metric_dictionary!$A$2:$D$200,4,FALSE),"{workload}")',
                    "approved" if instance_id in {"B0", "B4", "B5", "B6"} else "review",
                    "2026-05-27",
                ]
            )
    return rows


SHEETS: dict[str, list[list[object]]] = {
    "instructions": [
        ["SoC Cross-Die Database Import Template"],
        ["Keep sheet names and column names unchanged. IDs are stable keys used for upsert imports."],
        [""],
        ["Sheet", "Purpose"],
        ["projects", "Project master data"],
        ["scenarios", "Architecture scenarios linked to projects"],
        ["process_nodes", "Foundry/process reference data"],
        ["components", "Component/block hierarchy. parent_id may be blank only for root nodes."],
        ["tiers", "3D stack tier definitions linked to scenarios and process_nodes"],
        ["component_metrics", "Scenario-specific area, power, memory, logic, and other metrics"],
        ["metric_dictionary", "Allowed metric names and default unit/category/workload mapping"],
        [""],
        ["Metric maintenance note"],
        ["In component_metrics, fill scenario_id, instance_id, metric_name, metric_value, corner, confidence, and created_at. The id/unit/category/workload cells are formula-assisted and the backend can also fill them if formulas are not cached."],
    ],
    "metric_dictionary": METRIC_DICTIONARY,
    "projects": [
        ["id", "name", "product_family", "generation", "owner", "phase", "description", "created_at", "updated_at"],
        ["P001", "Mobile SoC Gen-A", "Flagship Mobile SoC", "Gen-A", "Architecture Team", "Architecture Planning", "Phase-1 architecture planning baseline.", "2026-05-27", "2026-05-27"],
        ["P002", "Mobile SoC Gen-B", "Flagship Mobile SoC", "Gen-B", "Product + Architecture", "Pre-Study", "Next-generation pre-study project.", "2026-05-27", "2026-05-27"],
    ],
    "scenarios": [
        ["id", "project_id", "name", "scenario_type", "process_combo", "description", "status", "created_at", "updated_at"],
        ["S1", "P001", "2D Baseline", "1 die", "N5 monolithic", "Current 2D planning baseline for cross-generation comparison.", "Low", "2026-05-27", "2026-05-27"],
        ["S2", "P001", "3DIC Option A", "3 tiers W2W", "N5 + N7 + N7", "Top N5 logic, middle N7 logic/memory, bottom N7 IO/PHY/PDN.", "High", "2026-05-27", "2026-05-27"],
        ["S3", "P001", "Cost-Reduced Option", "2 tiers W2W", "N7 + N7", "More conservative 2-tier split with lower process cost.", "Medium", "2026-05-27", "2026-05-27"],
    ],
    "process_nodes": [
        ["id", "foundry", "node_name", "logic_density_mtr_per_mm2", "sram_density_mb_per_mm2", "voltage_nominal", "cost_factor", "maturity_level", "description"],
        ["PN5", "TSMC", "N5", 171.3, 1.35, 0.75, 1.4, "Production", "High-performance logic process."],
        ["PN7", "TSMC", "N7", 113.9, 1.0, 0.8, 1.0, "Mature", "Mature logic and memory-friendly process."],
    ],
    "components": [
        ["id", "project_id", "parent_id", "name", "instance_type", "resource_type", "function_domain", "hierarchy_path", "description"],
        ["B0", "P001", "", "SOC_TOP", "top", "mixed", "SoC", "SOC_TOP", "tier=Split; confidence=approved"],
        ["B1", "P001", "B0", "CPU_CLUSTER", "subsystem", "logic", "Compute", "SOC_TOP/CPU_CLUSTER", "tier=T0; confidence=review"],
        ["B2", "P001", "B0", "GPU_TOP", "subsystem", "logic", "Graphics", "SOC_TOP/GPU_TOP", "tier=T0/T1; confidence=review"],
        ["B3", "P001", "B0", "NPU_TOP", "subsystem", "logic+memory", "AI", "SOC_TOP/NPU_TOP", "tier=T0/T1; confidence=draft"],
        ["B4", "P001", "B0", "ISP_TOP", "subsystem", "logic", "Camera", "SOC_TOP/ISP_TOP", "tier=T1; confidence=approved"],
        ["B5", "P001", "B0", "DDR_PHY", "phy", "phy_analog", "Memory IO", "SOC_TOP/DDR_PHY", "tier=T2; confidence=approved"],
        ["B6", "P001", "B0", "PCIE_USB_PHY", "phy", "phy_analog", "External IO", "SOC_TOP/PCIE_USB_PHY", "tier=T2; confidence=approved"],
        ["B7", "P001", "B1", "CPU_CORE_0", "block", "logic", "CPU", "SOC_TOP/CPU_CLUSTER/CPU_CORE_0", "tier=T0; confidence=review"],
        ["B8", "P001", "B3", "NPU_SRAM_BANKS", "macro_group", "memory", "AI Memory", "SOC_TOP/NPU_TOP/NPU_SRAM_BANKS", "tier=T1; confidence=draft"],
    ],
    "tiers": [
        ["id", "scenario_id", "tier_index", "name", "process_id", "role", "orientation", "thickness_um", "area_limit_mm2", "description"],
        ["T0", "S2", 0, "Top Tier", "PN5", "High-performance logic", "Face-down", 45, 28.2, "HB < 1um; utilization 72%; power 7.6 W"],
        ["T1", "S2", 1, "Middle Tier", "PN7", "Memory + medium logic", "Face-up / Face-to-face", 50, 31.4, "HB + TSV; utilization 66%; power 4.8 W"],
        ["T2", "S2", 2, "Bottom Tier", "PN7", "IO / PHY / PDN / logic", "Backside PDN", 60, 15.0, "TSV < 5um; utilization 58%; power 2.8 W"],
    ],
    "component_metrics": metric_rows(),
}


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 42)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def add_list_validation(ws, range_address: str, values: list[str]) -> None:
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(range_address)


def add_metric_validations(wb: Workbook) -> None:
    ws = wb["component_metrics"]
    max_row = max(ws.max_row + 40, 80)

    scenario_ids = [row[0] for row in SHEETS["scenarios"][1:]]
    component_ids = [row[0] for row in SHEETS["components"][1:]]
    metric_names = [row[0] for row in METRIC_DICTIONARY[1:]]
    metric_units = sorted({row[1] for row in METRIC_DICTIONARY[1:]})
    metric_categories = sorted({row[2] for row in METRIC_DICTIONARY[1:]})

    add_list_validation(ws, f"B2:B{max_row}", scenario_ids)
    add_list_validation(ws, f"C2:C{max_row}", component_ids)
    add_list_validation(ws, f"D2:D{max_row}", metric_names)
    add_list_validation(ws, f"F2:F{max_row}", metric_units)
    add_list_validation(ws, f"G2:G{max_row}", metric_categories)
    add_list_validation(ws, f"H2:H{max_row}", CORNER_VALUES)
    add_list_validation(ws, f"I2:I{max_row}", WORKLOAD_VALUES)
    add_list_validation(ws, f"J2:J{max_row}", CONFIDENCE_VALUES)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["E"].width = 14
    ws.auto_filter.ref = f"A1:K{ws.max_row}"


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in SHEETS.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
        style_sheet(ws)
        if len(rows) > 1 and len(rows[0]) > 1:
            table = Table(displayName=f"{name}_table", ref=f"A1:{ws.cell(row=1, column=len(rows[0])).column_letter}{len(rows)}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)
    add_metric_validations(wb)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
